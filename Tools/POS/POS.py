import sys
import sqlite3
import os
import shutil
import uuid
import hashlib
from datetime import datetime, timedelta
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewDialog
import random
import csv
from collections import defaultdict

# ==================== ថ្នាក់សម្រាប់គ្រប់គ្រងអ្នកប្រើប្រាស់ ====================
class User:
    def __init__(self, id, username, full_name, role, created_at=None):
        self.id = id
        self.username = username
        self.full_name = full_name
        self.role = role
        self.created_at = created_at
    
    def is_admin(self):
        return self.role == 'admin'
    
    def is_cashier(self):
        return self.role == 'cashier'

class CurrencyManager:
    def __init__(self):
        self.exchange_rate = 4000
        
    def usd_to_khr(self, usd_amount):
        return usd_amount * self.exchange_rate
    
    def khr_to_usd(self, khr_amount):
        return khr_amount / self.exchange_rate
    
    def format_currency(self, amount, currency='KHR', include_symbol=True):
        if currency == 'USD':
            if include_symbol:
                return f"${amount:,.2f}"
            else:
                return f"{amount:,.2f}"
        else:
            if include_symbol:
                return f"{amount:,.0f}៛"
            else:
                return f"{amount:,.0f}"
    
    def parse_currency(self, text):
        text = text.replace('$', '').replace('៛', '').replace(',', '').strip()
        try:
            return float(text)
        except ValueError:
            return 0.0

class BarcodeManager:
    @staticmethod
    def generate_barcode(length=6):
        barcode = str(random.randint(100000, 999999))
        return barcode
    
    @staticmethod
    def calculate_check_digit(barcode):
        return 0
    
    @staticmethod
    def generate_price_barcode(product_id, price):
        product_part = str(product_id).zfill(3)[-3:]
        price_part = str(int(price)).zfill(3)[-3:]
        return f"{product_part}{price_part}"
    
    @staticmethod
    def parse_price_barcode(barcode):
        if len(barcode) == 6:
            try:
                product_id = int(barcode[:3])
                price = int(barcode[3:]) * 1000
                return product_id, price
            except:
                return None, None
        return None, None

class BarcodePrintDialog(QDialog):
    def __init__(self, parent, product_data, quantity=1):
        super().__init__(parent)
        self.parent_window = parent
        self.product = product_data
        self.quantity = quantity
        self.setWindowTitle('បោះពុម្ព Barcode')
        self.setModal(True)
        self.setMinimumWidth(500)
        self.setMinimumHeight(400)
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout(self)
        
        title = QLabel("បោះពុម្ព Barcode សម្រាប់ទំនិញ")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4CAF50; padding: 10px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        info_group = QGroupBox("ព័ត៌មានទំនិញ")
        info_layout = QFormLayout()
        
        self.product_name = QLabel(self.product[1])
        self.product_name.setStyleSheet("font-weight: bold;")
        info_layout.addRow("ឈ្មោះទំនិញ:", self.product_name)
        
        self.product_barcode = QLabel(self.product[2] if self.product[2] else "មិនទាន់មាន")
        info_layout.addRow("Barcode:", self.product_barcode)
        
        self.product_price = QLabel(f"{self.product[3]:,.0f}៛")
        info_layout.addRow("តម្លៃ:", self.product_price)
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        settings_group = QGroupBox("ការកំណត់ការបោះពុម្ព")
        settings_layout = QFormLayout()
        
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 100)
        self.quantity_spin.setValue(self.quantity)
        self.quantity_spin.setSuffix(" ស្ទីគ័រ")
        self.quantity_spin.valueChanged.connect(self.update_preview)
        settings_layout.addRow("ចំនួន:", self.quantity_spin)
        
        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["A4", "A5"])
        self.page_size_combo.setCurrentText("A4")
        self.page_size_combo.currentTextChanged.connect(self.update_preview)
        settings_layout.addRow("ទំហំក្រដាស:", self.page_size_combo)
        
        self.barcode_type = QComboBox()
        self.barcode_type.addItems(["Barcode 6 ខ្ទង់", "Code 128"])
        self.barcode_type.currentTextChanged.connect(self.update_preview)
        settings_layout.addRow("ប្រភេទ Barcode:", self.barcode_type)
        
        self.include_price = QCheckBox("បង្ហាញតម្លៃនៅលើស្ទីគ័រ")
        self.include_price.setChecked(True)
        self.include_price.toggled.connect(self.update_preview)
        settings_layout.addRow("", self.include_price)
        
        self.label_size = QComboBox()
        self.label_size.addItems(["តូច (2.5cm x 1.5cm)", "មធ្យម (4cm x 2.5cm)", "ធំ (6cm x 4cm)"])
        self.label_size.currentTextChanged.connect(self.update_preview)
        settings_layout.addRow("ទំហំស្ទីគ័រ:", self.label_size)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        preview_group = QGroupBox("មើលជាមុន")
        preview_layout = QVBoxLayout()
        
        self.preview_widget = QWidget()
        self.preview_widget.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
        """)
        self.preview_widget.setMinimumHeight(200)
        
        self.preview_layout = QVBoxLayout(self.preview_widget)
        self.update_preview()
        
        preview_layout.addWidget(self.preview_widget)
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        button_layout = QHBoxLayout()
        
        self.print_btn = QPushButton("🖨️ បោះពុម្ព")
        self.print_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px 20px; font-weight: bold; font-size: 14px;")
        self.print_btn.clicked.connect(self.print_barcodes)
        
        self.print_preview_btn = QPushButton("📄 មើលមុនបោះពុម្ព")
        self.print_preview_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 10px 20px; font-weight: bold; font-size: 14px;")
        self.print_preview_btn.clicked.connect(self.show_print_preview)
        
        self.generate_new_btn = QPushButton("🔄 បង្កើត Barcode ថ្មី")
        self.generate_new_btn.setStyleSheet("background-color: #FF9800; color: white; padding: 10px 20px; font-weight: bold; font-size: 14px;")
        self.generate_new_btn.clicked.connect(self.generate_new_barcode)
        
        cancel_btn = QPushButton("បោះបង់")
        cancel_btn.setStyleSheet("background-color: #f44336; color: white; padding: 10px 20px; font-weight: bold; font-size: 14px;")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.print_btn)
        button_layout.addWidget(self.print_preview_btn)
        button_layout.addWidget(self.generate_new_btn)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def update_preview(self):
        for i in reversed(range(self.preview_layout.count())): 
            widget = self.preview_layout.itemAt(i)
            if widget and widget.widget():
                widget.widget().setParent(None)
        
        preview_widget = QWidget()
        preview_widget.setStyleSheet("background-color: white;")
        preview_layout = QVBoxLayout(preview_widget)
        preview_layout.setAlignment(Qt.AlignCenter)
        
        label_size = self.label_size.currentText()
        if "តូច" in label_size:
            label_width = 250
            label_height = 150
            barcode_font_size = 24
            price_font_size = 24
            name_font_size = 12
        elif "មធ្យម" in label_size:
            label_width = 400
            label_height = 250
            barcode_font_size = 32
            price_font_size = 32
            name_font_size = 14
        else:
            label_width = 600
            label_height = 400
            barcode_font_size = 48
            price_font_size = 48
            name_font_size = 16
        
        label_frame = QFrame()
        label_frame.setFixedSize(label_width, label_height)
        label_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 2px solid #4CAF50;
                border-radius: 5px;
            }
        """)
        
        frame_layout = QVBoxLayout(label_frame)
        frame_layout.setSpacing(5)
        frame_layout.setContentsMargins(10, 10, 10, 10)
        
        name_label = QLabel(self.product[1][:20] + "..." if len(self.product[1]) > 20 else self.product[1])
        name_label.setAlignment(Qt.AlignCenter)
        name_label.setStyleSheet(f"font-size: {name_font_size}px; font-weight: bold; color: #333;")
        name_label.setWordWrap(True)
        frame_layout.addWidget(name_label)
        
        barcode_text = self.product[2] if self.product[2] else "123456"
        barcode_label = QLabel(barcode_text)
        barcode_label.setAlignment(Qt.AlignCenter)
        barcode_label.setStyleSheet(f"font-family: 'Courier New', monospace; font-size: {barcode_font_size}px; font-weight: bold; letter-spacing: 5px;")
        frame_layout.addWidget(barcode_label)
        
        if self.include_price.isChecked():
            price_label = QLabel(f"{self.product[3]:,.0f}៛")
            price_label.setAlignment(Qt.AlignCenter)
            price_label.setStyleSheet(f"font-size: {price_font_size}px; font-weight: bold; color: #4CAF50;")
            frame_layout.addWidget(price_label)
        
        preview_layout.addWidget(label_frame, 0, Qt.AlignCenter)
        self.preview_layout.addWidget(preview_widget)
    
    def generate_new_barcode(self):
        new_barcode = BarcodeManager.generate_barcode(6)
        
        reply = QMessageBox.question(self, 'បញ្ជាក់', 
                                    f'តើចង់ប្តូរ Barcode ពី "{self.product[2]}" ទៅ "{new_barcode}" មែនទេ?',
                                    QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                self.parent_window.db.cursor.execute("""
                    UPDATE products SET barcode = ?, updated_at = ?
                    WHERE id = ?
                """, (new_barcode, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), self.product[0]))
                
                self.parent_window.db.conn.commit()
                
                self.product = (self.product[0], self.product[1], new_barcode, self.product[3], self.product[4])
                self.product_barcode.setText(new_barcode)
                self.update_preview()
                
                QMessageBox.information(self, 'ជោគជ័យ', 'បានបង្កើត Barcode ថ្មីរួចរាល់!')
                
                self.parent_window.load_products()
                if self.parent_window.current_user.is_admin():
                    self.parent_window.load_inventory()
                    
            except Exception as e:
                QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើត Barcode ថ្មី:\n{str(e)}')
    
    def print_barcodes(self):
        quantity = self.quantity_spin.value()
        include_price = self.include_price.isChecked()
        label_size = self.label_size.currentText()
        page_size = self.page_size_combo.currentText()
        
        printer = QPrinter(QPrinter.HighResolution)
        
        if page_size == "A4":
            printer.setPageSize(QPrinter.A4)
            page_margin = 5
        else:
            printer.setPageSize(QPrinter.A5)
            page_margin = 3
        
        printer.setPageMargins(page_margin, page_margin, page_margin, page_margin, QPrinter.Millimeter)
        
        print_dialog = QPrintDialog(printer, self)
        if print_dialog.exec_() == QDialog.Accepted:
            painter = QPainter()
            painter.begin(printer)
            
            if page_size == "A5":
                if "តូច" in label_size:
                    label_width_mm = 20
                    label_height_mm = 12
                    barcode_font_size = 6
                    price_font_size = 8
                    name_font_size = 4
                elif "មធ្យម" in label_size:
                    label_width_mm = 35
                    label_height_mm = 22
                    barcode_font_size = 10
                    price_font_size = 12
                    name_font_size = 5
                else:
                    label_width_mm = 50
                    label_height_mm = 35
                    barcode_font_size = 14
                    price_font_size = 16
                    name_font_size = 7
            else:
                if "តូច" in label_size:
                    label_width_mm = 25
                    label_height_mm = 15
                    barcode_font_size = 8
                    price_font_size = 10
                    name_font_size = 5
                elif "មធ្យម" in label_size:
                    label_width_mm = 40
                    label_height_mm = 25
                    barcode_font_size = 12
                    price_font_size = 14
                    name_font_size = 6
                else:
                    label_width_mm = 60
                    label_height_mm = 40
                    barcode_font_size = 16
                    price_font_size = 20
                    name_font_size = 8
            
            dpi = painter.device().logicalDpiX()
            label_width_px = int(label_width_mm * dpi / 25.4)
            label_height_px = int(label_height_mm * dpi / 25.4)
            
            page_width = painter.device().width()
            page_height = painter.device().height()
            
            spacing = 8 if page_size == "A5" else 15
            cols = max(1, page_width // (label_width_px + spacing))
            rows = max(1, page_height // (label_height_px + spacing))
            
            total_width = cols * (label_width_px + spacing) - spacing
            total_height = rows * (label_height_px + spacing) - spacing
            start_x = max(3, (page_width - total_width) // 2)
            start_y = max(3, (page_height - total_height) // 2)
            
            labels_printed = 0
            current_page = 1
            
            while labels_printed < quantity:
                if current_page > 1:
                    printer.newPage()
                
                for row in range(rows):
                    for col in range(cols):
                        if labels_printed >= quantity:
                            break
                        
                        x = start_x + col * (label_width_px + spacing)
                        y = start_y + row * (label_height_px + spacing)
                        
                        painter.setPen(QPen(Qt.black, 0.5))
                        painter.setBrush(QBrush(Qt.white))
                        painter.drawRect(x, y, label_width_px, label_height_px)
                        
                        painter.setPen(Qt.black)
                        name_font = QFont('Khmer OS', name_font_size)
                        painter.setFont(name_font)
                        name = self.product[1][:10] + "..." if len(self.product[1]) > 10 else self.product[1]
                        name_rect = QRect(x + 2, y + 2, label_width_px - 4, int(label_height_px * 0.2))
                        painter.drawText(name_rect, Qt.AlignCenter, name)
                        
                        painter.setPen(Qt.black)
                        barcode_font = QFont('Courier New', barcode_font_size)
                        barcode_font.setBold(True)
                        painter.setFont(barcode_font)
                        barcode_text = self.product[2] if self.product[2] else "000000"
                        barcode_rect = QRect(x + 2, y + int(label_height_px * 0.2), label_width_px - 4, int(label_height_px * 0.3))
                        painter.drawText(barcode_rect, Qt.AlignCenter, barcode_text)
                        
                        if include_price:
                            painter.setPen(QPen(QColor(76, 175, 80)))
                            price_font = QFont('Khmer OS', price_font_size)
                            price_font.setBold(True)
                            painter.setFont(price_font)
                            price_rect = QRect(x + 2, y + int(label_height_px * 0.55), label_width_px - 4, int(label_height_px * 0.25))
                            painter.drawText(price_rect, Qt.AlignCenter, f"{self.product[3]:,.0f}៛")
                        
                        labels_printed += 1
                    
                    if labels_printed >= quantity:
                        break
                
                current_page += 1
            
            painter.end()
            
            QMessageBox.information(self, 'ជោគជ័យ', 
                f'បានបោះពុម្ព {quantity} ស្ទីគ័ររួចរាល់!\n'
                f'ទំហំក្រដាស: {page_size}\n'
                f'ចំនួនស្ទីគ័រក្នុងមួយទំព័រ: {cols} x {rows} = {cols * rows}')
    
    def show_print_preview(self):
        printer = QPrinter(QPrinter.HighResolution)
        
        if self.page_size_combo.currentText() == "A4":
            printer.setPageSize(QPrinter.A4)
            page_margin = 5
        else:
            printer.setPageSize(QPrinter.A5)
            page_margin = 3
        
        printer.setPageMargins(page_margin, page_margin, page_margin, page_margin, QPrinter.Millimeter)
        
        preview = QPrintPreviewDialog(printer, self)
        preview.setWindowTitle(f"មើលមុនបោះពុម្ព Barcode - {self.page_size_combo.currentText()}")
        preview.paintRequested.connect(self.handle_paint_request)
        preview.exec_()
    
    def handle_paint_request(self, printer):
        painter = QPainter()
        painter.begin(printer)
        
        quantity = self.quantity_spin.value()
        include_price = self.include_price.isChecked()
        label_size = self.label_size.currentText()
        page_size = self.page_size_combo.currentText()
        
        if page_size == "A5":
            if "តូច" in label_size:
                label_width_mm = 20
                label_height_mm = 12
                barcode_font_size = 6
                price_font_size = 8
                name_font_size = 4
            elif "មធ្យម" in label_size:
                label_width_mm = 35
                label_height_mm = 22
                barcode_font_size = 10
                price_font_size = 12
                name_font_size = 5
            else:
                label_width_mm = 50
                label_height_mm = 35
                barcode_font_size = 14
                price_font_size = 16
                name_font_size = 7
        else:
            if "តូច" in label_size:
                label_width_mm = 25
                label_height_mm = 15
                barcode_font_size = 8
                price_font_size = 10
                name_font_size = 5
            elif "មធ្យម" in label_size:
                label_width_mm = 40
                label_height_mm = 25
                barcode_font_size = 12
                price_font_size = 14
                name_font_size = 6
            else:
                label_width_mm = 60
                label_height_mm = 40
                barcode_font_size = 16
                price_font_size = 20
                name_font_size = 8
        
        dpi = painter.device().logicalDpiX()
        label_width_px = int(label_width_mm * dpi / 25.4)
        label_height_px = int(label_height_mm * dpi / 25.4)
        
        page_width = painter.device().width()
        page_height = painter.device().height()
        
        spacing = 8 if page_size == "A5" else 15
        cols = max(1, page_width // (label_width_px + spacing))
        rows = max(1, page_height // (label_height_px + spacing))
        
        total_width = cols * (label_width_px + spacing) - spacing
        total_height = rows * (label_height_px + spacing) - spacing
        start_x = max(3, (page_width - total_width) // 2)
        start_y = max(3, (page_height - total_height) // 2)
        
        labels_printed = 0
        current_page = 1
        
        while labels_printed < quantity:
            if current_page > 1:
                printer.newPage()
            
            for row in range(rows):
                for col in range(cols):
                    if labels_printed >= quantity:
                        break
                    
                    x = start_x + col * (label_width_px + spacing)
                    y = start_y + row * (label_height_px + spacing)
                    
                    painter.setPen(QPen(Qt.black, 0.5))
                    painter.setBrush(QBrush(Qt.white))
                    painter.drawRect(x, y, label_width_px, label_height_px)
                    
                    painter.setPen(Qt.black)
                    name_font = QFont('Khmer OS', name_font_size)
                    painter.setFont(name_font)
                    name = self.product[1][:10] + "..." if len(self.product[1]) > 10 else self.product[1]
                    name_rect = QRect(x + 2, y + 2, label_width_px - 4, int(label_height_px * 0.2))
                    painter.drawText(name_rect, Qt.AlignCenter, name)
                    
                    painter.setPen(Qt.black)
                    barcode_font = QFont('Courier New', barcode_font_size)
                    barcode_font.setBold(True)
                    painter.setFont(barcode_font)
                    barcode_text = self.product[2] if self.product[2] else "000000"
                    barcode_rect = QRect(x + 2, y + int(label_height_px * 0.2), label_width_px - 4, int(label_height_px * 0.3))
                    painter.drawText(barcode_rect, Qt.AlignCenter, barcode_text)
                    
                    if include_price:
                        painter.setPen(QPen(QColor(76, 175, 80)))
                        price_font = QFont('Khmer OS', price_font_size)
                        price_font.setBold(True)
                        painter.setFont(price_font)
                        price_rect = QRect(x + 2, y + int(label_height_px * 0.55), label_width_px - 4, int(label_height_px * 0.25))
                        painter.drawText(price_rect, Qt.AlignCenter, f"{self.product[3]:,.0f}៛")
                    
                    labels_printed += 1
                
                if labels_printed >= quantity:
                    break
            
            current_page += 1
        
        painter.end()

class BulkBarcodeDialog(QDialog):
    def __init__(self, parent, db):
        super().__init__(parent)
        self.parent_window = parent
        self.db = db
        self.selected_products = []
        self.setWindowTitle('បោះពុម្ព Barcode ច្រើនមុខ')
        self.setModal(True)
        self.setMinimumWidth(800)
        self.setMinimumHeight(600)
        
        self.initUI()
        self.load_products()
    
    def initUI(self):
        layout = QVBoxLayout(self)
        
        title = QLabel("បោះពុម្ព Barcode ច្រើនមុខ")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #4CAF50; padding: 10px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        selection_group = QGroupBox("ជ្រើសរើសទំនិញ")
        selection_layout = QVBoxLayout()
        
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("ស្វែងរក:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("បញ្ចូលឈ្មោះ ឬកូដទំនិញ...")
        self.search_input.textChanged.connect(self.filter_products)
        search_layout.addWidget(self.search_input)
        selection_layout.addLayout(search_layout)
        
        self.product_table = QTableWidget()
        self.product_table.setColumnCount(6)
        self.product_table.setHorizontalHeaderLabels(['ជ្រើស', 'ID', 'ឈ្មោះទំនិញ', 'Barcode', 'តម្លៃ', 'បរិមាណ'])
        self.product_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.product_table.setAlternatingRowColors(True)
        
        self.product_table.setColumnWidth(0, 50)
        self.product_table.setColumnWidth(1, 50)
        self.product_table.setColumnWidth(2, 250)
        self.product_table.setColumnWidth(3, 100)
        self.product_table.setColumnWidth(4, 100)
        self.product_table.setColumnWidth(5, 80)
        
        selection_layout.addWidget(self.product_table)
        
        select_layout = QHBoxLayout()
        select_all_btn = QPushButton("ជ្រើសរើសទាំងអស់")
        select_all_btn.clicked.connect(self.select_all)
        select_layout.addWidget(select_all_btn)
        
        clear_all_btn = QPushButton("ឈូសចេញទាំងអស់")
        clear_all_btn.clicked.connect(self.clear_all)
        select_layout.addWidget(clear_all_btn)
        select_layout.addStretch()
        selection_layout.addLayout(select_layout)
        
        selection_group.setLayout(selection_layout)
        layout.addWidget(selection_group)
        
        settings_group = QGroupBox("ការកំណត់ការបោះពុម្ព")
        settings_layout = QFormLayout()
        
        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["A4", "A5"])
        self.page_size_combo.setCurrentText("A4")
        settings_layout.addRow("ទំហំក្រដាស:", self.page_size_combo)
        
        self.label_size = QComboBox()
        self.label_size.addItems(["តូច (2.5cm x 1.5cm)", "មធ្យម (4cm x 2.5cm)", "ធំ (6cm x 4cm)"])
        settings_layout.addRow("ទំហំស្ទីគ័រ:", self.label_size)
        
        self.include_price = QCheckBox("បង្ហាញតម្លៃនៅលើស្ទីគ័រ")
        self.include_price.setChecked(True)
        settings_layout.addRow("", self.include_price)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        button_layout = QHBoxLayout()
        
        self.print_btn = QPushButton("🖨️ បោះពុម្ព Barcode ដែលបានជ្រើស")
        self.print_btn.setStyleSheet("background-color: #4CAF50; padding: 10px; font-weight: bold;")
        self.print_btn.clicked.connect(self.print_selected)
        
        self.preview_btn = QPushButton("📄 មើលមុនបោះពុម្ព")
        self.preview_btn.setStyleSheet("background-color: #2196F3; padding: 10px; font-weight: bold;")
        self.preview_btn.clicked.connect(self.show_print_preview)
        
        cancel_btn = QPushButton("បោះបង់")
        cancel_btn.setStyleSheet("background-color: #f44336; padding: 10px; font-weight: bold;")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.print_btn)
        button_layout.addWidget(self.preview_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
    
    def load_products(self):
        try:
            self.db.cursor.execute("""
                SELECT id, name, barcode, selling_price_retail
                FROM products 
                WHERE active = 1
                ORDER BY name
            """)
            products = self.db.cursor.fetchall()
            
            self.product_table.setRowCount(len(products))
            for row, product in enumerate(products):
                check_item = QTableWidgetItem()
                check_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                check_item.setCheckState(Qt.Unchecked)
                self.product_table.setItem(row, 0, check_item)
                
                id_item = QTableWidgetItem(str(product[0]))
                id_item.setFlags(Qt.ItemIsEnabled)
                self.product_table.setItem(row, 1, id_item)
                
                name_item = QTableWidgetItem(product[1])
                name_item.setFlags(Qt.ItemIsEnabled)
                self.product_table.setItem(row, 2, name_item)
                
                barcode_item = QTableWidgetItem(product[2] if product[2] else "")
                barcode_item.setFlags(Qt.ItemIsEnabled)
                self.product_table.setItem(row, 3, barcode_item)
                
                price_item = QTableWidgetItem(f"{product[3]:,.0f}៛" if product[3] else "0៛")
                price_item.setFlags(Qt.ItemIsEnabled)
                price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.product_table.setItem(row, 4, price_item)
                
                spinbox = QSpinBox()
                spinbox.setMinimum(1)
                spinbox.setMaximum(999)
                spinbox.setValue(1)
                spinbox.setEnabled(False)
                self.product_table.setCellWidget(row, 5, spinbox)
                
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចផ្ទុកទិន្នន័យ: {str(e)}')
    
    def filter_products(self):
        search_text = self.search_input.text().lower()
        
        for row in range(self.product_table.rowCount()):
            name_item = self.product_table.item(row, 2)
            barcode_item = self.product_table.item(row, 3)
            
            match = False
            if name_item and search_text in name_item.text().lower():
                match = True
            if barcode_item and search_text in barcode_item.text().lower():
                match = True
            
            self.product_table.setRowHidden(row, not match)
    
    def select_all(self):
        for row in range(self.product_table.rowCount()):
            if not self.product_table.isRowHidden(row):
                check_item = self.product_table.item(row, 0)
                check_item.setCheckState(Qt.Checked)
                
                spinbox = self.product_table.cellWidget(row, 5)
                if spinbox:
                    spinbox.setEnabled(True)
    
    def clear_all(self):
        for row in range(self.product_table.rowCount()):
            check_item = self.product_table.item(row, 0)
            check_item.setCheckState(Qt.Unchecked)
            
            spinbox = self.product_table.cellWidget(row, 5)
            if spinbox:
                spinbox.setEnabled(False)
                spinbox.setValue(1)
    
    def print_selected(self):
        selected = []
        
        for row in range(self.product_table.rowCount()):
            check_item = self.product_table.item(row, 0)
            if check_item and check_item.checkState() == Qt.Checked:
                product_id = int(self.product_table.item(row, 1).text())
                product_name = self.product_table.item(row, 2).text()
                barcode = self.product_table.item(row, 3).text()
                price_text = self.product_table.item(row, 4).text().replace('៛', '').replace(',', '')
                price = int(price_text) if price_text else 0
                
                spinbox = self.product_table.cellWidget(row, 5)
                quantity = spinbox.value() if spinbox else 1
                
                selected.append((product_id, product_name, barcode, price, quantity))
        
        if not selected:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមជ្រើសរើសទំនិញយ៉ាងហោចណាស់មួយមុខ!')
            return
        
        self.bulk_print(selected)
    
    def show_print_preview(self):
        selected = []
        
        for row in range(self.product_table.rowCount()):
            check_item = self.product_table.item(row, 0)
            if check_item and check_item.checkState() == Qt.Checked:
                product_id = int(self.product_table.item(row, 1).text())
                product_name = self.product_table.item(row, 2).text()
                barcode = self.product_table.item(row, 3).text()
                price_text = self.product_table.item(row, 4).text().replace('៛', '').replace(',', '')
                price = int(price_text) if price_text else 0
                
                spinbox = self.product_table.cellWidget(row, 5)
                quantity = spinbox.value() if spinbox else 1
                
                selected.append((product_id, product_name, barcode, price, quantity))
        
        if not selected:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមជ្រើសរើសទំនិញយ៉ាងហោចណាស់មួយមុខ!')
            return
        
        printer = QPrinter(QPrinter.HighResolution)
        
        if self.page_size_combo.currentText() == "A4":
            printer.setPageSize(QPrinter.A4)
            page_margin = 5
        else:
            printer.setPageSize(QPrinter.A5)
            page_margin = 3
        
        printer.setPageMargins(page_margin, page_margin, page_margin, page_margin, QPrinter.Millimeter)
        
        preview = QPrintPreviewDialog(printer, self)
        preview.setWindowTitle(f"មើលមុនបោះពុម្ព Barcode - {self.page_size_combo.currentText()}")
        preview.paintRequested.connect(lambda p: self.handle_preview_paint(p, selected))
        preview.exec_()
    
    def handle_preview_paint(self, printer, selected_products):
        painter = QPainter()
        painter.begin(printer)
        
        label_size = self.label_size.currentText()
        include_price = self.include_price.isChecked()
        page_size = self.page_size_combo.currentText()
        
        if page_size == "A5":
            if "តូច" in label_size:
                label_width_mm = 20
                label_height_mm = 12
                barcode_font_size = 6
                price_font_size = 8
                name_font_size = 4
            elif "មធ្យម" in label_size:
                label_width_mm = 35
                label_height_mm = 22
                barcode_font_size = 10
                price_font_size = 12
                name_font_size = 5
            else:
                label_width_mm = 50
                label_height_mm = 35
                barcode_font_size = 14
                price_font_size = 16
                name_font_size = 7
        else:
            if "តូច" in label_size:
                label_width_mm = 25
                label_height_mm = 15
                barcode_font_size = 8
                price_font_size = 10
                name_font_size = 5
            elif "មធ្យម" in label_size:
                label_width_mm = 40
                label_height_mm = 25
                barcode_font_size = 12
                price_font_size = 14
                name_font_size = 6
            else:
                label_width_mm = 60
                label_height_mm = 40
                barcode_font_size = 16
                price_font_size = 20
                name_font_size = 8
        
        dpi = painter.device().logicalDpiX()
        label_width_px = int(label_width_mm * dpi / 25.4)
        label_height_px = int(label_height_mm * dpi / 25.4)
        
        page_width = painter.device().width()
        page_height = painter.device().height()
        
        spacing = 8 if page_size == "A5" else 15
        cols = max(1, page_width // (label_width_px + spacing))
        
        total_width = cols * (label_width_px + spacing) - spacing
        start_x = max(3, (page_width - total_width) // 2)
        
        labels_printed = 0
        current_row = 0
        current_col = 0
        page_margin = 3 if page_size == "A5" else 5
        
        for product_id, name, barcode, price, quantity in selected_products:
            for _ in range(quantity):
                x = start_x + current_col * (label_width_px + spacing)
                y = page_margin * dpi / 25.4 + current_row * (label_height_px + spacing)
                
                if y + label_height_px > page_height - (page_margin * dpi / 25.4):
                    painter.end()
                    printer.newPage()
                    painter.begin(printer)
                    current_row = 0
                    current_col = 0
                    x = start_x
                    y = page_margin * dpi / 25.4
                
                painter.setPen(QPen(Qt.black, 0.5))
                painter.setBrush(QBrush(Qt.white))
                painter.drawRect(x, y, label_width_px, label_height_px)
                
                painter.setPen(Qt.black)
                name_font = QFont('Khmer OS', name_font_size)
                painter.setFont(name_font)
                short_name = name[:10] + "..." if len(name) > 10 else name
                name_rect = QRect(x + 2, y + 2, label_width_px - 4, int(label_height_px * 0.2))
                painter.drawText(name_rect, Qt.AlignCenter, short_name)
                
                painter.setPen(Qt.black)
                barcode_font = QFont('Courier New', barcode_font_size)
                barcode_font.setBold(True)
                painter.setFont(barcode_font)
                barcode_rect = QRect(x + 2, y + int(label_height_px * 0.2), label_width_px - 4, int(label_height_px * 0.3))
                painter.drawText(barcode_rect, Qt.AlignCenter, barcode[:6] if barcode else "000000")
                
                if include_price:
                    painter.setPen(QPen(QColor(76, 175, 80)))
                    price_font = QFont('Khmer OS', price_font_size)
                    price_font.setBold(True)
                    painter.setFont(price_font)
                    price_rect = QRect(x + 2, y + int(label_height_px * 0.55), label_width_px - 4, int(label_height_px * 0.25))
                    painter.drawText(price_rect, Qt.AlignCenter, f"{price:,.0f}៛")
                
                labels_printed += 1
                current_col += 1
                if current_col >= cols:
                    current_col = 0
                    current_row += 1
        
        painter.end()
    
    def bulk_print(self, selected_products):
        printer = QPrinter(QPrinter.HighResolution)
        
        if self.page_size_combo.currentText() == "A4":
            printer.setPageSize(QPrinter.A4)
            page_margin = 5
        else:
            printer.setPageSize(QPrinter.A5)
            page_margin = 3
        
        printer.setPageMargins(page_margin, page_margin, page_margin, page_margin, QPrinter.Millimeter)
        
        print_dialog = QPrintDialog(printer, self)
        if print_dialog.exec_() == QDialog.Accepted:
            painter = QPainter()
            painter.begin(printer)
            
            label_size = self.label_size.currentText()
            include_price = self.include_price.isChecked()
            page_size = self.page_size_combo.currentText()
            
            if page_size == "A5":
                if "តូច" in label_size:
                    label_width_mm = 20
                    label_height_mm = 12
                    barcode_font_size = 6
                    price_font_size = 8
                    name_font_size = 4
                elif "មធ្យម" in label_size:
                    label_width_mm = 35
                    label_height_mm = 22
                    barcode_font_size = 10
                    price_font_size = 12
                    name_font_size = 5
                else:
                    label_width_mm = 50
                    label_height_mm = 35
                    barcode_font_size = 14
                    price_font_size = 16
                    name_font_size = 7
            else:
                if "តូច" in label_size:
                    label_width_mm = 25
                    label_height_mm = 15
                    barcode_font_size = 8
                    price_font_size = 10
                    name_font_size = 5
                elif "មធ្យម" in label_size:
                    label_width_mm = 40
                    label_height_mm = 25
                    barcode_font_size = 12
                    price_font_size = 14
                    name_font_size = 6
                else:
                    label_width_mm = 60
                    label_height_mm = 40
                    barcode_font_size = 16
                    price_font_size = 20
                    name_font_size = 8
            
            dpi = painter.device().logicalDpiX()
            label_width_px = int(label_width_mm * dpi / 25.4)
            label_height_px = int(label_height_mm * dpi / 25.4)
            
            page_width = painter.device().width()
            page_height = painter.device().height()
            
            spacing = 8 if page_size == "A5" else 15
            cols = max(1, page_width // (label_width_px + spacing))
            
            total_width = cols * (label_width_px + spacing) - spacing
            start_x = max(3, (page_width - total_width) // 2)
            
            labels_printed = 0
            current_row = 0
            current_col = 0
            
            for product_id, name, barcode, price, quantity in selected_products:
                for _ in range(quantity):
                    x = start_x + current_col * (label_width_px + spacing)
                    y = page_margin * dpi / 25.4 + current_row * (label_height_px + spacing)
                    
                    if y + label_height_px > page_height - (page_margin * dpi / 25.4):
                        painter.end()
                        printer.newPage()
                        painter.begin(printer)
                        current_row = 0
                        current_col = 0
                        x = start_x
                        y = page_margin * dpi / 25.4
                    
                    painter.setPen(QPen(Qt.black, 0.5))
                    painter.setBrush(QBrush(Qt.white))
                    painter.drawRect(x, y, label_width_px, label_height_px)
                    
                    painter.setPen(Qt.black)
                    name_font = QFont('Khmer OS', name_font_size)
                    painter.setFont(name_font)
                    short_name = name[:10] + "..." if len(name) > 10 else name
                    name_rect = QRect(x + 2, y + 2, label_width_px - 4, int(label_height_px * 0.2))
                    painter.drawText(name_rect, Qt.AlignCenter, short_name)
                    
                    painter.setPen(Qt.black)
                    barcode_font = QFont('Courier New', barcode_font_size)
                    barcode_font.setBold(True)
                    painter.setFont(barcode_font)
                    barcode_rect = QRect(x + 2, y + int(label_height_px * 0.2), label_width_px - 4, int(label_height_px * 0.3))
                    painter.drawText(barcode_rect, Qt.AlignCenter, barcode[:6] if barcode else "000000")
                    
                    if include_price:
                        painter.setPen(QPen(QColor(76, 175, 80)))
                        price_font = QFont('Khmer OS', price_font_size)
                        price_font.setBold(True)
                        painter.setFont(price_font)
                        price_rect = QRect(x + 2, y + int(label_height_px * 0.55), label_width_px - 4, int(label_height_px * 0.25))
                        painter.drawText(price_rect, Qt.AlignCenter, f"{price:,.0f}៛")
                    
                    labels_printed += 1
                    current_col += 1
                    if current_col >= cols:
                        current_col = 0
                        current_row += 1
            
            painter.end()
            
            QMessageBox.information(self, 'ជោគជ័យ', 
                f'បានបោះពុម្ព Barcode {labels_printed} ស្ទីគ័ររួចរាល់!\n'
                f'ទំហំក្រដាស: {self.page_size_combo.currentText()}')

# ==================== ថ្នាក់សម្រាប់គ្រប់គ្រងអ្នកប្រើប្រាស់ ====================
class RegisterDialog(QDialog):
    def __init__(self, db, current_user=None):
        super().__init__()
        self.db = db
        self.current_user = current_user
        self.setWindowTitle('ចុះឈ្មោះប្រើប្រាស់ប្រព័ន្ធ')
        self.setModal(True)
        self.setFixedSize(500, 600)
        
        font = QFont('Khmer OS', 12)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
                font-weight: 500;
                padding: 2px 0;
            }
            QLineEdit {
                padding: 12px 15px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
                margin-bottom: 5px;
                min-height: 20px;
            }
            QLineEdit:hover {
                border-color: #b0b0b0;
            }
            QLineEdit:focus {
                border-color: #4CAF50;
                background-color: #ffffff;
            }
            QLineEdit::placeholder {
                color: #a0a0a0;
                font-style: italic;
                font-size: 13px;
            }
            QPushButton {
                padding: 12px 20px;
                border: none;
                border-radius: 8px;
                font-size: 15px;
                font-weight: bold;
                min-height: 40px;
            }
            QPushButton#registerBtn {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton#registerBtn:hover {
                background-color: #45a049;
            }
            QPushButton#cancelBtn {
                background-color: #f44336;
                color: white;
            }
            QPushButton#cancelBtn:hover {
                background-color: #e53935;
            }
            QPushButton#loginLink {
                background-color: transparent;
                color: #2196F3;
                font-size: 14px;
                border: none;
                padding: 8px;
            }
            QPushButton#loginLink:hover {
                color: #1976D2;
                text-decoration: underline;
                background-color: rgba(33, 150, 243, 0.1);
            }
            QComboBox {
                padding: 12px 15px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
                min-width: 150px;
            }
            QComboBox:hover {
                border-color: #b0b0b0;
            }
            QComboBox:focus {
                border-color: #4CAF50;
            }
        """)
        
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(40, 30, 40, 30)
        
        title_widget = QWidget()
        title_layout = QHBoxLayout(title_widget)
        title_layout.setContentsMargins(0, 0, 0, 0)
        
        icon_label = QLabel("📝")
        icon_label.setStyleSheet("font-size: 30px; color: #4CAF50;")
        title_layout.addWidget(icon_label)
        
        title_label = QLabel("ចុះឈ្មោះប្រើប្រាស់")
        title_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #4CAF50; padding: 10px 0;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        layout.addWidget(title_widget)
        
        name_widget = QWidget()
        name_layout = QHBoxLayout(name_widget)
        name_layout.setContentsMargins(0, 0, 0, 0)
        name_layout.setSpacing(10)
        
        name_icon = QLabel("👤")
        name_icon.setStyleSheet("font-size: 20px; color: #666;")
        name_layout.addWidget(name_icon)
        
        name_container = QVBoxLayout()
        name_container.setSpacing(2)
        name_container.addWidget(QLabel("ឈ្មោះពេញ:"))
        self.fullname_input = QLineEdit()
        self.fullname_input.setPlaceholderText("ឧ. សុខ ដារ៉ា")
        name_container.addWidget(self.fullname_input)
        name_layout.addLayout(name_container)
        
        layout.addWidget(name_widget)
        
        username_widget = QWidget()
        username_layout = QHBoxLayout(username_widget)
        username_layout.setContentsMargins(0, 0, 0, 0)
        username_layout.setSpacing(10)
        
        username_icon = QLabel("🔑")
        username_icon.setStyleSheet("font-size: 20px; color: #666;")
        username_layout.addWidget(username_icon)
        
        username_container = QVBoxLayout()
        username_container.setSpacing(2)
        username_container.addWidget(QLabel("ឈ្មោះអ្នកប្រើប្រាស់:"))
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("បញ្ចូលឈ្មោះសម្រាប់ចូលប្រើ...")
        username_container.addWidget(self.username_input)
        username_layout.addLayout(username_container)
        
        layout.addWidget(username_widget)
        
        password_widget = QWidget()
        password_layout = QHBoxLayout(password_widget)
        password_layout.setContentsMargins(0, 0, 0, 0)
        password_layout.setSpacing(10)
        
        password_icon = QLabel("🔒")
        password_icon.setStyleSheet("font-size: 20px; color: #666;")
        password_layout.addWidget(password_icon)
        
        password_container = QVBoxLayout()
        password_container.setSpacing(2)
        password_container.addWidget(QLabel("ពាក្យសម្ងាត់:"))
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("យ៉ាងហោចណាស់ 6 តួអក្សរ...")
        self.password_input.setEchoMode(QLineEdit.Password)
        password_container.addWidget(self.password_input)
        password_layout.addLayout(password_container)
        
        layout.addWidget(password_widget)
        
        confirm_widget = QWidget()
        confirm_layout = QHBoxLayout(confirm_widget)
        confirm_layout.setContentsMargins(0, 0, 0, 0)
        confirm_layout.setSpacing(10)
        
        confirm_icon = QLabel("✓")
        confirm_icon.setStyleSheet("font-size: 20px; color: #666;")
        confirm_layout.addWidget(confirm_icon)
        
        confirm_container = QVBoxLayout()
        confirm_container.setSpacing(2)
        confirm_container.addWidget(QLabel("បញ្ជាក់ពាក្យសម្ងាត់:"))
        self.confirm_input = QLineEdit()
        self.confirm_input.setPlaceholderText("វាយពាក្យសម្ងាត់ម្តងទៀត...")
        self.confirm_input.setEchoMode(QLineEdit.Password)
        confirm_container.addWidget(self.confirm_input)
        confirm_layout.addLayout(confirm_container)
        
        layout.addWidget(confirm_widget)
        
        self.role_widget = QWidget()
        role_layout = QHBoxLayout(self.role_widget)
        role_layout.setContentsMargins(0, 0, 0, 0)
        role_layout.setSpacing(10)
        
        role_icon = QLabel("⚙️")
        role_icon.setStyleSheet("font-size: 20px; color: #666;")
        role_layout.addWidget(role_icon)
        
        role_container = QVBoxLayout()
        role_container.setSpacing(2)
        role_container.addWidget(QLabel("តួនាទី:"))
        self.role_combo = QComboBox()
        self.role_combo.addItem("អ្នកលក់", "cashier")
        self.role_combo.addItem("អ្នកគ្រប់គ្រង", "admin")
        role_container.addWidget(self.role_combo)
        role_layout.addLayout(role_container)
        
        if self.current_user and self.current_user.is_admin():
            layout.addWidget(self.role_widget)
        else:
            self.role_widget.hide()
            self.role = 'cashier'
        
        layout.addStretch()
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.register_btn = QPushButton("ចុះឈ្មោះ")
        self.register_btn.setObjectName("registerBtn")
        self.register_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.register_btn.clicked.connect(self.register)
        button_layout.addWidget(self.register_btn)
        
        self.cancel_btn = QPushButton("បោះបង់")
        self.cancel_btn.setObjectName("cancelBtn")
        self.cancel_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
        
        if not self.current_user:
            self.login_link = QPushButton("មានគណនីរួចហើយ? ចុចទីនេះដើម្បីចូលប្រើប្រាស់")
            self.login_link.setObjectName("loginLink")
            self.login_link.setCursor(QCursor(Qt.PointingHandCursor))
            self.login_link.clicked.connect(self.go_to_login)
            layout.addWidget(self.login_link)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #f44336; font-size: 14px; padding: 8px; background-color: #ffebee; border-radius: 4px;")
        self.status_label.hide()
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
    
    def register(self):
        fullname = self.fullname_input.text().strip()
        username = self.username_input.text().strip()
        password = self.password_input.text()
        confirm = self.confirm_input.text()
        
        if self.current_user and self.current_user.is_admin():
            role = self.role_combo.currentData()
        else:
            role = 'cashier'
        
        if not fullname or not username or not password:
            self.show_error("សូមបញ្ចូលព័ត៌មានឱ្យបានពេញលេញ!")
            return
        
        if password != confirm:
            self.show_error("ពាក្យសម្ងាត់មិនត្រូវគ្នា!")
            return
        
        if len(password) < 6:
            self.show_error("ពាក្យសម្ងាត់ត្រូវមានយ៉ាងហោចណាស់ 6 តួអក្សរ!")
            return
        
        try:
            self.db.cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
            if self.db.cursor.fetchone():
                self.show_error("ឈ្មោះអ្នកប្រើប្រាស់នេះមានរួចហើយ!")
                return
            
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.db.cursor.execute("""
                INSERT INTO users (username, password, full_name, role, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (username, hashed_password, fullname, role, 1, current_time, current_time))
            
            self.db.conn.commit()
            
            QMessageBox.information(self, "ជោគជ័យ", "បានចុះឈ្មោះរួចរាល់! សូមចូលប្រើប្រាស់ប្រព័ន្ធ។")
            self.accept()
            
        except Exception as e:
            self.show_error(f"មានបញ្ហា: {str(e)}")
    
    def go_to_login(self):
        self.done(2)
    
    def show_error(self, message):
        self.status_label.setText(message)
        self.status_label.show()
        QTimer.singleShot(3000, lambda: self.status_label.hide())

class LoginDialog(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.current_user = None
        self.setWindowTitle('ចូលប្រើប្រាស់ប្រព័ន្ធ')
        self.setModal(True)
        self.setFixedSize(450, 550)
        
        font = QFont('Khmer OS', 12)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f0f0;
            }
            QLabel {
                color: #333;
                font-size: 14px;
                padding: 2px 0;
            }
            QLineEdit {
                padding: 12px 15px;
                border: 2px solid #ddd;
                border-radius: 5px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
                min-height: 20px;
            }
            QLineEdit:focus {
                border-color: #4CAF50;
            }
            QLineEdit::placeholder {
                color: #999;
                font-style: italic;
                font-size: 13px;
            }
            QPushButton {
                padding: 12px;
                border: none;
                border-radius: 5px;
                font-size: 16px;
                font-weight: bold;
                min-height: 45px;
            }
            QPushButton#loginBtn {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton#loginBtn:hover {
                background-color: #45a049;
            }
            QPushButton#cancelBtn {
                background-color: #f44336;
                color: white;
            }
            QPushButton#cancelBtn:hover {
                background-color: #da190b;
            }
            QPushButton#registerLink {
                background-color: transparent;
                color: #2196F3;
                text-decoration: underline;
                font-size: 15px;
                border: none;
            }
            QPushButton#registerLink:hover {
                color: #1976D2;
                background-color: rgba(33, 150, 243, 0.1);
            }
        """)
        
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(40, 40, 40, 40)
        
        title_label = QLabel("🏪 ប្រព័ន្ធគ្រប់គ្រងការលក់")
        title_label.setStyleSheet("font-size: 22px; font-weight: bold; color: #4CAF50; qproperty-alignment: AlignCenter; padding: 20px;")
        layout.addWidget(title_label)
        
        layout.addWidget(QLabel("ឈ្មោះអ្នកប្រើប្រាស់:"))
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("បញ្ចូលឈ្មោះអ្នកប្រើប្រាស់...")
        layout.addWidget(self.username_input)
        
        layout.addWidget(QLabel("ពាក្យសម្ងាត់:"))
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("បញ្ចូលពាក្យសម្ងាត់...")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.returnPressed.connect(self.login)
        layout.addWidget(self.password_input)
        
        layout.addStretch()
        
        button_layout = QHBoxLayout()
        
        self.login_btn = QPushButton("ចូលប្រើប្រាស់")
        self.login_btn.setObjectName("loginBtn")
        self.login_btn.clicked.connect(self.login)
        button_layout.addWidget(self.login_btn)
        
        self.cancel_btn = QPushButton("បោះបង់")
        self.cancel_btn.setObjectName("cancelBtn")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.register_link = QPushButton("មិនទាន់មានគណនី? ចុចទីនេះដើម្បីចុះឈ្មោះ")
        self.register_link.setObjectName("registerLink")
        self.register_link.clicked.connect(self.go_to_register)
        layout.addWidget(self.register_link)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #f44336; font-size: 14px; qproperty-alignment: AlignCenter; padding: 5px;")
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
    
    def login(self):
        username = self.username_input.text().strip()
        password = self.password_input.text()
        
        if not username or not password:
            self.show_error("សូមបញ្ចូលឈ្មោះអ្នកប្រើប្រាស់ និងពាក្យសម្ងាត់!")
            return
        
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        
        self.db.cursor.execute("""
            SELECT id, username, full_name, role, created_at 
            FROM users 
            WHERE username = ? AND password = ? AND active = 1
        """, (username, hashed_password))
        
        user_data = self.db.cursor.fetchone()
        
        if user_data:
            self.current_user = User(user_data[0], user_data[1], user_data[2], user_data[3], user_data[4])
            self.accept()
        else:
            self.show_error("ឈ្មោះអ្នកប្រើប្រាស់ ឬពាក្យសម្ងាត់មិនត្រឹមត្រូវ!")
    
    def go_to_register(self):
        self.done(2)
    
    def show_error(self, message):
        self.status_label.setText(message)
        QTimer.singleShot(3000, lambda: self.status_label.clear())

class WelcomeDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('សូមស្វាគមន៍')
        self.setModal(True)
        self.setFixedSize(550, 450)
        
        font = QFont('Khmer OS', 12)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f0f0;
            }
            QLabel {
                color: #333;
                font-size: 16px;
            }
            QPushButton {
                padding: 15px 30px;
                border: none;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
                min-width: 220px;
                min-height: 50px;
            }
            QPushButton#loginBtn {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton#loginBtn:hover {
                background-color: #45a049;
            }
            QPushButton#registerBtn {
                background-color: #2196F3;
                color: white;
            }
            QPushButton#registerBtn:hover {
                background-color: #1976D2;
            }
        """)
        
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        layout.setSpacing(30)
        layout.setContentsMargins(50, 50, 50, 50)
        
        icon_label = QLabel("🏪")
        icon_label.setStyleSheet("font-size: 90px; qproperty-alignment: AlignCenter;")
        layout.addWidget(icon_label)
        
        welcome_label = QLabel("សូមស្វាគមន៍មកកាន់\nប្រព័ន្ធគ្រប់គ្រងការលក់សម្ភារៈសិក្សា")
        welcome_label.setStyleSheet("font-size: 22px; font-weight: bold; color: #4CAF50; qproperty-alignment: AlignCenter;")
        welcome_label.setWordWrap(True)
        layout.addWidget(welcome_label)
        
        layout.addStretch()
        
        self.login_btn = QPushButton("ចូលប្រើប្រាស់")
        self.login_btn.setObjectName("loginBtn")
        self.login_btn.clicked.connect(lambda: self.done(1))
        layout.addWidget(self.login_btn)
        
        self.register_btn = QPushButton("ចុះឈ្មោះថ្មី")
        self.register_btn.setObjectName("registerBtn")
        self.register_btn.clicked.connect(lambda: self.done(2))
        layout.addWidget(self.register_btn)
        
        self.setLayout(layout)

# ==================== ថ្នាក់ StudyMaterialsDB ====================
class StudyMaterialsDB:
    def __init__(self):
        self.db_path = 'study_materials.db'
        self.conn = None
        self.cursor = None
        self.currency_manager = CurrencyManager()
        self.connect_db()
        self.create_image_folder()
    
    def connect_db(self):
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        self.create_tables()
    
    def create_tables(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                full_name TEXT NOT NULL,
                role TEXT DEFAULT 'cashier',
                active BOOLEAN DEFAULT 1,
                created_at TEXT,
                updated_at TEXT
            )
        ''')
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT,
                barcode TEXT UNIQUE,
                cost_price REAL DEFAULT 0,
                selling_price_retail REAL DEFAULT 0,
                selling_price_wholesale REAL DEFAULT 0,
                wholesale_min_qty INTEGER DEFAULT 10,
                discount_percent REAL DEFAULT 0,
                discount_amount REAL DEFAULT 0,
                has_discount BOOLEAN DEFAULT 0,
                stock_retail INTEGER DEFAULT 0,
                stock_wholesale INTEGER DEFAULT 0,
                min_stock INTEGER DEFAULT 5,
                description TEXT,
                image_path TEXT,
                taxable BOOLEAN DEFAULT 1,
                discountable BOOLEAN DEFAULT 1,
                active BOOLEAN DEFAULT 1,
                created_at TEXT,
                updated_at TEXT
            )
        ''')
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                total_amount REAL,
                discount_amount REAL DEFAULT 0,
                final_amount REAL,
                payment_method TEXT,
                received_amount REAL DEFAULT 0,
                change_amount REAL DEFAULT 0,
                customer_name TEXT,
                customer_phone TEXT,
                user_id INTEGER,
                exchange_rate INTEGER DEFAULT 4000,
                sale_type TEXT DEFAULT 'retail',
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS sale_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER,
                product_id INTEGER,
                quantity INTEGER,
                original_price REAL,
                discount_percent REAL DEFAULT 0,
                discount_amount REAL DEFAULT 0,
                final_price REAL,
                subtotal REAL,
                sale_type TEXT DEFAULT 'retail',
                FOREIGN KEY (sale_id) REFERENCES sales (id),
                FOREIGN KEY (product_id) REFERENCES products (id)
            )
        ''')
        self.conn.commit()
        
        self.upgrade_all_tables()
        
        try:
            self.cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode)")
            self.conn.commit()
        except Exception as e:
            print(f"Note: Could not create unique index: {e}")
        
        self.create_default_admin()
        
        self.cursor.execute("SELECT COUNT(*) FROM products")
        if self.cursor.fetchone()[0] == 0:
            self.insert_sample_data()
    
    def create_default_admin(self):
        self.cursor.execute("SELECT COUNT(*) FROM users")
        if self.cursor.fetchone()[0] == 0:
            hashed_password = hashlib.sha256("admin123".encode()).hexdigest()
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            self.cursor.execute("""
                INSERT INTO users (username, password, full_name, role, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, ("admin", hashed_password, "អ្នកគ្រប់គ្រង", "admin", 1, current_time, current_time))
            
            hashed_password2 = hashlib.sha256("cashier123".encode()).hexdigest()
            self.cursor.execute("""
                INSERT INTO users (username, password, full_name, role, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, ("cashier", hashed_password2, "អ្នកលក់", "cashier", 1, current_time, current_time))
            
            self.conn.commit()
            print("Default users created: admin/admin123, cashier/cashier123")
    
    def upgrade_all_tables(self):
        tables_to_upgrade = {
            'products': [
                ('barcode', 'TEXT'),
                ('cost_price', 'REAL DEFAULT 0'),
                ('selling_price_retail', 'REAL DEFAULT 0'),
                ('selling_price_wholesale', 'REAL DEFAULT 0'),
                ('wholesale_min_qty', 'INTEGER DEFAULT 10'),
                ('discount_percent', 'REAL DEFAULT 0'),
                ('discount_amount', 'REAL DEFAULT 0'),
                ('has_discount', 'BOOLEAN DEFAULT 0'),
                ('stock_retail', 'INTEGER DEFAULT 0'),
                ('stock_wholesale', 'INTEGER DEFAULT 0'),
                ('min_stock', 'INTEGER DEFAULT 5'),
                ('description', 'TEXT'),
                ('image_path', 'TEXT'),
                ('taxable', 'BOOLEAN DEFAULT 1'),
                ('discountable', 'BOOLEAN DEFAULT 1'),
                ('active', 'BOOLEAN DEFAULT 1'),
                ('created_at', 'TEXT'),
                ('updated_at', 'TEXT')
            ],
            'sales': [
                ('discount_amount', 'REAL DEFAULT 0'),
                ('final_amount', 'REAL'),
                ('received_amount', 'REAL DEFAULT 0'),
                ('change_amount', 'REAL DEFAULT 0'),
                ('customer_name', 'TEXT'),
                ('customer_phone', 'TEXT'),
                ('user_id', 'INTEGER'),
                ('exchange_rate', 'INTEGER DEFAULT 4000'),
                ('sale_type', 'TEXT DEFAULT "retail"')
            ],
            'sale_items': [
                ('original_price', 'REAL'),
                ('discount_percent', 'REAL DEFAULT 0'),
                ('discount_amount', 'REAL DEFAULT 0'),
                ('final_price', 'REAL'),
                ('subtotal', 'REAL'),
                ('sale_type', 'TEXT DEFAULT "retail"')
            ]
        }
        
        for table_name, columns in tables_to_upgrade.items():
            self.upgrade_table(table_name, columns)
    
    def upgrade_table(self, table_name, columns):
        try:
            self.cursor.execute(f"PRAGMA table_info({table_name})")
            existing_columns = [column[1] for column in self.cursor.fetchall()]
            
            for col_name, col_def in columns:
                if col_name not in existing_columns:
                    try:
                        print(f"Adding {col_name} column to {table_name} table...")
                        self.cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_def}")
                        self.conn.commit()
                        print(f"✓ {col_name} column added successfully")
                    except Exception as e:
                        print(f"Error adding {col_name} column: {e}")
        except Exception as e:
            print(f"Error upgrading {table_name} table: {e}")
    
    def create_image_folder(self):
        if not os.path.exists('product_images'):
            os.makedirs('product_images')
    
    def insert_sample_data(self):
        sample_products = [
            ('សៀវភៅគណិតវិទ្យាថ្នាក់ទី១០', 'សៀវភៅ', '123456', 20000, 32000, 28000, 5, 10, 3200, 1, 50, 100, 10, 'សៀវភៅសិក្សាគណិតវិទ្យាថ្នាក់ទី១០', None, 1, 1, 1),
            ('ប៊ិកខ្មៅ', 'ប៊ិក', '234567', 800, 2000, 1500, 20, 5, 100, 1, 200, 500, 50, 'ប៊ិកខ្មៅសម្រាប់សរសេរ', None, 1, 1, 1),
            ('សៀវភៅកត់ត្រា 100ទំព័រ', 'សៀវភៅកត់ត្រា', '345678', 3200, 6000, 5000, 10, 0, 0, 0, 150, 300, 30, 'សៀវភៅកត់ត្រា 100ទំព័រ គម្របរឹង', None, 1, 1, 1),
            ('ខ្មៅដែរ HB', 'ខ្មៅដៃ', '456789', 400, 1000, 800, 50, 0, 0, 0, 300, 1000, 100, 'ខ្មៅដែរ HB សម្រាប់សរសេរ', None, 1, 1, 1),
            ('បន្ទាត់ 30cm', 'បន្ទាត់', '567890', 1200, 2400, 2000, 15, 0, 0, 0, 80, 200, 20, 'បន្ទាត់ផ្លាស្ទិច 30cm', None, 1, 1, 1),
        ]
        
        for product in sample_products:
            try:
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.cursor.execute('''
                    INSERT INTO products (
                        name, category, barcode, cost_price, selling_price_retail, selling_price_wholesale,
                        wholesale_min_qty, discount_percent, discount_amount, has_discount,
                        stock_retail, stock_wholesale, min_stock,
                        description, image_path, taxable, discountable, active,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', product + (current_time, current_time))
            except sqlite3.IntegrityError:
                pass
            except Exception as e:
                print(f"Error inserting sample data: {e}")
        
        self.conn.commit()

# ==================== ថ្នាក់ MainWindow ====================
class MainWindow(QMainWindow):
    def __init__(self, current_user):
        super().__init__()
        self.db = StudyMaterialsDB()
        self.currency_manager = CurrencyManager()
        self.current_user = current_user
        self.cart = []
        self.sale_type = 'retail'
        self.selected_product_image = None
        self.barcode_input_mode = 'manual'
        self.initUI()
        
        role_text = "អ្នកគ្រប់គ្រង" if current_user.is_admin() else "អ្នកលក់"
        self.setWindowTitle(f'ប្រព័ន្ធគ្រប់គ្រងការលក់សម្ភារៈសិក្សា - {current_user.full_name} ({role_text})')
    
    def initUI(self):
        self.setGeometry(100, 100, 1500, 800)
        
        font = QFont('Khmer OS', 10)
        self.setFont(font)
        
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTabWidget::pane {
                border: 1px solid #cccccc;
                background-color: white;
                border-radius: 5px;
            }
            QTabBar::tab {
                background-color: #e0e0e0;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
                font-size: 13px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 13px;
                font-weight: bold;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QTableWidget {
                gridline-color: #d0d0d0;
                selection-background-color: #e3f2fd;
                alternate-background-color: #f9f9f9;
                font-size: 12px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 12px;
            }
            QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox {
                padding: 8px;
                border: 1px solid #cccccc;
                border-radius: 4px;
                font-size: 12px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                min-height: 30px;
            }
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                border: 2px solid #4CAF50;
            }
            QLineEdit::placeholder {
                color: #999;
                font-style: italic;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #cccccc;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
                font-size: 13px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QMenuBar {
                background-color: #4CAF50;
                color: white;
                font-size: 13px;
            }
            QMenuBar::item {
                background-color: transparent;
                padding: 8px 16px;
            }
            QMenuBar::item:selected {
                background-color: #45a049;
            }
            QMenu {
                background-color: white;
                border: 1px solid #cccccc;
                font-size: 12px;
            }
            QMenu::item:selected {
                background-color: #e8f5e8;
            }
        """)
        
        menubar = self.menuBar()
        
        file_menu = menubar.addMenu('ឯកសារ')
        
        change_pass_action = QAction('ប្តូរពាក្យសម្ងាត់', self)
        change_pass_action.triggered.connect(self.change_password)
        file_menu.addAction(change_pass_action)
        
        file_menu.addSeparator()
        
        logout_action = QAction('ចាកចេញ', self)
        logout_action.triggered.connect(self.logout)
        file_menu.addAction(logout_action)
        
        if self.current_user.is_admin():
            user_menu = menubar.addMenu('អ្នកប្រើប្រាស់')
            manage_users_action = QAction('គ្រប់គ្រងអ្នកប្រើប្រាស់', self)
            manage_users_action.triggered.connect(self.manage_users)
            user_menu.addAction(manage_users_action)
            
            register_action = QAction('បង្កើតគណនីថ្មី', self)
            register_action.triggered.connect(self.register_new_user)
            user_menu.addAction(register_action)
            
            barcode_menu = menubar.addMenu('🏷️ Barcode')
            bulk_barcode_action = QAction('បោះពុម្ព Barcode ច្រើនមុខ', self)
            bulk_barcode_action.triggered.connect(self.bulk_print_barcodes)
            barcode_menu.addAction(bulk_barcode_action)
        
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        info_widget = QWidget()
        info_widget.setStyleSheet("background-color: #e8f5e8; border-radius: 5px; padding: 8px;")
        info_layout = QHBoxLayout(info_widget)
        info_layout.setContentsMargins(10, 5, 10, 5)
        
        user_info = QLabel(f"អ្នកប្រើប្រាស់: {self.current_user.full_name} ({'អ្នកគ្រប់គ្រង' if self.current_user.is_admin() else 'អ្នកលក់'})")
        user_info.setStyleSheet("font-weight: bold; color: #4CAF50; font-size: 13px;")
        info_layout.addWidget(user_info)
        
        info_layout.addStretch()
        
        rate_label = QLabel(f"អត្រាប្តូរប្រាក់: 1$ = {self.currency_manager.exchange_rate:,.0f}៛")
        rate_label.setStyleSheet("color: #2196F3; font-weight: bold; font-size: 13px;")
        info_layout.addWidget(rate_label)
        
        logout_btn = QPushButton("ចាកចេញ")
        logout_btn.setStyleSheet("background-color: #f44336; padding: 5px 15px; font-size: 12px;")
        logout_btn.clicked.connect(self.logout)
        info_layout.addWidget(logout_btn)
        
        layout.addWidget(info_widget)
        
        self.tabs = QTabWidget()
        self.tabs.addTab(self.create_sales_tab(), "🏪 លក់ទំនិញ")
        
        if self.current_user.is_admin():
            self.tabs.addTab(self.create_inventory_tab(), "📦 ស្តុកទំនិញ")
            self.tabs.addTab(self.create_reports_tab(), "📊 របាយការណ៍")
        
        layout.addWidget(self.tabs)
        
        self.statusBar().showMessage('រួចរាល់')
        self.statusBar().setStyleSheet("background-color: #4CAF50; color: white; padding: 5px; font-size: 12px;")
    
    def change_password(self):
        dialog = ChangePasswordDialog(self.db, self.current_user.id)
        dialog.exec_()
    
    def logout(self):
        reply = QMessageBox.question(self, 'បញ្ជាក់', 'តើអ្នកពិតជាចង់ចាកចេញមែនទេ?',
                                    QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.close()
            QTimer.singleShot(100, lambda: start_auth_flow())
    
    def manage_users(self):
        dialog = UserManagementDialog(self, self.db)
        dialog.exec_()
    
    def register_new_user(self):
        dialog = RegisterDialog(self.db, self.current_user)
        dialog.exec_()
    
    def bulk_print_barcodes(self):
        dialog = BulkBarcodeDialog(self, self.db)
        dialog.exec_()
    
    def create_sales_tab(self):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)
        
        left_panel = QWidget()
        left_panel.setMinimumWidth(650)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(8)
        
        sale_type_group = QGroupBox("ប្រភេទលក់")
        sale_type_group.setStyleSheet("""
            QGroupBox { 
                font-size: 11px; 
                font-weight: bold;
                color: #2196F3;
                border: 1px solid #2196F3;
                border-radius: 3px;
                margin-top: 5px;
                padding-top: 6px;
                max-height: 60px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        sale_type_layout = QHBoxLayout()
        sale_type_layout.setSpacing(15)
        sale_type_layout.setContentsMargins(5, 2, 5, 2)
        
        self.retail_radio = QRadioButton("លក់រាយ")
        self.retail_radio.setChecked(True)
        self.retail_radio.setStyleSheet("font-size: 11px;")
        self.retail_radio.toggled.connect(self.on_sale_type_changed)
        sale_type_layout.addWidget(self.retail_radio)
        
        self.wholesale_radio = QRadioButton("លក់ដុំ")
        self.wholesale_radio.setStyleSheet("font-size: 11px;")
        self.wholesale_radio.toggled.connect(self.on_sale_type_changed)
        sale_type_layout.addWidget(self.wholesale_radio)
        
        sale_type_layout.addStretch()
        sale_type_group.setLayout(sale_type_layout)
        left_layout.addWidget(sale_type_group)
        
        barcode_mode_group = QGroupBox("របៀបបញ្ចូល Barcode")
        barcode_mode_group.setStyleSheet("""
            QGroupBox { 
                font-size: 11px; 
                font-weight: bold;
                color: #FF9800;
                border: 1px solid #FF9800;
                border-radius: 3px;
                margin-top: 5px;
                padding-top: 6px;
                max-height: 60px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        barcode_mode_layout = QHBoxLayout()
        barcode_mode_layout.setSpacing(15)
        barcode_mode_layout.setContentsMargins(5, 2, 5, 2)
        
        self.manual_mode_radio = QRadioButton("បញ្ចូលដោយដៃ")
        self.manual_mode_radio.setChecked(True)
        self.manual_mode_radio.setStyleSheet("font-size: 11px;")
        self.manual_mode_radio.toggled.connect(self.on_barcode_mode_changed)
        barcode_mode_layout.addWidget(self.manual_mode_radio)
        
        self.scan_mode_radio = QRadioButton("ប្រើក្បាលស្កេន")
        self.scan_mode_radio.setStyleSheet("font-size: 11px;")
        self.scan_mode_radio.toggled.connect(self.on_barcode_mode_changed)
        barcode_mode_layout.addWidget(self.scan_mode_radio)
        
        barcode_mode_layout.addStretch()
        barcode_mode_group.setLayout(barcode_mode_layout)
        left_layout.addWidget(barcode_mode_group)
        
        search_group = QGroupBox("ស្វែងរកទំនិញ")
        search_group.setStyleSheet("""
            QGroupBox { 
                font-size: 11px; 
                font-weight: bold;
                color: #4CAF50;
                border: 1px solid #4CAF50;
                border-radius: 3px;
                margin-top: 5px;
                padding-top: 6px;
                max-height: 90px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        search_layout = QGridLayout()
        search_layout.setVerticalSpacing(3)
        search_layout.setHorizontalSpacing(8)
        search_layout.setContentsMargins(5, 2, 5, 2)
        
        search_layout.addWidget(QLabel("ឈ្មោះ/កូដ:"), 0, 0)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("បញ្ចូលឈ្មោះ ឬកូដទំនិញ...")
        self.search_input.setStyleSheet("font-size: 11px; padding: 4px;")
        self.search_input.textChanged.connect(self.search_products)
        search_layout.addWidget(self.search_input, 0, 1)
        
        search_layout.addWidget(QLabel("ប្រភេទ:"), 1, 0)
        self.category_filter = QComboBox()
        self.category_filter.setStyleSheet("font-size: 11px; padding: 4px;")
        self.category_filter.addItem("ទាំងអស់")
        self.category_filter.addItems(['សៀវភៅ', 'សៀវភៅកត់ត្រា', 'ប៊ិក', 'ខ្មៅដៃ', 'បន្ទាត់', 'កាបូប', 'ម៉ាស៊ីនគិតលេខ', 'ក្រដាស'])
        self.category_filter.currentTextChanged.connect(self.search_products)
        search_layout.addWidget(self.category_filter, 1, 1)
        
        search_btn = QPushButton("ស្វែងរក")
        search_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogContentsView))
        search_btn.setStyleSheet("font-size: 11px; padding: 4px; min-height: 20px;")
        search_btn.clicked.connect(self.search_products)
        search_layout.addWidget(search_btn, 0, 2, 2, 1)
        
        search_group.setLayout(search_layout)
        left_layout.addWidget(search_group)
        
        self.product_table = QTableWidget()
        self.product_table.setColumnCount(9)
        self.product_table.setHorizontalHeaderLabels(['ID', 'កូដ ៦ខ្ទង់', 'ឈ្មោះទំនិញ', 'ប្រភេទ', 'តម្លៃលក់', 'បញ្ចុះតម្លៃ', 'ស្តុករាយ', 'ស្តុកដុំ', 'រូបភាព'])
        self.product_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.product_table.setAlternatingRowColors(True)
        self.product_table.doubleClicked.connect(self.add_to_cart)
        
        self.product_table.setColumnWidth(0, 40)
        self.product_table.setColumnWidth(1, 80)
        self.product_table.setColumnWidth(2, 180)
        self.product_table.setColumnWidth(3, 90)
        self.product_table.setColumnWidth(4, 100)
        self.product_table.setColumnWidth(5, 80)
        self.product_table.setColumnWidth(6, 70)
        self.product_table.setColumnWidth(7, 70)
        self.product_table.setColumnWidth(8, 90)
        
        self.product_table.verticalHeader().setDefaultSectionSize(35)
        
        left_layout.addWidget(self.product_table)
        
        barcode_group = QGroupBox("បញ្ចូល Barcode ៦ខ្ទង់")
        barcode_group.setStyleSheet("QGroupBox { font-size: 11px; }")
        barcode_layout = QHBoxLayout()
        barcode_layout.setSpacing(5)
        barcode_layout.setContentsMargins(5, 2, 5, 2)
        
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("បញ្ចូល Barcode ៦ខ្ទង់ ដោយដៃ ឬប្រើក្បាលស្កេន...")
        self.barcode_input.setMaxLength(6)
        self.barcode_input.setValidator(QRegExpValidator(QRegExp("[0-9]{0,6}")))
        self.barcode_input.returnPressed.connect(self.add_by_barcode)
        self.barcode_input.setMinimumHeight(30)
        self.barcode_input.setStyleSheet("font-size: 12px; font-weight: bold;")
        barcode_layout.addWidget(self.barcode_input, 3)
        
        self.add_barcode_btn = QPushButton("➕ បន្ថែម")
        self.add_barcode_btn.setStyleSheet("background-color: #4CAF50; font-size: 11px; padding: 4px; min-height: 20px;")
        self.add_barcode_btn.clicked.connect(self.add_by_barcode)
        barcode_layout.addWidget(self.add_barcode_btn, 1)
        
        self.clear_barcode_btn = QPushButton("🗑️ សម្អាត")
        self.clear_barcode_btn.setStyleSheet("background-color: #f44336; font-size: 11px; padding: 4px; min-height: 20px;")
        self.clear_barcode_btn.clicked.connect(self.clear_barcode_input)
        barcode_layout.addWidget(self.clear_barcode_btn, 1)
        
        barcode_group.setLayout(barcode_layout)
        left_layout.addWidget(barcode_group)
        
        self.barcode_mode_label = QLabel("របៀបបច្ចុប្បន្ន: បញ្ចូលដោយដៃ (Barcode ៦ខ្ទង់)")
        self.barcode_mode_label.setStyleSheet("color: #2196F3; font-size: 11px; padding: 2px;")
        left_layout.addWidget(self.barcode_mode_label)
        
        right_panel = QWidget()
        right_panel.setMinimumWidth(500)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(8)
        
        image_group = QGroupBox("រូបភាពទំនិញដែលបានជ្រើស")
        image_group.setStyleSheet("""
            QGroupBox { 
                font-size: 12px; 
                font-weight: bold;
                color: #4CAF50;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                margin-top: 8px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        image_layout = QVBoxLayout()
        image_layout.setContentsMargins(10, 10, 10, 10)
        image_layout.setSpacing(5)
        
        self.selected_product_image_label = QLabel()
        self.selected_product_image_label.setAlignment(Qt.AlignCenter)
        self.selected_product_image_label.setMinimumHeight(200)
        self.selected_product_image_label.setMaximumHeight(220)
        self.selected_product_image_label.setMinimumWidth(200)
        self.selected_product_image_label.setMaximumWidth(220)
        self.selected_product_image_label.setStyleSheet("""
            QLabel {
                border: 2px solid #4CAF50;
                border-radius: 5px;
                background-color: #f9f9f9;
                padding: 5px;
                font-size: 13px;
                font-weight: bold;
                color: #666;
            }
        """)
        self.selected_product_image_label.setText("គ្មានរូបភាព\n\nចុចទ្វេដងលើទំនិញ\nដើម្បីមើលរូបភាព")
        
        image_layout.addWidget(self.selected_product_image_label, 0, Qt.AlignCenter)
        
        self.selected_product_name = QLabel("")
        self.selected_product_name.setAlignment(Qt.AlignCenter)
        self.selected_product_name.setStyleSheet("""
            QLabel {
                font-size: 12px;
                font-weight: bold;
                color: #2196F3;
                padding: 5px;
                background-color: #e3f2fd;
                border-radius: 3px;
            }
        """)
        image_layout.addWidget(self.selected_product_name)
        
        image_group.setLayout(image_layout)
        right_layout.addWidget(image_group)
        
        self.sale_type_label = QLabel("ប្រភេទលក់: លក់រាយ")
        self.sale_type_label.setStyleSheet("font-size: 11px; font-weight: bold; color: #2196F3; padding: 2px;")
        right_layout.addWidget(self.sale_type_label)
        
        cart_title = QLabel("🛒 កន្ត្រកទំនិញ")
        cart_title.setStyleSheet("font-size: 13px; font-weight: bold; color: #4CAF50; padding: 5px; background-color: #e8f5e8; border-radius: 3px;")
        right_layout.addWidget(cart_title)
        
        self.cart_table = QTableWidget()
        self.cart_table.setColumnCount(8)
        self.cart_table.setHorizontalHeaderLabels(['ID', 'ឈ្មោះ', 'ប្រភេទ', 'តម្លៃដើម', 'បញ្ចុះ', 'តម្លៃចុង', 'បរិមាណ', ''])
        
        self.cart_table.setColumnWidth(0, 30)
        self.cart_table.setColumnWidth(1, 100)
        self.cart_table.setColumnWidth(2, 40)
        self.cart_table.setColumnWidth(3, 65)
        self.cart_table.setColumnWidth(4, 55)
        self.cart_table.setColumnWidth(5, 65)
        self.cart_table.setColumnWidth(6, 60)
        self.cart_table.setColumnWidth(7, 20)
        
        self.cart_table.verticalHeader().setDefaultSectionSize(25)
        self.cart_table.setAlternatingRowColors(True)
        self.cart_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        right_layout.addWidget(self.cart_table)
        
        total_group = QGroupBox("សង្ខេបតម្លៃ")
        total_group.setStyleSheet("""
            QGroupBox { 
                font-size: 12px; 
                font-weight: bold;
                color: #4CAF50;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        total_layout = QGridLayout(total_group)
        total_layout.setSpacing(8)
        total_layout.setContentsMargins(10, 10, 10, 10)

        total_layout.addWidget(QLabel("ដើម:"), 0, 0)
        self.subtotal_label = QLabel("0៛")
        self.subtotal_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.subtotal_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                font-weight: bold;
                color: #333;
                background-color: #f5f5f5;
                padding: 5px 10px;
                border-radius: 3px;
            }
        """)
        total_layout.addWidget(self.subtotal_label, 0, 1)

        total_layout.addWidget(QLabel("បញ្ចុះតម្លៃ:"), 0, 2)
        self.total_discount_label = QLabel("0៛")
        self.total_discount_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.total_discount_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                font-weight: bold;
                color: #f44336;
                background-color: #ffebee;
                padding: 5px 10px;
                border-radius: 3px;
            }
        """)
        total_layout.addWidget(self.total_discount_label, 0, 3)

        total_layout.addWidget(QLabel("សរុបរៀល:"), 1, 0)
        self.total_khr_label = QLabel("0៛")
        self.total_khr_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.total_khr_label.setStyleSheet("""
            QLabel {
                font-size: 15px;
                font-weight: bold;
                color: #4CAF50;
                background-color: #e8f5e8;
                padding: 6px 12px;
                border-radius: 3px;
                border: 1px solid #4CAF50;
            }
        """)
        total_layout.addWidget(self.total_khr_label, 1, 1)

        total_layout.addWidget(QLabel("សរុបដុល្លា:"), 1, 2)
        self.total_usd_label = QLabel("$0.00")
        self.total_usd_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.total_usd_label.setStyleSheet("""
            QLabel {
                font-size: 15px;
                font-weight: bold;
                color: #2196F3;
                background-color: #e3f2fd;
                padding: 6px 12px;
                border-radius: 3px;
                border: 1px solid #2196F3;
            }
        """)
        total_layout.addWidget(self.total_usd_label, 1, 3)

        total_layout.setColumnStretch(0, 1)
        total_layout.setColumnStretch(1, 2)
        total_layout.setColumnStretch(2, 1)
        total_layout.setColumnStretch(3, 2)

        right_layout.addWidget(total_group)
        
        customer_group = QGroupBox("ព័ត៌មានអតិថិជន")
        customer_group.setStyleSheet("""
            QGroupBox { 
                font-size: 10px; 
                font-weight: bold;
                color: #FF9800;
                border: 1px solid #FF9800;
                border-radius: 3px;
                margin-top: 5px;
                padding-top: 6px;
                max-height: 85px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        customer_layout = QFormLayout()
        customer_layout.setSpacing(2)
        customer_layout.setContentsMargins(5, 2, 5, 2)
        
        self.customer_name = QLineEdit()
        self.customer_name.setPlaceholderText("ឈ្មោះអតិថិជន (បើមាន)...")
        self.customer_name.setStyleSheet("""
            QLineEdit {
                font-size: 10px;
                padding: 3px;
                border: 1px solid #ddd;
                border-radius: 3px;
                min-height: 18px;
                max-height: 22px;
            }
        """)
        customer_layout.addRow("ឈ្មោះ:", self.customer_name)
        
        self.customer_phone = QLineEdit()
        self.customer_phone.setPlaceholderText("លេខទូរស័ព្ទ (បើមាន)...")
        self.customer_phone.setStyleSheet("""
            QLineEdit {
                font-size: 10px;
                padding: 3px;
                border: 1px solid #ddd;
                border-radius: 3px;
                min-height: 18px;
                max-height: 22px;
            }
        """)
        customer_layout.addRow("ទូរស័ព្ទ:", self.customer_phone)
        
        customer_group.setLayout(customer_layout)
        right_layout.addWidget(customer_group)
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(5)
        
        self.remove_btn = QPushButton("❌ ដក")
        self.remove_btn.setStyleSheet("background-color: #f44336; padding: 3px; font-size: 10px; min-width: 40px; min-height: 20px;")
        self.remove_btn.clicked.connect(self.remove_from_cart)
        
        self.clear_btn = QPushButton("🔄 សម្អាត")
        self.clear_btn.setStyleSheet("background-color: #FF9800; padding: 3px; font-size: 10px; min-width: 40px; min-height: 20px;")
        self.clear_btn.clicked.connect(self.clear_cart)
        
        self.checkout_btn = QPushButton("💵 បញ្ចប់")
        self.checkout_btn.setStyleSheet("background-color: #2196F3; padding: 3px; font-size: 10px; min-width: 40px; min-height: 20px;")
        self.checkout_btn.clicked.connect(self.checkout)
        
        button_layout.addWidget(self.remove_btn)
        button_layout.addWidget(self.clear_btn)
        button_layout.addWidget(self.checkout_btn)
        
        right_layout.addLayout(button_layout)
        
        layout.addWidget(left_panel, 60)
        layout.addWidget(right_panel, 40)
        
        self.load_products()
        
        return widget
    
    def on_barcode_mode_changed(self):
        if self.manual_mode_radio.isChecked():
            self.barcode_input_mode = 'manual'
            self.barcode_mode_label.setText("របៀបបច្ចុប្បន្ន: បញ្ចូលដោយដៃ (Barcode ៦ខ្ទង់)")
            self.barcode_input.setPlaceholderText("បញ្ចូល Barcode ៦ខ្ទង់ ដោយដៃ...")
        else:
            self.barcode_input_mode = 'scan'
            self.barcode_mode_label.setText("របៀបបច្ចុប្បន្ន: ប្រើក្បាលស្កេន (Barcode ៦ខ្ទង់)")
            self.barcode_input.setPlaceholderText("សូមស្កេន Barcode ៦ខ្ទង់...")
        
        self.clear_barcode_input()
    
    def clear_barcode_input(self):
        self.barcode_input.clear()
        self.barcode_input.setFocus()
    
    def on_sale_type_changed(self):
        if self.retail_radio.isChecked():
            self.sale_type = 'retail'
            self.sale_type_label.setText("ប្រភេទលក់: លក់រាយ")
        else:
            self.sale_type = 'wholesale'
            self.sale_type_label.setText("ប្រភេទលក់: លក់ដុំ")
        
        self.load_products()
        self.clear_cart()
    
    def load_products(self):
        try:
            if self.sale_type == 'retail':
                self.db.cursor.execute("""
                    SELECT id, barcode, name, category, selling_price_retail, 
                           has_discount, discount_percent, discount_amount,
                           stock_retail, stock_wholesale, image_path 
                    FROM products 
                    WHERE stock_retail > 0 AND active = 1
                    ORDER BY name
                """)
            else:
                self.db.cursor.execute("""
                    SELECT id, barcode, name, category, selling_price_wholesale, 
                           has_discount, discount_percent, discount_amount,
                           stock_retail, stock_wholesale, image_path 
                    FROM products 
                    WHERE stock_wholesale > 0 AND active = 1
                    ORDER BY name
                """)
            
            products = self.db.cursor.fetchall()
            
            self.product_table.setRowCount(len(products))
            for row, product in enumerate(products):
                id_item = QTableWidgetItem(str(product[0]))
                self.product_table.setItem(row, 0, id_item)
                
                barcode = str(product[1] if product[1] else '')
                if len(barcode) > 6:
                    barcode = barcode[-6:]
                barcode_item = QTableWidgetItem(barcode)
                self.product_table.setItem(row, 1, barcode_item)
                
                name_item = QTableWidgetItem(str(product[2]))
                self.product_table.setItem(row, 2, name_item)
                
                cat_item = QTableWidgetItem(str(product[3] if product[3] else ''))
                self.product_table.setItem(row, 3, cat_item)
                
                price = product[4] or 0
                price_item = QTableWidgetItem(f"{price:,.0f}៛")
                price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.product_table.setItem(row, 4, price_item)
                
                has_discount = product[5]
                discount_percent = product[6] or 0
                discount_amount = product[7] or 0
                
                if has_discount:
                    if discount_percent > 0:
                        discount_text = f"{discount_percent:.0f}%"
                    else:
                        discount_text = f"{discount_amount:,.0f}៛"
                    discount_item = QTableWidgetItem(discount_text)
                    discount_item.setForeground(QBrush(QColor(244, 67, 54)))
                else:
                    discount_item = QTableWidgetItem("-")
                discount_item.setTextAlignment(Qt.AlignCenter)
                self.product_table.setItem(row, 5, discount_item)
                
                retail_stock = product[8] or 0
                stock_retail_item = QTableWidgetItem(str(retail_stock))
                stock_retail_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if retail_stock <= 5:
                    stock_retail_item.setForeground(QBrush(QColor(244, 67, 54)))
                self.product_table.setItem(row, 6, stock_retail_item)
                
                wholesale_stock = product[9] or 0
                stock_wholesale_item = QTableWidgetItem(str(wholesale_stock))
                stock_wholesale_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if wholesale_stock <= 5:
                    stock_wholesale_item.setForeground(QBrush(QColor(244, 67, 54)))
                self.product_table.setItem(row, 7, stock_wholesale_item)
                
                if len(product) > 10 and product[10] and os.path.exists(product[10]):
                    item = QTableWidgetItem("✅")
                    item.setForeground(QBrush(QColor(0, 128, 0)))
                else:
                    item = QTableWidgetItem("❌")
                    item.setForeground(QBrush(QColor(255, 0, 0)))
                self.product_table.setItem(row, 8, item)
                    
            self.statusBar().showMessage(f'បានផ្ទុកទំនិញ {len(products)} មុខ')
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចផ្ទុកទិន្នន័យ: {str(e)}')
    
    def search_products(self):
        search_term = self.search_input.text()
        category = self.category_filter.currentText()
        
        try:
            if self.sale_type == 'retail':
                query = """
                    SELECT id, barcode, name, category, selling_price_retail, 
                           has_discount, discount_percent, discount_amount,
                           stock_retail, stock_wholesale, image_path 
                    FROM products 
                    WHERE stock_retail > 0 AND active = 1
                """
            else:
                query = """
                    SELECT id, barcode, name, category, selling_price_wholesale, 
                           has_discount, discount_percent, discount_amount,
                           stock_retail, stock_wholesale, image_path 
                    FROM products 
                    WHERE stock_wholesale > 0 AND active = 1
                """
            
            params = []
            
            if search_term:
                query += " AND (name LIKE ? OR barcode LIKE ?)"
                params.extend([f'%{search_term}%', f'%{search_term}%'])
            
            if category != "ទាំងអស់":
                query += " AND category = ?"
                params.append(category)
            
            query += " ORDER BY name"
            
            self.db.cursor.execute(query, params)
            products = self.db.cursor.fetchall()
            
            self.product_table.setRowCount(len(products))
            for row, product in enumerate(products):
                id_item = QTableWidgetItem(str(product[0]))
                self.product_table.setItem(row, 0, id_item)
                
                barcode = str(product[1] if product[1] else '')
                if len(barcode) > 6:
                    barcode = barcode[-6:]
                barcode_item = QTableWidgetItem(barcode)
                self.product_table.setItem(row, 1, barcode_item)
                
                name_item = QTableWidgetItem(str(product[2]))
                self.product_table.setItem(row, 2, name_item)
                
                cat_item = QTableWidgetItem(str(product[3] if product[3] else ''))
                self.product_table.setItem(row, 3, cat_item)
                
                price = product[4] or 0
                price_item = QTableWidgetItem(f"{price:,.0f}៛")
                price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.product_table.setItem(row, 4, price_item)
                
                has_discount = product[5]
                discount_percent = product[6] or 0
                discount_amount = product[7] or 0
                
                if has_discount:
                    if discount_percent > 0:
                        discount_text = f"{discount_percent:.0f}%"
                    else:
                        discount_text = f"{discount_amount:,.0f}៛"
                    discount_item = QTableWidgetItem(discount_text)
                    discount_item.setForeground(QBrush(QColor(244, 67, 54)))
                else:
                    discount_item = QTableWidgetItem("-")
                discount_item.setTextAlignment(Qt.AlignCenter)
                self.product_table.setItem(row, 5, discount_item)
                
                retail_stock = product[8] or 0
                stock_retail_item = QTableWidgetItem(str(retail_stock))
                stock_retail_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if retail_stock <= 5:
                    stock_retail_item.setForeground(QBrush(QColor(244, 67, 54)))
                self.product_table.setItem(row, 6, stock_retail_item)
                
                wholesale_stock = product[9] or 0
                stock_wholesale_item = QTableWidgetItem(str(wholesale_stock))
                stock_wholesale_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if wholesale_stock <= 5:
                    stock_wholesale_item.setForeground(QBrush(QColor(244, 67, 54)))
                self.product_table.setItem(row, 7, stock_wholesale_item)
                
                if len(product) > 10 and product[10] and os.path.exists(product[10]):
                    item = QTableWidgetItem("✅")
                    item.setForeground(QBrush(QColor(0, 128, 0)))
                else:
                    item = QTableWidgetItem("❌")
                    item.setForeground(QBrush(QColor(255, 0, 0)))
                self.product_table.setItem(row, 8, item)
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចស្វែងរក: {str(e)}')
    
    def add_to_cart(self):
        current_row = self.product_table.currentRow()
        if current_row >= 0:
            try:
                product_id = int(self.product_table.item(current_row, 0).text())
                barcode = self.product_table.item(current_row, 1).text()
                name = self.product_table.item(current_row, 2).text()
                price_text = self.product_table.item(current_row, 4).text().replace('៛', '').replace(',', '')
                price = int(float(price_text)) if price_text else 0
                
                discount_text = self.product_table.item(current_row, 5).text()
                has_discount = discount_text != "-"
                
                if self.sale_type == 'retail':
                    stock = int(self.product_table.item(current_row, 6).text())
                else:
                    stock = int(self.product_table.item(current_row, 7).text())
                
                self.db.cursor.execute("""
                    SELECT discount_percent, discount_amount, has_discount, 
                           selling_price_retail, selling_price_wholesale, wholesale_min_qty
                    FROM products WHERE id = ?
                """, (product_id,))
                product_details = self.db.cursor.fetchone()
                
                discount_percent = product_details[0] or 0
                discount_amount = product_details[1] or 0
                has_discount_db = product_details[2]
                wholesale_min = product_details[5] or 1
                
                final_price = price
                if has_discount_db and self.sale_type == 'retail':
                    if discount_percent > 0:
                        final_price = price - (price * discount_percent / 100)
                    elif discount_amount > 0:
                        final_price = price - discount_amount
                
                for item in self.cart:
                    if item['id'] == product_id and item['sale_type'] == self.sale_type:
                        new_quantity = item['quantity'] + 1
                        
                        if self.sale_type == 'wholesale' and new_quantity < wholesale_min:
                            QMessageBox.warning(self, 'ព្រមាន', 
                                              f'{name}\nការលក់ដុំត្រូវការបរិមាណយ៉ាងតិច {wholesale_min} ដុំ!')
                            return
                        
                        if new_quantity <= stock:
                            item['quantity'] = new_quantity
                            item['subtotal'] = item['final_price'] * new_quantity
                            self.update_cart_display()
                            self.statusBar().showMessage(f'បានបន្ថែម {name} មួយទៀត')
                        else:
                            QMessageBox.warning(self, 'ព្រមាន', 
                                              f'{name}\nទំនិញមិនគ្រប់គ្រាន់ក្នុងស្តុក!\nស្តុកនៅសល់: {stock}')
                        return
                
                if self.sale_type == 'wholesale' and 1 < wholesale_min:
                    QMessageBox.warning(self, 'ព្រមាន', 
                                      f'{name}\nការលក់ដុំត្រូវការបរិមាណយ៉ាងតិច {wholesale_min} ដុំ!')
                    return
                
                self.cart.append({
                    'id': product_id,
                    'barcode': barcode,
                    'name': name,
                    'original_price': price,
                    'final_price': final_price,
                    'discount_percent': discount_percent if has_discount_db and self.sale_type == 'retail' else 0,
                    'discount_amount': discount_amount if has_discount_db and self.sale_type == 'retail' else 0,
                    'has_discount': has_discount_db and self.sale_type == 'retail',
                    'quantity': 1,
                    'stock': stock,
                    'sale_type': self.sale_type,
                    'subtotal': final_price
                })
                self.update_cart_display()
                self.statusBar().showMessage(f'បានបន្ថែម {name} ទៅកន្ត្រក')
                
            except Exception as e:
                QMessageBox.critical(self, 'កំហុស', f'មិនអាចបន្ថែមទំនិញ: {str(e)}')
    
    def add_by_barcode(self):
        barcode = self.barcode_input.text().strip()
        if not barcode:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមបញ្ចូល Barcode!')
            return
        
        if len(barcode) != 6:
            QMessageBox.warning(self, 'ព្រមាន', 'Barcode ត្រូវមាន ៦ខ្ទង់!')
            self.barcode_input.setFocus()
            self.barcode_input.selectAll()
            return
        
        try:
            if self.sale_type == 'retail':
                self.db.cursor.execute("""
                    SELECT id, barcode, name, category, selling_price_retail, 
                           discount_percent, discount_amount, has_discount,
                           stock_retail, stock_wholesale, image_path, wholesale_min_qty
                    FROM products 
                    WHERE barcode = ? AND stock_retail > 0 AND active = 1
                """, (barcode,))
            else:
                self.db.cursor.execute("""
                    SELECT id, barcode, name, category, selling_price_wholesale, 
                           discount_percent, discount_amount, has_discount,
                           stock_retail, stock_wholesale, image_path, wholesale_min_qty
                    FROM products 
                    WHERE barcode = ? AND stock_wholesale > 0 AND active = 1
                """, (barcode,))
            
            product = self.db.cursor.fetchone()
            
            if product:
                product_id = product[0]
                name = product[2]
                price = product[4] or 0
                
                if self.sale_type == 'retail':
                    stock = product[8] or 0
                else:
                    stock = product[9] or 0
                
                discount_percent = product[5] or 0
                discount_amount = product[6] or 0
                has_discount = product[7]
                wholesale_min = product[11] or 1
                
                final_price = price
                if has_discount and self.sale_type == 'retail':
                    if discount_percent > 0:
                        final_price = price - (price * discount_percent / 100)
                    elif discount_amount > 0:
                        final_price = price - discount_amount
                
                for item in self.cart:
                    if item['id'] == product_id and item['sale_type'] == self.sale_type:
                        new_quantity = item['quantity'] + 1
                        
                        if self.sale_type == 'wholesale' and new_quantity < wholesale_min:
                            QMessageBox.warning(self, 'ព្រមាន', 
                                              f'{name}\nការលក់ដុំត្រូវការបរិមាណយ៉ាងតិច {wholesale_min} ដុំ!')
                            return
                        
                        if new_quantity <= stock:
                            item['quantity'] = new_quantity
                            item['subtotal'] = item['final_price'] * new_quantity
                            self.update_cart_display()
                            self.barcode_input.clear()
                            self.statusBar().showMessage(f'បានបន្ថែម {name} មួយទៀត')
                            
                            if self.scan_mode_radio.isChecked():
                                self.barcode_input.setFocus()
                        else:
                            QMessageBox.warning(self, 'ព្រមាន', 
                                              f'{name}\nទំនិញមិនគ្រប់គ្រាន់ក្នុងស្តុក!')
                        return
                
                if self.sale_type == 'wholesale' and 1 < wholesale_min:
                    QMessageBox.warning(self, 'ព្រមាន', 
                                      f'{name}\nការលក់ដុំត្រូវការបរិមាណយ៉ាងតិច {wholesale_min} ដុំ!')
                    return
                
                self.cart.append({
                    'id': product_id,
                    'barcode': product[1],
                    'name': name,
                    'original_price': price,
                    'final_price': final_price,
                    'discount_percent': discount_percent if has_discount and self.sale_type == 'retail' else 0,
                    'discount_amount': discount_amount if has_discount and self.sale_type == 'retail' else 0,
                    'has_discount': has_discount and self.sale_type == 'retail',
                    'quantity': 1,
                    'stock': stock,
                    'sale_type': self.sale_type,
                    'subtotal': final_price
                })
                self.update_cart_display()
                self.barcode_input.clear()
                self.statusBar().showMessage(f'បានបន្ថែម {name} ទៅកន្ត្រក')
                
                if product[10] and os.path.exists(product[10]):
                    pixmap = QPixmap(product[10])
                    if not pixmap.isNull():
                        scaled_pixmap = pixmap.scaled(
                            190, 180, 
                            Qt.KeepAspectRatio, 
                            Qt.SmoothTransformation
                        )
                        self.selected_product_image_label.setPixmap(scaled_pixmap)
                        self.selected_product_image_label.setText("")
                
                if self.scan_mode_radio.isChecked():
                    self.barcode_input.setFocus()
            else:
                QMessageBox.warning(self, 'មិនមានទំនិញ', 
                                   f'រកមិនឃើញទំនិញដែលមានកូដ: {barcode}')
                self.barcode_input.clear()
                if self.scan_mode_radio.isChecked():
                    self.barcode_input.setFocus()
                
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចស្វែងរក Barcode: {str(e)}')
    
    def update_cart_display(self):
        self.cart_table.setRowCount(len(self.cart))
        total_original = 0
        total_discount = 0
        total_final = 0
        
        for row, item in enumerate(self.cart):
            id_item = QTableWidgetItem(str(item['id']))
            id_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            id_item.setFont(QFont('Khmer OS', 8))
            self.cart_table.setItem(row, 0, id_item)
            
            name_item = QTableWidgetItem(item['name'][:15] + "..." if len(item['name']) > 15 else item['name'])
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            name_item.setFont(QFont('Khmer OS', 8))
            self.cart_table.setItem(row, 1, name_item)
            
            type_item = QTableWidgetItem("រាយ" if item['sale_type'] == 'retail' else "ដុំ")
            type_item.setTextAlignment(Qt.AlignCenter)
            if item['sale_type'] == 'wholesale':
                type_item.setForeground(QBrush(QColor(255, 152, 0)))
            type_item.setFont(QFont('Khmer OS', 8))
            self.cart_table.setItem(row, 2, type_item)
            
            orig_price_item = QTableWidgetItem(f"{item['original_price']:,}៛")
            orig_price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if item['has_discount']:
                orig_price_item.setForeground(QBrush(QColor(128, 128, 128)))
                font = QFont('Khmer OS', 7)
                font.setStrikeOut(True)
                orig_price_item.setFont(font)
            else:
                orig_price_item.setFont(QFont('Khmer OS', 7))
            self.cart_table.setItem(row, 3, orig_price_item)
            
            if item['has_discount']:
                if item['discount_percent'] > 0:
                    discount_text = f"{item['discount_percent']:.0f}%"
                else:
                    discount_text = f"{item['discount_amount']:,}៛"
                discount_item = QTableWidgetItem(discount_text)
                discount_item.setForeground(QBrush(QColor(244, 67, 54)))
            else:
                discount_item = QTableWidgetItem("-")
            discount_item.setTextAlignment(Qt.AlignCenter)
            discount_item.setFont(QFont('Khmer OS', 7))
            self.cart_table.setItem(row, 4, discount_item)
            
            final_price_item = QTableWidgetItem(f"{item['final_price']:,}៛")
            final_price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if item['has_discount']:
                final_price_item.setForeground(QBrush(QColor(76, 175, 80)))
                font = QFont('Khmer OS', 8)
                font.setBold(True)
                final_price_item.setFont(font)
            else:
                final_price_item.setFont(QFont('Khmer OS', 7))
            self.cart_table.setItem(row, 5, final_price_item)
            
            spinbox = QSpinBox()
            spinbox.setMinimum(1)
            spinbox.setMaximum(item['stock'])
            spinbox.setValue(item['quantity'])
            spinbox.valueChanged.connect(lambda v, r=row: self.update_cart_quantity(r, v))
            spinbox.setStyleSheet("font-size: 8px; padding: 0px; min-height: 15px; max-height: 18px;")
            self.cart_table.setCellWidget(row, 6, spinbox)
            
            remove_btn = QPushButton("✕")
            remove_btn.setStyleSheet("background-color: #f44336; color: white; border: none; border-radius: 2px; font-weight: bold; font-size: 8px; min-width: 14px; max-width: 14px; min-height: 14px; max-height: 14px;")
            remove_btn.clicked.connect(lambda checked, r=row: self.remove_cart_item(r))
            self.cart_table.setCellWidget(row, 7, remove_btn)
            
            item_subtotal = item['final_price'] * item['quantity']
            total_final += item_subtotal
            total_original += item['original_price'] * item['quantity']
            total_discount += (item['original_price'] - item['final_price']) * item['quantity']
        
        self.total_khr = total_final
        self.total_usd = self.currency_manager.khr_to_usd(total_final)
        self.total_original = total_original
        self.total_discount = total_discount
        
        self.subtotal_label.setText(f"{total_original:,.0f}៛")
        self.total_discount_label.setText(f"{total_discount:,.0f}៛")
        self.total_khr_label.setText(f"{total_final:,.0f}៛")
        self.total_usd_label.setText(f"${self.total_usd:.2f}")
        
        for row in range(self.cart_table.rowCount()):
            self.cart_table.setRowHeight(row, 18)
    
    def update_cart_quantity(self, row, quantity):
        if row < len(self.cart):
            item = self.cart[row]
            
            if item['sale_type'] == 'wholesale':
                self.db.cursor.execute("SELECT wholesale_min_qty FROM products WHERE id = ?", (item['id'],))
                min_qty = self.db.cursor.fetchone()[0]
                if quantity < min_qty:
                    QMessageBox.warning(self, 'ព្រមាន', 
                                      f'{item["name"]}\nការលក់ដុំត្រូវការបរិមាណយ៉ាងតិច {min_qty} ដុំ!')
                    return
            
            self.cart[row]['quantity'] = quantity
            self.cart[row]['subtotal'] = self.cart[row]['final_price'] * quantity
            self.update_cart_display()
    
    def remove_cart_item(self, row):
        if row < len(self.cart):
            removed_item = self.cart.pop(row)
            self.update_cart_display()
            self.statusBar().showMessage(f'បានដក {removed_item["name"]} ចេញពីកន្ត្រក')
    
    def remove_from_cart(self):
        current_row = self.cart_table.currentRow()
        if current_row >= 0:
            self.remove_cart_item(current_row)
    
    def clear_cart(self):
        if self.cart:
            reply = QMessageBox.question(self, 'បញ្ជាក់', 
                                        'តើអ្នកពិតជាចង់សម្អាតកន្ត្រកទាំងមូលមែនទេ?',
                                        QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.cart.clear()
                self.update_cart_display()
                self.statusBar().showMessage('បានសម្អាតកន្ត្រក')
    
    def checkout(self):
        if not self.cart:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមបន្ថែមទំនិញទៅកន្ត្រកមុនពេលបញ្ចប់ការលក់!')
            return
        
        for item in self.cart:
            if item['sale_type'] == 'wholesale':
                self.db.cursor.execute("SELECT wholesale_min_qty FROM products WHERE id = ?", (item['id'],))
                min_qty = self.db.cursor.fetchone()[0]
                if item['quantity'] < min_qty:
                    QMessageBox.warning(self, 'ព្រមាន', 
                                      f'{item["name"]}\nការលក់ដុំត្រូវការបរិមាណយ៉ាងតិច {min_qty} ដុំ!')
                    return
        
        payment_dialog = QDialog(self)
        payment_dialog.setWindowTitle('បង់ប្រាក់')
        payment_dialog.setModal(True)
        payment_dialog.setMinimumWidth(500)
        
        font = QFont('Khmer OS', 11)
        payment_dialog.setFont(font)
        
        layout = QVBoxLayout(payment_dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        summary_group = QGroupBox("សង្ខេបការលក់")
        summary_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 12px; }")
        summary_layout = QFormLayout()
        summary_layout.setSpacing(8)
        
        summary_layout.addRow("ប្រភេទលក់:", QLabel("លក់រាយ" if self.sale_type == 'retail' else "លក់ដុំ"))
        summary_layout.addRow("តម្លៃសរុបដើម:", QLabel(f"{self.total_original:,.0f}៛"))
        summary_layout.addRow("បញ្ចុះតម្លៃសរុប:", QLabel(f"{self.total_discount:,.0f}៛"))
        summary_layout.addRow("តម្លៃសរុបចុងក្រោយ (រៀល):", QLabel(f"{self.total_khr:,.0f}៛"))
        summary_layout.addRow("តម្លៃសរុបចុងក្រោយ (ដុល្លារ):", QLabel(f"${self.total_usd:.2f}"))
        
        self.payment_method = QComboBox()
        self.payment_method.addItems(['សាច់ប្រាក់', 'ឥណទាន', 'ABA Pay', 'Wing', 'True Money'])
        summary_layout.addRow("វិធីបង់ប្រាក់:", self.payment_method)
        
        summary_group.setLayout(summary_layout)
        layout.addWidget(summary_group)
        
        payment_group = QGroupBox("ទូទាត់ប្រាក់")
        payment_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 12px; }")
        payment_layout = QFormLayout()
        payment_layout.setSpacing(8)
        
        self.amount_received = QLineEdit()
        self.amount_received.setPlaceholderText("បញ្ចូលទឹកប្រាក់ជារៀល...")
        self.amount_received.textChanged.connect(self.calculate_change)
        payment_layout.addRow("ទឹកប្រាក់ទទួល:", self.amount_received)
        
        self.change_label = QLabel("0៛")
        self.change_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #4CAF50;")
        payment_layout.addRow("ប្រាក់អាប់:", self.change_label)
        
        payment_group.setLayout(payment_layout)
        layout.addWidget(payment_group)
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        confirm_btn = QPushButton("បញ្ជាក់ការលក់")
        confirm_btn.setStyleSheet("background-color: #4CAF50; color: white;")
        confirm_btn.clicked.connect(lambda: self.process_payment(payment_dialog))
        
        cancel_btn = QPushButton("បោះបង់")
        cancel_btn.setStyleSheet("background-color: #f44336; color: white;")
        cancel_btn.clicked.connect(payment_dialog.reject)
        
        button_layout.addWidget(confirm_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        payment_dialog.exec_()
    
    def calculate_change(self):
        try:
            received_text = self.amount_received.text().replace(',', '')
            if received_text:
                received = int(float(received_text))
                change = received - self.total_khr
                self.change_label.setText(f"{change:,}៛")
            else:
                self.change_label.setText("0៛")
        except ValueError:
            self.change_label.setText("មិនត្រឹមត្រូវ")
    
    def process_payment(self, dialog):
        try:
            received_text = self.amount_received.text().replace(',', '')
            if not received_text:
                QMessageBox.warning(dialog, 'ព្រមាន', 'សូមបញ្ចូលទឹកប្រាក់!')
                return
            
            received = int(float(received_text))
            
            if received < self.total_khr:
                QMessageBox.warning(dialog, 'ព្រមាន', 'ទឹកប្រាក់មិនគ្រប់!')
                return
            
            change = received - self.total_khr
            
            date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            method = self.payment_method.currentText()
            customer_name = self.customer_name.text() or None
            customer_phone = self.customer_phone.text() or None
            
            self.db.cursor.execute("""
                INSERT INTO sales (date, total_amount, discount_amount, final_amount,
                                  payment_method, received_amount, change_amount, 
                                  customer_name, customer_phone, user_id, exchange_rate, sale_type)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (date, self.total_original, self.total_discount, self.total_khr,
                  method, received, change, customer_name, customer_phone, 
                  self.current_user.id, self.currency_manager.exchange_rate, self.sale_type))
            
            sale_id = self.db.cursor.lastrowid
            
            for item in self.cart:
                self.db.cursor.execute("""
                    INSERT INTO sale_items (sale_id, product_id, quantity, 
                                          original_price, discount_percent, discount_amount,
                                          final_price, subtotal, sale_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (sale_id, item['id'], item['quantity'], 
                      item['original_price'], item['discount_percent'], item['discount_amount'],
                      item['final_price'], item['subtotal'], item['sale_type']))
                
                if item['sale_type'] == 'retail':
                    self.db.cursor.execute("""
                        UPDATE products SET stock_retail = stock_retail - ?, updated_at = ?
                        WHERE id = ?
                    """, (item['quantity'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), item['id']))
                else:
                    self.db.cursor.execute("""
                        UPDATE products SET stock_wholesale = stock_wholesale - ?, updated_at = ?
                        WHERE id = ?
                    """, (item['quantity'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), item['id']))
            
            self.db.conn.commit()
            
            self.show_receipt(sale_id, date, method, received, change)
            
            self.cart.clear()
            self.update_cart_display()
            self.customer_name.clear()
            self.customer_phone.clear()
            self.load_products()
            
            if self.current_user.is_admin():
                self.load_inventory()
            
            dialog.accept()
            self.statusBar().showMessage('ការលក់បានបញ្ចប់ដោយជោគជ័យ')
            
        except ValueError:
            QMessageBox.warning(dialog, 'កំហុស', 'សូមបញ្ចូលលេខឱ្យបានត្រឹមត្រូវ!')
        except Exception as e:
            QMessageBox.critical(dialog, 'កំហុស', f'មានបញ្ហាក្នុងការរក្សាទុក:\n{str(e)}')
            import traceback
            traceback.print_exc()
    
    def show_receipt(self, sale_id, date, method, received, change):
        receipt_dialog = QDialog(self)
        receipt_dialog.setWindowTitle('វិក្កយបត្រ')
        receipt_dialog.setMinimumWidth(600)
        receipt_dialog.setMinimumHeight(600)
        
        font = QFont('Khmer OS', 10)
        receipt_dialog.setFont(font)
        
        layout = QVBoxLayout(receipt_dialog)
        layout.setContentsMargins(10, 10, 10, 10)
        
        receipt_text = QTextEdit()
        receipt_text.setReadOnly(True)
        receipt_text.setStyleSheet("font-family: 'Khmer OS', 'Courier New', monospace; font-size: 11px; background-color: white; padding: 10px; border: 1px solid #ddd; border-radius: 5px;")
        
        total_usd = self.currency_manager.khr_to_usd(self.total_khr)
        received_usd = self.currency_manager.khr_to_usd(received)
        change_usd = self.currency_manager.khr_to_usd(change)
        
        receipt = f"""
╔══════════════════════════════════════════════════════╗
║              ហាងលក់សម្ភារៈសិក្សា                      ║
╠══════════════════════════════════════════════════════╣
║ លេខវិក្កយបត្រ: {sale_id:06d}
║ កាលបរិច្ឆេទ: {date}
║ ប្រភេទលក់: {'លក់រាយ' if self.sale_type == 'retail' else 'លក់ដុំ'}
║ វិធីបង់ប្រាក់: {method}
║ អត្រាប្តូរប្រាក់: 1$ = {self.currency_manager.exchange_rate:,.0f}៛
"""
        
        if self.customer_name.text():
            receipt += f"║ អតិថិជន: {self.customer_name.text()}\n"
        if self.customer_phone.text():
            receipt += f"║ ទូរស័ព្ទ: {self.customer_phone.text()}\n"
        
        receipt += "╠══════════════════════════════════════════════════════╣\n"
        receipt += "║ បញ្ជីទំនិញ:\n"
        
        for item in self.cart:
            sale_type_text = "រាយ" if item['sale_type'] == 'retail' else "ដុំ"
            if item['has_discount']:
                receipt += f"""║ {item['name'][:15]:15} ({sale_type_text})
║   {item['quantity']:3} x {item['original_price']:6,}៛ = {item['original_price'] * item['quantity']:8,}៛
║   បញ្ចុះ: {item['discount_percent']:.0f}% ({item['discount_amount']:,.0f}៛)
║   ចុងក្រោយ: {item['subtotal']:8,}៛
"""
            else:
                receipt += f"""║ {item['name'][:15]:15} ({sale_type_text})
║   {item['quantity']:3} x {item['final_price']:6,}៛ = {item['subtotal']:8,}៛
"""
        
        receipt += "╠══════════════════════════════════════════════════════╣\n"
        receipt += f"║ តម្លៃសរុបដើម: {self.total_original:19,}៛\n"
        receipt += f"║ បញ្ចុះតម្លៃសរុប: {self.total_discount:17,}៛\n"
        receipt += f"║ តម្លៃសរុបចុងក្រោយ: {self.total_khr:15,}៛\n"
        receipt += f"║ តម្លៃសរុបចុងក្រោយ: ${total_usd:14.2f}\n"
        receipt += "╠══════════════════════════════════════════════════════╣\n"
        receipt += f"║ ទឹកប្រាក់ទទួល: {received:17,}៛\n"
        receipt += f"║ ប្រាក់អាប់: {change:19,}៛\n"
        receipt += f"║ ទឹកប្រាក់ទទួល: ${received_usd:15.2f}\n"
        receipt += f"║ ប្រាក់អាប់: ${change_usd:17.2f}\n"
        receipt += "╚══════════════════════════════════════════════════════╝\n"
        receipt += "          សូមអរគុណចំពោះការទិញរបស់លោកអ្នក!"
        
        receipt_text.setText(receipt)
        layout.addWidget(receipt_text)
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        print_btn = QPushButton("🖨️ បោះពុម្ព")
        print_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 5px 15px; font-weight: bold; font-size: 11px;")
        print_btn.clicked.connect(lambda: self.print_receipt(receipt))
        
        preview_btn = QPushButton("📄 មើលមុនបោះពុម្ព")
        preview_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 5px 15px; font-weight: bold; font-size: 11px;")
        preview_btn.clicked.connect(lambda: self.preview_receipt(receipt))
        
        close_btn = QPushButton("បិទ")
        close_btn.setStyleSheet("background-color: #f44336; color: white; padding: 5px 15px; font-weight: bold; font-size: 11px;")
        close_btn.clicked.connect(receipt_dialog.accept)
        
        button_layout.addWidget(print_btn)
        button_layout.addWidget(preview_btn)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        receipt_dialog.exec_()
    
    def preview_receipt(self, receipt):
        printer = QPrinter(QPrinter.HighResolution)
        
        preview_dialog = QPrintPreviewDialog(printer, self)
        preview_dialog.setWindowTitle("មើលមុនបោះពុម្ពវិក្កយបត្រ")
        
        page_size_combo = QComboBox()
        page_size_combo.addItems(["A4", "A5"])
        page_size_combo.setCurrentText("A5")
        
        toolbar = preview_dialog.findChild(QToolBar)
        if toolbar:
            toolbar.addSeparator()
            toolbar.addWidget(QLabel("  ទំហំក្រដាស: "))
            toolbar.addWidget(page_size_combo)
        
        def handle_paint(printer):
            if page_size_combo.currentText() == "A4":
                printer.setPageSize(QPrinter.A4)
                page_margin = 5
            else:
                printer.setPageSize(QPrinter.A5)
                page_margin = 3
            
            printer.setPageMargins(page_margin, page_margin, page_margin, page_margin, QPrinter.Millimeter)
            
            painter = QPainter()
            painter.begin(printer)
            
            font_size = 9 if page_size_combo.currentText() == "A5" else 11
            font = QFont('Khmer OS', font_size)
            painter.setFont(font)
            
            page_rect = printer.pageRect()
            start_x = 10
            line_height = font_size + 5
            margin = 10
            
            lines = receipt.split('\n')
            y_pos = margin + 20
            
            for line in lines:
                if y_pos > page_rect.height() - margin:
                    printer.newPage()
                    y_pos = margin + 20
                
                painter.drawText(start_x, y_pos, line)
                y_pos += line_height
            
            painter.end()
        
        preview_dialog.paintRequested.connect(handle_paint)
        preview_dialog.exec_()
    
    def print_receipt(self, receipt):
        dialog = QDialog(self)
        dialog.setWindowTitle("ជ្រើសរើសទំហំក្រដាស")
        dialog.setModal(True)
        dialog.setFixedSize(300, 150)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        
        layout.addWidget(QLabel("សូមជ្រើសរើសទំហំក្រដាសសម្រាប់បោះពុម្ពវិក្កយបត្រ:"))
        
        page_size_combo = QComboBox()
        page_size_combo.addItems(["A5 (148 x 210 mm)", "A4 (210 x 297 mm)"])
        page_size_combo.setCurrentIndex(0)
        layout.addWidget(page_size_combo)
        
        button_layout = QHBoxLayout()
        
        ok_btn = QPushButton("យល់ព្រម")
        ok_btn.clicked.connect(dialog.accept)
        
        cancel_btn = QPushButton("បោះបង់")
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        selected_size = page_size_combo.currentText()
        
        printer = QPrinter(QPrinter.HighResolution)
        
        if "A5" in selected_size:
            printer.setPageSize(QPrinter.A5)
            page_margin = 3
        else:
            printer.setPageSize(QPrinter.A4)
            page_margin = 5
        
        printer.setPageMargins(page_margin, page_margin, page_margin, page_margin, QPrinter.Millimeter)
        
        print_dialog = QPrintDialog(printer, self)
        print_dialog.setWindowTitle(f"បោះពុម្ពវិក្កយបត្រ - {selected_size}")
        
        if print_dialog.exec_() == QDialog.Accepted:
            painter = QPainter()
            painter.begin(printer)
            
            font_size = 9 if "A5" in selected_size else 11
            font = QFont('Khmer OS', font_size)
            painter.setFont(font)
            
            page_rect = printer.pageRect()
            start_x = 10
            line_height = font_size + 5
            margin = 10
            
            lines = receipt.split('\n')
            y_pos = margin + 20
            
            for line in lines:
                if y_pos > page_rect.height() - margin:
                    printer.newPage()
                    y_pos = margin + 20
                
                painter.drawText(start_x, y_pos, line)
                y_pos += line_height
            
            painter.end()
            
            QMessageBox.information(self, 'ជោគជ័យ', 
                f'បានបញ្ជូនវិក្កយបត្រទៅម៉ាស៊ីនព្រីនរួចរាល់!\n\n' +
                f'ទំហំក្រដាស: {selected_size}')
    
    def create_inventory_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)
        
        add_btn = QPushButton("➕ បន្ថែមទំនិញថ្មី")
        add_btn.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        add_btn.clicked.connect(self.add_product_dialog)
        
        edit_btn = QPushButton("✏️ កែប្រែទំនិញ")
        edit_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        edit_btn.setStyleSheet("background-color: #FF9800;")
        edit_btn.clicked.connect(self.edit_product_dialog)
        
        delete_btn = QPushButton("🗑️ លុបទំនិញ")
        delete_btn.setIcon(self.style().standardIcon(QStyle.SP_TrashIcon))
        delete_btn.setStyleSheet("background-color: #f44336;")
        delete_btn.clicked.connect(self.delete_product)
        
        print_barcode_btn = QPushButton("🏷️ បោះពុម្ព Barcode")
        print_barcode_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        print_barcode_btn.setStyleSheet("background-color: #FF9800;")
        print_barcode_btn.clicked.connect(self.print_product_barcode)
        
        bulk_barcode_btn = QPushButton("📑 បោះពុម្ព Barcode ច្រើនមុខ")
        bulk_barcode_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        bulk_barcode_btn.setStyleSheet("background-color: #9C27B0;")
        bulk_barcode_btn.clicked.connect(self.bulk_print_barcodes)
        
        refresh_btn = QPushButton("🔄 ធ្វើឱ្យថ្មី")
        refresh_btn.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        refresh_btn.clicked.connect(self.load_inventory)
        
        export_btn = QPushButton("📥 នាំចេញ")
        export_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        export_btn.clicked.connect(self.export_inventory)
        
        toolbar.addWidget(add_btn)
        toolbar.addWidget(edit_btn)
        toolbar.addWidget(delete_btn)
        toolbar.addWidget(print_barcode_btn)
        toolbar.addWidget(bulk_barcode_btn)
        toolbar.addWidget(refresh_btn)
        toolbar.addWidget(export_btn)
        toolbar.addStretch()
        
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("ស្វែងរក:"))
        self.inventory_search = QLineEdit()
        self.inventory_search.setPlaceholderText("បញ្ចូលឈ្មោះ ឬកូដទំនិញ...")
        self.inventory_search.textChanged.connect(self.filter_inventory)
        search_layout.addWidget(self.inventory_search)
        
        layout.addLayout(toolbar)
        layout.addLayout(search_layout)
        
        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(17)
        self.inventory_table.setHorizontalHeaderLabels([
            'ID', 'កូដ ៦ខ្ទង់', 'ឈ្មោះ', 'ប្រភេទ', 'តម្លៃដើម', 
            'តម្លៃរាយ', 'តម្លៃដុំ', 'បរិមាណដុំតិច',
            'បញ្ចុះតម្លៃ%', 'បញ្ចុះតម្លៃ៛', 'មានបញ្ចុះតម្លៃ',
            'ស្តុករាយ', 'ស្តុកដុំ', 'ស្តុកអប្បបរមា', 
            'ស្ថានភាព', 'រូបភាព', 'កែប្រែចុងក្រោយ'
        ])
        self.inventory_table.setAlternatingRowColors(True)
        self.inventory_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.inventory_table.doubleClicked.connect(self.edit_product_dialog)
        
        self.inventory_table.setColumnWidth(0, 40)
        self.inventory_table.setColumnWidth(1, 80)
        self.inventory_table.setColumnWidth(2, 150)
        self.inventory_table.setColumnWidth(3, 80)
        self.inventory_table.setColumnWidth(4, 80)
        self.inventory_table.setColumnWidth(5, 80)
        self.inventory_table.setColumnWidth(6, 80)
        self.inventory_table.setColumnWidth(7, 80)
        self.inventory_table.setColumnWidth(8, 70)
        self.inventory_table.setColumnWidth(9, 70)
        self.inventory_table.setColumnWidth(10, 80)
        self.inventory_table.setColumnWidth(11, 70)
        self.inventory_table.setColumnWidth(12, 70)
        self.inventory_table.setColumnWidth(13, 80)
        self.inventory_table.setColumnWidth(14, 70)
        self.inventory_table.setColumnWidth(15, 70)
        self.inventory_table.setColumnWidth(16, 110)
        
        self.inventory_table.verticalHeader().setDefaultSectionSize(35)
        
        layout.addWidget(self.inventory_table)
        
        summary_layout = QHBoxLayout()
        summary_layout.setSpacing(15)
        summary_layout.addWidget(QLabel("សង្ខេបស្តុក:"))
        
        self.total_products_label = QLabel("ទំនិញសរុប: 0")
        self.total_products_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        summary_layout.addWidget(self.total_products_label)
        
        self.low_stock_label = QLabel("ស្តុកជិតអស់: 0")
        self.low_stock_label.setStyleSheet("color: #f44336; font-weight: bold; font-size: 12px;")
        summary_layout.addWidget(self.low_stock_label)
        
        self.out_of_stock_label = QLabel("អស់ស្តុក: 0")
        self.out_of_stock_label.setStyleSheet("color: #ff9800; font-weight: bold; font-size: 12px;")
        summary_layout.addWidget(self.out_of_stock_label)
        
        summary_layout.addStretch()
        layout.addLayout(summary_layout)
        
        self.load_inventory()
        
        return widget
    
    def load_inventory(self):
        if not self.current_user.is_admin():
            return
            
        try:
            self.db.cursor.execute("""
                SELECT id, barcode, name, category, cost_price, 
                       selling_price_retail, selling_price_wholesale, wholesale_min_qty,
                       discount_percent, discount_amount, has_discount,
                       stock_retail, stock_wholesale, min_stock,
                       CASE WHEN active = 1 THEN 'សកម្ម' ELSE 'មិនសកម្ម' END,
                       image_path, updated_at
                FROM products 
                ORDER BY id
            """)
            products = self.db.cursor.fetchall()
            
            self.inventory_table.setRowCount(len(products))
            total_products = 0
            low_stock = 0
            out_of_stock = 0
            
            for row, product in enumerate(products):
                for col, value in enumerate(product):
                    if col == 1:
                        barcode = str(value if value else '')
                        if len(barcode) > 6:
                            barcode = barcode[-6:]
                        item = QTableWidgetItem(barcode)
                    elif col in [4, 5, 6]:
                        try:
                            val = int(value) if value else 0
                            item = QTableWidgetItem(f"{val:,}៛")
                            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        except:
                            item = QTableWidgetItem("0៛")
                    elif col == 8:
                        item = QTableWidgetItem(f"{value:.0f}%" if value else "0%")
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    elif col == 9:
                        item = QTableWidgetItem(f"{value:,.0f}៛" if value else "0៛")
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    elif col == 10:
                        item = QTableWidgetItem("បាទ/ចាស" if value else "ទេ")
                        if value:
                            item.setForeground(QBrush(QColor(76, 175, 80)))
                    elif col in [11, 12, 13]:
                        item = QTableWidgetItem(str(value if value else '0'))
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        
                        if col == 11 or col == 12:
                            try:
                                stock = int(value) if value else 0
                                min_stock_val = int(product[13]) if product[13] else 5
                                
                                if stock <= 0:
                                    item.setBackground(QColor(255, 235, 238))
                                    if col == 11:
                                        out_of_stock += 1
                                elif stock <= min_stock_val:
                                    item.setBackground(QColor(255, 243, 224))
                                    if col == 11:
                                        low_stock += 1
                            except:
                                pass
                    elif col == 15:
                        if value and os.path.exists(str(value)):
                            item = QTableWidgetItem("✅")
                            item.setForeground(QBrush(QColor(0, 128, 0)))
                        else:
                            item = QTableWidgetItem("❌")
                            item.setForeground(QBrush(QColor(255, 0, 0)))
                    else:
                        item = QTableWidgetItem(str(value if value else ''))
                    
                    if col < self.inventory_table.columnCount():
                        self.inventory_table.setItem(row, col, item)
                
                total_products += 1
            
            self.total_products_label.setText(f"ទំនិញសរុប: {total_products}")
            self.low_stock_label.setText(f"ស្តុកជិតអស់: {low_stock}")
            self.out_of_stock_label.setText(f"អស់ស្តុក: {out_of_stock}")
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចផ្ទុកស្តុក: {str(e)}')
    
    def filter_inventory(self):
        if not self.current_user.is_admin():
            return
            
        search_term = self.inventory_search.text().lower()
        
        for row in range(self.inventory_table.rowCount()):
            match = False
            for col in [1, 2]:
                item = self.inventory_table.item(row, col)
                if item and search_term in item.text().lower():
                    match = True
                    break
            self.inventory_table.setRowHidden(row, not match)
    
    def add_product_dialog(self):
        if not self.current_user.is_admin():
            QMessageBox.warning(self, 'ព្រមាន', 'អ្នកមិនមានសិទ្ធិប្រើប្រាស់មុខងារនេះទេ!')
            return
        
        dialog = AddProductDialog(self)
        dialog.exec_()
    
    def edit_product_dialog(self):
        if not self.current_user.is_admin():
            QMessageBox.warning(self, 'ព្រមាន', 'អ្នកមិនមានសិទ្ធិប្រើប្រាស់មុខងារនេះទេ!')
            return
            
        current_row = self.inventory_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមជ្រើសរើសទំនិញដែលចង់កែប្រែ!')
            return
        
        try:
            product_id_item = self.inventory_table.item(current_row, 0)
            if not product_id_item:
                QMessageBox.warning(self, 'ព្រមាន', 'មិនអាចយក ID ទំនិញបានទេ!')
                return
                
            product_id = int(product_id_item.text())
            
            self.db.cursor.execute("SELECT * FROM products WHERE id = ?", (product_id,))
            product = self.db.cursor.fetchone()
            
            if product:
                dialog = AddProductDialog(self, product_data=product)
                
                if dialog.exec_() == QDialog.Accepted:
                    self.load_inventory()
                    self.load_products()
                    self.statusBar().showMessage(f'កែប្រែទំនិញ "{product[1]}" រួចរាល់')
            else:
                QMessageBox.warning(self, 'ព្រមាន', 'រកមិនឃើញទំនិញក្នុងប្រព័ន្ធ!')
                
        except ValueError as e:
            QMessageBox.critical(self, 'កំហុស', f'ទិន្នន័យ ID មិនត្រឹមត្រូវ: {str(e)}')
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មានបញ្ហាក្នុងការកែប្រែទំនិញ:\n{str(e)}')
            import traceback
            traceback.print_exc()
    
    def print_product_barcode(self):
        if not self.current_user.is_admin():
            QMessageBox.warning(self, 'ព្រមាន', 'អ្នកមិនមានសិទ្ធិប្រើប្រាស់មុខងារនេះទេ!')
            return
            
        current_row = self.inventory_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមជ្រើសរើសទំនិញដែលចង់បោះពុម្ព Barcode!')
            return
        
        product_id = int(self.inventory_table.item(current_row, 0).text())
        product_name = self.inventory_table.item(current_row, 2).text()
        barcode = self.inventory_table.item(current_row, 1).text()
        price_text = self.inventory_table.item(current_row, 5).text().replace('៛', '').replace(',', '')
        price = int(price_text) if price_text else 0
        
        image_item = self.inventory_table.item(current_row, 15)
        image_path = None
        if image_item and "✅" in image_item.text():
            self.db.cursor.execute("SELECT image_path FROM products WHERE id = ?", (product_id,))
            result = self.db.cursor.fetchone()
            if result:
                image_path = result[0]
        
        product_data = (product_id, product_name, barcode, price, image_path)
        
        dialog = BarcodePrintDialog(self, product_data, 1)
        dialog.exec_()
    
    def delete_product(self):
        if not self.current_user.is_admin():
            QMessageBox.warning(self, 'ព្រមាន', 'អ្នកមិនមានសិទ្ធិប្រើប្រាស់មុខងារនេះទេ!')
            return
            
        current_row = self.inventory_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមជ្រើសរើសទំនិញដែលចង់លុប!')
            return
        
        product_id = int(self.inventory_table.item(current_row, 0).text())
        product_name = self.inventory_table.item(current_row, 2).text()
        
        image_item = self.inventory_table.item(current_row, 15)
        has_image = image_item and "✅" in image_item.text()
        
        reply = QMessageBox.question(self, 'បញ្ជាក់ការលុប',
                                   f'តើអ្នកពិតជាចង់លុបទំនិញ "{product_name}" មែនទេ?\n' +
                                   ('(រូបភាពនឹងត្រូវលុបផងដែរ)' if has_image else ''),
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                self.db.cursor.execute("SELECT image_path FROM products WHERE id = ?", (product_id,))
                result = self.db.cursor.fetchone()
                if result and result[0] and os.path.exists(result[0]):
                    try:
                        os.remove(result[0])
                    except:
                        pass
                
                self.db.cursor.execute("DELETE FROM products WHERE id = ?", (product_id,))
                self.db.conn.commit()
                self.load_inventory()
                self.load_products()
                QMessageBox.information(self, 'ជោគជ័យ', 'លុបទំនិញរួចរាល់!')
            except Exception as e:
                QMessageBox.critical(self, 'កំហុស', f'មិនអាចលុបទំនិញ:\n{str(e)}')
    
    def export_inventory(self):
        if not self.current_user.is_admin():
            QMessageBox.warning(self, 'ព្រមាន', 'អ្នកមិនមានសិទ្ធិប្រើប្រាស់មុខងារនេះទេ!')
            return
            
        try:
            import csv
            
            filename, _ = QFileDialog.getSaveFileName(self, 'រក្សាទុកឯកសារ', 
                                                      'inventory.csv', 
                                                      'CSV Files (*.csv)')
            
            if filename:
                with open(filename, 'w', newline='', encoding='utf-8-sig') as file:
                    writer = csv.writer(file)
                    
                    headers = []
                    for col in range(self.inventory_table.columnCount()):
                        headers.append(self.inventory_table.horizontalHeaderItem(col).text())
                    writer.writerow(headers)
                    
                    for row in range(self.inventory_table.rowCount()):
                        if not self.inventory_table.isRowHidden(row):
                            row_data = []
                            for col in range(self.inventory_table.columnCount()):
                                item = self.inventory_table.item(row, col)
                                row_data.append(item.text() if item else '')
                            writer.writerow(row_data)
                
                QMessageBox.information(self, 'ជោគជ័យ', f'បាននាំចេញទិន្នន័យទៅ {filename}')
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចនាំចេញទិន្នន័យ:\n{str(e)}')
    
    def create_reports_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        controls_group = QGroupBox("ជ្រើសរើសរបាយការណ៍")
        controls_group.setStyleSheet("QGroupBox { font-size: 13px; }")
        controls_layout = QGridLayout()
        controls_layout.setVerticalSpacing(10)
        controls_layout.setHorizontalSpacing(15)
        
        controls_layout.addWidget(QLabel("ប្រភេទរបាយការណ៍:"), 0, 0)
        self.report_type = QComboBox()
        self.report_type.addItems([
            'របាយការណ៍លក់ប្រចាំថ្ងៃ', 
            'របាយការណ៍លក់ប្រចាំខែ',
            'របាយការណ៍ស្តុក', 
            'របាយការណ៍ប្រាក់ចំណេញ',
            'របាយការណ៍លក់រាយ', 
            'របាយការណ៍លក់ដុំ',
            'របាយការណ៍ប្រចាំថ្ងៃ (លក់រាយ+ដុំ)',
            'របាយការណ៍ប្រចាំខែ (លក់រាយ+ដុំ)'
        ])
        controls_layout.addWidget(self.report_type, 0, 1)
        
        controls_layout.addWidget(QLabel("ចាប់ពីថ្ងៃ:"), 1, 0)
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-30))
        self.start_date.setCalendarPopup(True)
        controls_layout.addWidget(self.start_date, 1, 1)
        
        controls_layout.addWidget(QLabel("ដល់ថ្ងៃ:"), 2, 0)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        controls_layout.addWidget(self.end_date, 2, 1)
        
        generate_btn = QPushButton("📊 បង្កើតរបាយការណ៍")
        generate_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        generate_btn.clicked.connect(self.generate_report)
        generate_btn.setMinimumHeight(40)
        controls_layout.addWidget(generate_btn, 3, 0, 1, 2)
        
        controls_group.setLayout(controls_layout)
        layout.addWidget(controls_group)
        
        self.report_text = QTextEdit()
        self.report_text.setReadOnly(True)
        self.report_text.setStyleSheet("font-family: 'Khmer OS', 'Courier New', monospace; font-size: 12px; background-color: white; border: 1px solid #cccccc; border-radius: 5px; padding: 10px;")
        layout.addWidget(self.report_text)
        
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        
        pdf_btn = QPushButton("📄 នាំចេញជា PDF")
        pdf_btn.clicked.connect(self.export_report_pdf)
        pdf_btn.setMinimumHeight(35)
        export_layout.addWidget(pdf_btn)
        
        excel_btn = QPushButton("📊 នាំចេញជា Excel")
        excel_btn.clicked.connect(self.export_report_excel)
        excel_btn.setMinimumHeight(35)
        export_layout.addWidget(excel_btn)
        
        layout.addLayout(export_layout)
        
        return widget
    
    def generate_report(self):
        if not self.current_user.is_admin():
            QMessageBox.warning(self, 'ព្រមាន', 'អ្នកមិនមានសិទ្ធិប្រើប្រាស់មុខងារនេះទេ!')
            return
            
        report_type = self.report_type.currentText()
        start = self.start_date.date().toString('yyyy-MM-dd')
        end = self.end_date.date().toString('yyyy-MM-dd')
        
        try:
            if report_type == 'របាយការណ៍លក់ប្រចាំថ្ងៃ':
                self.generate_daily_sales_report(start, end)
            elif report_type == 'របាយការណ៍លក់ប្រចាំខែ':
                self.generate_monthly_sales_report(start, end)
            elif report_type == 'របាយការណ៍ស្តុក':
                self.generate_stock_report()
            elif report_type == 'របាយការណ៍ប្រាក់ចំណេញ':
                self.generate_profit_report(start, end)
            elif report_type == 'របាយការណ៍លក់រាយ':
                self.generate_sales_by_type_report(start, end, 'retail')
            elif report_type == 'របាយការណ៍លក់ដុំ':
                self.generate_sales_by_type_report(start, end, 'wholesale')
            elif report_type == 'របាយការណ៍ប្រចាំថ្ងៃ (លក់រាយ+ដុំ)':
                self.generate_daily_combined_report(start, end)
            elif report_type == 'របាយការណ៍ប្រចាំខែ (លក់រាយ+ដុំ)':
                self.generate_monthly_combined_report(start, end)
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍:\n{str(e)}')
    
    def generate_daily_sales_report(self, start, end):
        try:
            self.db.cursor.execute("""
                SELECT date(s.date) as sale_date, 
                       COUNT(DISTINCT s.id) as num_sales,
                       SUM(s.total_amount) as total_original,
                       SUM(s.discount_amount) as total_discount,
                       SUM(s.final_amount) as total_final,
                       AVG(s.final_amount) as avg_sale,
                       SUM(CASE WHEN s.sale_type = 'retail' THEN s.final_amount ELSE 0 END) as retail_total,
                       SUM(CASE WHEN s.sale_type = 'wholesale' THEN s.final_amount ELSE 0 END) as wholesale_total,
                       COUNT(CASE WHEN s.sale_type = 'retail' THEN 1 END) as retail_count,
                       COUNT(CASE WHEN s.sale_type = 'wholesale' THEN 1 END) as wholesale_count,
                       GROUP_CONCAT(DISTINCT s.payment_method) as payment_methods
                FROM sales s
                WHERE date(s.date) BETWEEN ? AND ?
                GROUP BY date(s.date)
                ORDER BY sale_date DESC
            """, (start, end))
            
            daily_sales = self.db.cursor.fetchall()
            
            self.db.cursor.execute("""
                SELECT SUM(s.total_amount) as total_original,
                       SUM(s.discount_amount) as total_discount,
                       SUM(s.final_amount) as total_final,
                       COUNT(DISTINCT s.id) as total_sales,
                       COUNT(DISTINCT si.product_id) as products_sold,
                       SUM(si.quantity) as total_items
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ?
            """, (start, end))
            
            totals = self.db.cursor.fetchone()
            
            report = "=" * 80 + "\n"
            report += "            របាយការណ៍លក់ប្រចាំថ្ងៃ\n"
            report += f"            ចាប់ពី {start} ដល់ {end}\n"
            report += "=" * 80 + "\n\n"
            
            if totals and totals[3] > 0:
                report += f"សង្ខេបសរុប:\n"
                report += f"  ចំនួនវិក្កយបត្រសរុប: {totals[3]}\n"
                report += f"  ចំនួនទំនិញដែលបានលក់: {totals[5] if totals[5] else 0} ដុំ\n"
                report += f"  ចំនួនប្រភេទទំនិញ: {totals[4] if totals[4] else 0} មុខ\n"
                report += f"  តម្លៃសរុបដើម: {totals[0]:,.0f}៛\n"
                report += f"  បញ្ចុះតម្លៃសរុប: {totals[1]:,.0f}៛\n"
                report += f"  តម្លៃសរុបចុងក្រោយ: {totals[2]:,.0f}៛\n"
                report += f"  តម្លៃសរុបចុងក្រោយ: ${self.currency_manager.khr_to_usd(totals[2]):.2f}\n\n"
                
                report += "-" * 80 + "\n"
                report += f"{'ថ្ងៃ':<12} {'វិក្កយបត្រ':<10} {'លក់រាយ':<10} {'លក់ដុំ':<10} {'តម្លៃដើម':<15} {'បញ្ចុះ':<12} {'តម្លៃចុង':<15}\n"
                report += "-" * 80 + "\n"
                
                for sale in daily_sales:
                    if sale[0]:
                        report += f"{sale[0]:<12} {sale[1]:<10} {sale[8] if sale[8] else 0:<10} {sale[9] if sale[9] else 0:<10} {sale[2]:<15,}៛ {sale[3]:<12,}៛ {sale[4]:<15,}៛\n"
            else:
                report += "មិនមានទិន្នន័យលក់ក្នុងអំឡុងពេលនេះទេ។\n"
            
            report += "=" * 80 + "\n"
            
            self.report_text.setText(report)
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍ប្រចាំថ្ងៃ:\n{str(e)}')
    
    def generate_monthly_sales_report(self, start, end):
        try:
            self.db.cursor.execute("""
                SELECT strftime('%Y-%m', date) as month,
                       COUNT(DISTINCT id) as num_sales,
                       SUM(total_amount) as total_original,
                       SUM(discount_amount) as total_discount,
                       SUM(final_amount) as total_final,
                       AVG(final_amount) as avg_sale,
                       SUM(CASE WHEN sale_type = 'retail' THEN final_amount ELSE 0 END) as retail_total,
                       SUM(CASE WHEN sale_type = 'wholesale' THEN final_amount ELSE 0 END) as wholesale_total,
                       COUNT(CASE WHEN sale_type = 'retail' THEN 1 END) as retail_count,
                       COUNT(CASE WHEN sale_type = 'wholesale' THEN 1 END) as wholesale_count
                FROM sales
                WHERE date(date) BETWEEN ? AND ?
                GROUP BY strftime('%Y-%m', date)
                ORDER BY month DESC
            """, (start, end))
            
            monthly_sales = self.db.cursor.fetchall()
            
            self.db.cursor.execute("""
                SELECT SUM(s.total_amount) as total_original,
                       SUM(s.discount_amount) as total_discount,
                       SUM(s.final_amount) as total_final,
                       COUNT(DISTINCT s.id) as total_sales,
                       COUNT(DISTINCT si.product_id) as products_sold,
                       SUM(si.quantity) as total_items,
                       COUNT(DISTINCT strftime('%Y-%m', s.date)) as num_months
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ?
            """, (start, end))
            
            totals = self.db.cursor.fetchone()
            
            report = "=" * 80 + "\n"
            report += "            របាយការណ៍លក់ប្រចាំខែ\n"
            report += f"            ចាប់ពី {start} ដល់ {end}\n"
            report += "=" * 80 + "\n\n"
            
            if totals and totals[3] > 0:
                report += f"សង្ខេបសរុប:\n"
                report += f"  ចំនួនខែ: {totals[6]}\n"
                report += f"  ចំនួនវិក្កយបត្រសរុប: {totals[3]}\n"
                report += f"  ចំនួនទំនិញដែលបានលក់: {totals[5] if totals[5] else 0} ដុំ\n"
                report += f"  ចំនួនប្រភេទទំនិញ: {totals[4] if totals[4] else 0} មុខ\n"
                report += f"  តម្លៃសរុបដើម: {totals[0]:,.0f}៛\n"
                report += f"  បញ្ចុះតម្លៃសរុប: {totals[1]:,.0f}៛\n"
                report += f"  តម្លៃសរុបចុងក្រោយ: {totals[2]:,.0f}៛\n"
                report += f"  តម្លៃសរុបចុងក្រោយ: ${self.currency_manager.khr_to_usd(totals[2]):.2f}\n"
                if totals[3] > 0:
                    report += f"  តម្លៃមធ្យមក្នុងមួយវិក្កយបត្រ: {totals[2] / totals[3]:,.0f}៛\n\n"
                
                report += "-" * 80 + "\n"
                report += f"{'ខែ':<8} {'វិក្កយបត្រ':<10} {'លក់រាយ':<10} {'លក់ដុំ':<10} {'តម្លៃដើម':<15} {'បញ្ចុះ':<12} {'តម្លៃចុង':<15}\n"
                report += "-" * 80 + "\n"
                
                for sale in monthly_sales:
                    if sale[0]:
                        report += f"{sale[0]:<8} {sale[1]:<10} {sale[8] if sale[8] else 0:<10} {sale[9] if sale[9] else 0:<10} {sale[2]:<15,}៛ {sale[3]:<12,}៛ {sale[4]:<15,}៛\n"
            else:
                report += "មិនមានទិន្នន័យលក់ក្នុងអំឡុងពេលនេះទេ។\n"
            
            report += "=" * 80 + "\n"
            
            self.report_text.setText(report)
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍ប្រចាំខែ:\n{str(e)}')
    
    def generate_stock_report(self):
        try:
            self.db.cursor.execute("""
                SELECT id, name, category, barcode, cost_price, selling_price_retail,
                       selling_price_wholesale, stock_retail, stock_wholesale, min_stock,
                       (stock_retail + stock_wholesale) as total_stock,
                       (stock_retail * selling_price_retail + stock_wholesale * selling_price_wholesale) as total_value
                FROM products
                WHERE active = 1
                ORDER BY total_stock ASC
            """)
            
            products = self.db.cursor.fetchall()
            
            self.db.cursor.execute("""
                SELECT COUNT(*) as total_products,
                       SUM(stock_retail + stock_wholesale) as total_items,
                       SUM(stock_retail * cost_price + stock_wholesale * cost_price) as total_cost,
                       SUM(stock_retail * selling_price_retail + stock_wholesale * selling_price_wholesale) as total_value,
                       SUM(CASE WHEN stock_retail + stock_wholesale <= min_stock THEN 1 ELSE 0 END) as low_stock,
                       SUM(CASE WHEN stock_retail + stock_wholesale = 0 THEN 1 ELSE 0 END) as out_of_stock
                FROM products
                WHERE active = 1
            """)
            
            totals = self.db.cursor.fetchone()
            
            report = "=" * 80 + "\n"
            report += "                    របាយការណ៍ស្តុកទំនិញ\n"
            report += "=" * 80 + "\n\n"
            
            if totals and totals[0] > 0:
                report += f"សង្ខេបស្តុក:\n"
                report += f"  ចំនួនទំនិញសរុប: {totals[0]} មុខ\n"
                report += f"  ចំនួនដុំសរុប: {totals[1] if totals[1] else 0} ដុំ\n"
                report += f"  តម្លៃស្តុក (ថ្លៃដើម): {totals[2]:,.0f}៛\n"
                report += f"  តម្លៃស្តុក (ថ្លៃលក់): {totals[3]:,.0f}៛\n"
                report += f"  ប្រាក់ចំណេញសក្តានុពល: {(totals[3] - totals[2]):,.0f}៛\n"
                report += f"  ទំនិញជិតអស់: {totals[4]} មុខ\n"
                report += f"  ទំនិញអស់ស្តុក: {totals[5]} មុខ\n\n"
                
                report += "-" * 80 + "\n"
                report += f"{'ID':<5} {'ឈ្មោះ':<20} {'ស្តុករាយ':<10} {'ស្តុកដុំ':<10} {'សរុប':<8} {'តម្លៃដើម':<12} {'តម្លៃលក់':<12}\n"
                report += "-" * 80 + "\n"
                
                for p in products:
                    status = ""
                    if p[10] <= p[9]:
                        status = "⚠️"
                    if p[10] == 0:
                        status = "❌"
                    
                    barcode = str(p[3] if p[3] else '')
                    if len(barcode) > 6:
                        barcode = barcode[-6:]
                    
                    report += f"{p[0]:<5} {p[1][:20]:<20} {p[7]:<10} {p[8]:<10} {p[10]:<8} {p[4]:<12,}៛ {p[5]:<12,}៛ {status}\n"
            else:
                report += "មិនមានទិន្នន័យស្តុកទេ។\n"
            
            report += "=" * 80 + "\n"
            
            self.report_text.setText(report)
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍ស្តុក:\n{str(e)}')
    
    def generate_profit_report(self, start, end):
        try:
            self.db.cursor.execute("""
                SELECT s.date,
                       si.product_id,
                       p.name,
                       si.quantity,
                       si.sale_type,
                       si.original_price,
                       p.cost_price,
                       si.final_price,
                       (si.final_price - p.cost_price) * si.quantity as profit,
                       si.discount_amount,
                       si.discount_percent
                FROM sales s
                JOIN sale_items si ON s.id = si.sale_id
                JOIN products p ON si.product_id = p.id
                WHERE date(s.date) BETWEEN ? AND ?
                ORDER BY s.date DESC
            """, (start, end))
            
            items = self.db.cursor.fetchall()
            
            self.db.cursor.execute("""
                SELECT COUNT(DISTINCT s.id) as total_sales,
                       SUM(s.final_amount) as total_revenue,
                       SUM(p.cost_price * si.quantity) as total_cost,
                       SUM((si.final_price - p.cost_price) * si.quantity) as total_profit,
                       SUM(si.quantity) as total_items,
                       COUNT(DISTINCT si.product_id) as products_sold,
                       SUM(si.discount_amount * si.quantity) as total_discount
                FROM sales s
                JOIN sale_items si ON s.id = si.sale_id
                JOIN products p ON si.product_id = p.id
                WHERE date(s.date) BETWEEN ? AND ?
            """, (start, end))
            
            totals = self.db.cursor.fetchone()
            
            report = "=" * 80 + "\n"
            report += "                របាយការណ៍ប្រាក់ចំណេញ\n"
            report += f"                ចាប់ពី {start} ដល់ {end}\n"
            report += "=" * 80 + "\n\n"
            
            if totals and totals[0] > 0:
                report += f"សង្ខេបប្រាក់ចំណេញ:\n"
                report += f"  ចំនួនវិក្កយបត្រ: {totals[0]}\n"
                report += f"  ចំនួនទំនិញដែលបានលក់: {totals[4] if totals[4] else 0} ដុំ\n"
                report += f"  ចំនួនប្រភេទទំនិញ: {totals[5] if totals[5] else 0} មុខ\n"
                report += f"  បញ្ចុះតម្លៃសរុប: {totals[6] if totals[6] else 0:,.0f}៛\n"
                report += f"  ចំណូលសរុប: {totals[1]:,.0f}៛\n"
                report += f"  ថ្លៃដើមសរុប: {totals[2]:,.0f}៛\n"
                report += f"  ប្រាក់ចំណេញសរុប: {totals[3]:,.0f}៛\n"
                if totals[1] > 0:
                    report += f"  រឹមប្រាក់ចំណេញ: {(totals[3] / totals[1] * 100):.2f}%\n\n"
                
                report += "-" * 80 + "\n"
                report += f"{'ថ្ងៃ':<12} {'ឈ្មោះទំនិញ':<20} {'បរិមាណ':<8} {'ប្រភេទ':<8} {'ថ្លៃដើម':<12} {'ថ្លៃលក់':<12} {'ចំណេញ':<12}\n"
                report += "-" * 80 + "\n"
                
                for item in items:
                    profit_per_item = (item[7] - item[6]) * item[3]
                    report += f"{item[0][:10]:<12} {item[2][:20]:<20} {item[3]:<8} {item[4]:<8} {item[6]:<12,}៛ {item[7]:<12,}៛ {profit_per_item:<12,}៛\n"
            else:
                report += "មិនមានទិន្នន័យលក់ក្នុងអំឡុងពេលនេះទេ។\n"
            
            report += "=" * 80 + "\n"
            
            self.report_text.setText(report)
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍ប្រាក់ចំណេញ:\n{str(e)}')
    
    def generate_sales_by_type_report(self, start, end, sale_type):
        try:
            type_name = "លក់រាយ" if sale_type == 'retail' else "លក់ដុំ"
            
            self.db.cursor.execute("""
                SELECT date(s.date) as sale_date,
                       s.id as sale_id,
                       s.total_amount,
                       s.discount_amount,
                       s.final_amount,
                       s.payment_method,
                       u.full_name as cashier,
                       COUNT(si.id) as items_count,
                       SUM(si.quantity) as total_quantity
                FROM sales s
                JOIN users u ON s.user_id = u.id
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ? AND s.sale_type = ?
                GROUP BY s.id
                ORDER BY s.date DESC
            """, (start, end, sale_type))
            
            sales = self.db.cursor.fetchall()
            
            self.db.cursor.execute("""
                SELECT COUNT(DISTINCT s.id) as total_sales,
                       SUM(s.final_amount) as total_revenue,
                       AVG(s.final_amount) as avg_sale,
                       SUM(si.quantity) as total_items,
                       COUNT(DISTINCT si.product_id) as products_sold,
                       SUM(s.discount_amount) as total_discount,
                       COUNT(DISTINCT s.user_id) as num_cashiers
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ? AND s.sale_type = ?
            """, (start, end, sale_type))
            
            totals = self.db.cursor.fetchone()
            
            report = "=" * 80 + "\n"
            report += f"            របាយការណ៍{type_name}\n"
            report += f"            ចាប់ពី {start} ដល់ {end}\n"
            report += "=" * 80 + "\n\n"
            
            if totals and totals[0] > 0:
                report += f"សង្ខេប{type_name}:\n"
                report += f"  ចំនួនវិក្កយបត្រ: {totals[0]}\n"
                report += f"  ចំនួនទំនិញដែលបានលក់: {totals[3] if totals[3] else 0} ដុំ\n"
                report += f"  ចំនួនប្រភេទទំនិញ: {totals[4] if totals[4] else 0} មុខ\n"
                report += f"  បញ្ចុះតម្លៃសរុប: {totals[5] if totals[5] else 0:,.0f}៛\n"
                report += f"  ចំណូលសរុប: {totals[1]:,.0f}៛\n"
                report += f"  ចំណូលសរុប: ${self.currency_manager.khr_to_usd(totals[1]):.2f}\n"
                report += f"  តម្លៃមធ្យមក្នុងមួយវិក្កយបត្រ: {totals[2]:,.0f}៛\n"
                report += f"  ចំនួនអ្នកលក់: {totals[6]}\n\n"
                
                report += "-" * 80 + "\n"
                report += f"{'ថ្ងៃ':<12} {'លេខវិក្កយបត្រ':<15} {'ចំនួនដុំ':<10} {'បញ្ចុះ':<12} {'តម្លៃចុង':<15} {'វិធីបង់':<15}\n"
                report += "-" * 80 + "\n"
                
                for sale in sales:
                    if sale[0]:
                        report += f"{sale[0]:<12} {sale[1]:06d}{'':<9} {sale[8]:<10} {sale[3]:<12,}៛ {sale[4]:<15,}៛ {sale[5]:<15}\n"
            else:
                report += f"មិនមានទិន្នន័យ{type_name}ក្នុងអំឡុងពេលនេះទេ។\n"
            
            report += "=" * 80 + "\n"
            
            self.report_text.setText(report)
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍{type_name}:\n{str(e)}')
    
    def generate_daily_combined_report(self, start, end):
        try:
            self.db.cursor.execute("""
                SELECT date(s.date) as sale_date,
                       COUNT(DISTINCT s.id) as total_sales,
                       SUM(CASE WHEN s.sale_type = 'retail' THEN 1 ELSE 0 END) as retail_count,
                       SUM(CASE WHEN s.sale_type = 'wholesale' THEN 1 ELSE 0 END) as wholesale_count,
                       SUM(CASE WHEN s.sale_type = 'retail' THEN s.final_amount ELSE 0 END) as retail_amount,
                       SUM(CASE WHEN s.sale_type = 'wholesale' THEN s.final_amount ELSE 0 END) as wholesale_amount,
                       SUM(s.final_amount) as total_amount,
                       SUM(s.discount_amount) as total_discount,
                       SUM(si.quantity) as total_items
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ?
                GROUP BY date(s.date)
                ORDER BY sale_date DESC
            """, (start, end))
            
            daily_data = self.db.cursor.fetchall()
            
            self.db.cursor.execute("""
                SELECT COUNT(DISTINCT s.id) as total_sales,
                       SUM(s.final_amount) as total_amount,
                       SUM(s.discount_amount) as total_discount,
                       SUM(si.quantity) as total_items,
                       AVG(s.final_amount) as avg_sale,
                       SUM(CASE WHEN s.sale_type = 'retail' THEN s.final_amount ELSE 0 END) as total_retail,
                       SUM(CASE WHEN s.sale_type = 'wholesale' THEN s.final_amount ELSE 0 END) as total_wholesale,
                       COUNT(DISTINCT CASE WHEN s.sale_type = 'retail' THEN s.id END) as retail_count,
                       COUNT(DISTINCT CASE WHEN s.sale_type = 'wholesale' THEN s.id END) as wholesale_count
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ?
            """, (start, end))
            
            totals = self.db.cursor.fetchone()
            
            report = "=" * 80 + "\n"
            report += "            របាយការណ៍ប្រចាំថ្ងៃ (លក់រាយ+លក់ដុំ)\n"
            report += f"            ចាប់ពី {start} ដល់ {end}\n"
            report += "=" * 80 + "\n\n"
            
            if totals and totals[0] > 0:
                report += f"សង្ខេបសរុប:\n"
                report += f"  ចំនួនវិក្កយបត្រសរុប: {totals[0]}\n"
                report += f"    - លក់រាយ: {totals[7]} វិក្កយបត្រ\n"
                report += f"    - លក់ដុំ: {totals[8]} វិក្កយបត្រ\n"
                report += f"  ចំនួនទំនិញដែលបានលក់: {totals[3] if totals[3] else 0} ដុំ\n"
                report += f"  បញ្ចុះតម្លៃសរុប: {totals[2]:,.0f}៛\n"
                report += f"  ចំណូលសរុប: {totals[1]:,.0f}៛\n"
                report += f"  ចំណូលសរុប: ${self.currency_manager.khr_to_usd(totals[1]):.2f}\n"
                report += f"    - លក់រាយ: {totals[5]:,.0f}៛ (${self.currency_manager.khr_to_usd(totals[5]):.2f})\n"
                report += f"    - លក់ដុំ: {totals[6]:,.0f}៛ (${self.currency_manager.khr_to_usd(totals[6]):.2f})\n"
                report += f"  តម្លៃមធ្យមក្នុងមួយវិក្កយបត្រ: {totals[4]:,.0f}៛\n\n"
                
                report += "-" * 80 + "\n"
                report += f"{'ថ្ងៃ':<12} {'វិក្កយបត្រ':<10} {'លក់រាយ':<10} {'លក់ដុំ':<10} {'តម្លៃរាយ':<12} {'តម្លៃដុំ':<12} {'សរុប':<12}\n"
                report += "-" * 80 + "\n"
                
                for data in daily_data:
                    if data[0]:
                        report += f"{data[0]:<12} {data[1]:<10} {data[2]:<10} {data[3]:<10} {data[4]:<12,}៛ {data[5]:<12,}៛ {data[6]:<12,}៛\n"
            else:
                report += "មិនមានទិន្នន័យលក់ក្នុងអំឡុងពេលនេះទេ។\n"
            
            report += "=" * 80 + "\n"
            
            self.report_text.setText(report)
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍ប្រចាំថ្ងៃ:\n{str(e)}')
    
    def generate_monthly_combined_report(self, start, end):
        try:
            self.db.cursor.execute("""
                SELECT strftime('%Y-%m', s.date) as month,
                       COUNT(DISTINCT s.id) as total_sales,
                       SUM(CASE WHEN s.sale_type = 'retail' THEN 1 ELSE 0 END) as retail_count,
                       SUM(CASE WHEN s.sale_type = 'wholesale' THEN 1 ELSE 0 END) as wholesale_count,
                       SUM(CASE WHEN s.sale_type = 'retail' THEN s.final_amount ELSE 0 END) as retail_amount,
                       SUM(CASE WHEN s.sale_type = 'wholesale' THEN s.final_amount ELSE 0 END) as wholesale_amount,
                       SUM(s.final_amount) as total_amount,
                       SUM(s.discount_amount) as total_discount,
                       SUM(si.quantity) as total_items
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ?
                GROUP BY strftime('%Y-%m', s.date)
                ORDER BY month DESC
            """, (start, end))
            
            monthly_data = self.db.cursor.fetchall()
            
            self.db.cursor.execute("""
                SELECT COUNT(DISTINCT s.id) as total_sales,
                       SUM(s.final_amount) as total_amount,
                       SUM(s.discount_amount) as total_discount,
                       SUM(si.quantity) as total_items,
                       AVG(s.final_amount) as avg_sale,
                       SUM(CASE WHEN s.sale_type = 'retail' THEN s.final_amount ELSE 0 END) as total_retail,
                       SUM(CASE WHEN s.sale_type = 'wholesale' THEN s.final_amount ELSE 0 END) as total_wholesale,
                       COUNT(DISTINCT CASE WHEN s.sale_type = 'retail' THEN s.id END) as retail_count,
                       COUNT(DISTINCT CASE WHEN s.sale_type = 'wholesale' THEN s.id END) as wholesale_count,
                       COUNT(DISTINCT strftime('%Y-%m', s.date)) as num_months
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE date(s.date) BETWEEN ? AND ?
            """, (start, end))
            
            totals = self.db.cursor.fetchone()
            
            report = "=" * 80 + "\n"
            report += "            របាយការណ៍ប្រចាំខែ (លក់រាយ+លក់ដុំ)\n"
            report += f"            ចាប់ពី {start} ដល់ {end}\n"
            report += "=" * 80 + "\n\n"
            
            if totals and totals[0] > 0:
                report += f"សង្ខេបសរុប ({totals[9]} ខែ):\n"
                report += f"  ចំនួនវិក្កយបត្រសរុប: {totals[0]}\n"
                report += f"    - លក់រាយ: {totals[7]} វិក្កយបត្រ\n"
                report += f"    - លក់ដុំ: {totals[8]} វិក្កយបត្រ\n"
                report += f"  ចំនួនទំនិញដែលបានលក់: {totals[3] if totals[3] else 0} ដុំ\n"
                report += f"  បញ្ចុះតម្លៃសរុប: {totals[2]:,.0f}៛\n"
                report += f"  ចំណូលសរុប: {totals[1]:,.0f}៛\n"
                report += f"  ចំណូលសរុប: ${self.currency_manager.khr_to_usd(totals[1]):.2f}\n"
                report += f"    - លក់រាយ: {totals[5]:,.0f}៛ (${self.currency_manager.khr_to_usd(totals[5]):.2f})\n"
                report += f"    - លក់ដុំ: {totals[6]:,.0f}៛ (${self.currency_manager.khr_to_usd(totals[6]):.2f})\n"
                report += f"  តម្លៃមធ្យមក្នុងមួយវិក្កយបត្រ: {totals[4]:,.0f}៛\n\n"
                
                report += "-" * 80 + "\n"
                report += f"{'ខែ':<8} {'វិក្កយបត្រ':<10} {'លក់រាយ':<10} {'លក់ដុំ':<10} {'តម្លៃរាយ':<12} {'តម្លៃដុំ':<12} {'សរុប':<12}\n"
                report += "-" * 80 + "\n"
                
                for data in monthly_data:
                    if data[0]:
                        report += f"{data[0]:<8} {data[1]:<10} {data[2]:<10} {data[3]:<10} {data[4]:<12,}៛ {data[5]:<12,}៛ {data[6]:<12,}៛\n"
            else:
                report += "មិនមានទិន្នន័យលក់ក្នុងអំឡុងពេលនេះទេ។\n"
            
            report += "=" * 80 + "\n"
            
            self.report_text.setText(report)
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចបង្កើតរបាយការណ៍ប្រចាំខែ:\n{str(e)}')
    
    def export_report_pdf(self):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            filename, _ = QFileDialog.getSaveFileName(self, 'រក្សាទុក PDF', 
                                                      f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf', 
                                                      'PDF Files (*.pdf)')
            
            if filename:
                doc = SimpleDocTemplate(filename, pagesize=A4)
                styles = getSampleStyleSheet()
                story = []
                
                title = Paragraph(f"<para alignment='center'><b>របាយការណ៍</b></para>", styles['Title'])
                story.append(title)
                story.append(Spacer(1, 0.25*inch))
                
                date_text = Paragraph(f"<para alignment='center'>ចាប់ពី {self.start_date.date().toString('yyyy-MM-dd')} ដល់ {self.end_date.date().toString('yyyy-MM-dd')}</para>", styles['Normal'])
                story.append(date_text)
                story.append(Spacer(1, 0.25*inch))
                
                report_content = self.report_text.toPlainText()
                
                lines = report_content.split('\n')
                for line in lines:
                    if line.strip():
                        p = Paragraph(line.replace('-', ' ').replace('=', ' '), styles['Normal'])
                        story.append(p)
                        story.append(Spacer(1, 0.1*inch))
                
                doc.build(story)
                
                QMessageBox.information(self, 'ជោគជ័យ', f'បាននាំចេញរបាយការណ៍ទៅ {filename}')
                
        except ImportError:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមដំឡើង reportlab ជាមុនសិន:\npip install reportlab')
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចនាំចេញ PDF:\n{str(e)}')
    
    def export_report_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            filename, _ = QFileDialog.getSaveFileName(self, 'រក្សាទុក Excel', 
                                                      f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', 
                                                      'Excel Files (*.xlsx)')
            
            if filename:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "របាយការណ៍"
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                lines = self.report_text.toPlainText().split('\n')
                
                for i, line in enumerate(lines, 1):
                    ws.cell(row=i, column=1, value=line)
                    
                    if '=' in line or '-' in line:
                        ws.cell(row=i, column=1).font = Font(bold=True)
                    
                    if i == 1:
                        ws.cell(row=i, column=1).font = Font(size=16, bold=True)
                        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
                    
                    if 'សរុប' in line:
                        ws.cell(row=i, column=1).font = Font(bold=True, color="4CAF50")
                    
                    if 'មិនមាន' in line:
                        ws.cell(row=i, column=1).font = Font(color="FF0000")
                    
                    ws.cell(row=i, column=1).border = thin_border
                
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 100)
                    ws.column_dimensions[column].width = adjusted_width
                
                wb.save(filename)
                
                QMessageBox.information(self, 'ជោគជ័យ', f'បាននាំចេញរបាយការណ៍ទៅ {filename}')
                
        except ImportError:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមដំឡើង openpyxl ជាមុនសិន:\npip install openpyxl')
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចនាំចេញ Excel:\n{str(e)}')

# ==================== ថ្នាក់ ChangePasswordDialog ====================
class ChangePasswordDialog(QDialog):
    def __init__(self, db, user_id):
        super().__init__()
        self.db = db
        self.user_id = user_id
        self.setWindowTitle('ប្តូរពាក្យសម្ងាត់')
        self.setModal(True)
        self.setFixedSize(450, 400)
        
        font = QFont('Khmer OS', 12)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
                padding: 2px 0;
            }
            QLineEdit {
                padding: 12px 15px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
                min-height: 20px;
            }
            QLineEdit:focus {
                border-color: #4CAF50;
            }
            QLineEdit::placeholder {
                color: #999;
                font-style: italic;
            }
            QPushButton {
                padding: 12px;
                border: none;
                border-radius: 5px;
                font-size: 15px;
                font-weight: bold;
                min-height: 40px;
            }
            QPushButton[text="រក្សាទុក"] {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton[text="រក្សាទុក"]:hover {
                background-color: #45a049;
            }
            QPushButton[text="បោះបង់"] {
                background-color: #f44336;
                color: white;
            }
            QPushButton[text="បោះបង់"]:hover {
                background-color: #da190b;
            }
        """)
        
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        
        title = QLabel("ប្តូរពាក្យសម្ងាត់")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4CAF50; qproperty-alignment: AlignCenter; padding: 10px;")
        layout.addWidget(title)
        
        layout.addWidget(QLabel("ពាក្យសម្ងាត់បច្ចុប្បន្ន:"))
        self.current_password = QLineEdit()
        self.current_password.setEchoMode(QLineEdit.Password)
        self.current_password.setPlaceholderText("បញ្ចូលពាក្យសម្ងាត់បច្ចុប្បន្ន...")
        layout.addWidget(self.current_password)
        
        layout.addWidget(QLabel("ពាក្យសម្ងាត់ថ្មី:"))
        self.new_password = QLineEdit()
        self.new_password.setEchoMode(QLineEdit.Password)
        self.new_password.setPlaceholderText("បញ្ចូលពាក្យសម្ងាត់ថ្មី...")
        layout.addWidget(self.new_password)
        
        layout.addWidget(QLabel("បញ្ជាក់ពាក្យសម្ងាត់ថ្មី:"))
        self.confirm_password = QLineEdit()
        self.confirm_password.setEchoMode(QLineEdit.Password)
        self.confirm_password.setPlaceholderText("បញ្ចូលពាក្យសម្ងាត់ម្តងទៀត...")
        layout.addWidget(self.confirm_password)
        
        layout.addStretch()
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        save_btn = QPushButton("រក្សាទុក")
        save_btn.clicked.connect(self.save_password)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton("បោះបង់")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #f44336; font-size: 13px; padding: 5px; background-color: #ffebee; border-radius: 4px;")
        self.status_label.hide()
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
    
    def save_password(self):
        current = self.current_password.text()
        new = self.new_password.text()
        confirm = self.confirm_password.text()
        
        if not current or not new or not confirm:
            self.show_error("សូមបញ្ចូលព័ត៌មានឱ្យបានពេញលេញ!")
            return
        
        if new != confirm:
            self.show_error("ពាក្យសម្ងាត់ថ្មីមិនត្រូវគ្នា!")
            return
        
        if len(new) < 6:
            self.show_error("ពាក្យសម្ងាត់ត្រូវមានយ៉ាងហោចណាស់ 6 តួអក្សរ!")
            return
        
        hashed_current = hashlib.sha256(current.encode()).hexdigest()
        self.db.cursor.execute("SELECT password FROM users WHERE id = ?", (self.user_id,))
        db_password = self.db.cursor.fetchone()[0]
        
        if hashed_current != db_password:
            self.show_error("ពាក្យសម្ងាត់បច្ចុប្បន្នមិនត្រឹមត្រូវ!")
            return
        
        hashed_new = hashlib.sha256(new.encode()).hexdigest()
        self.db.cursor.execute("UPDATE users SET password = ?, updated_at = ? WHERE id = ?", 
                              (hashed_new, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), self.user_id))
        self.db.conn.commit()
        
        QMessageBox.information(self, "ជោគជ័យ", "បានប្តូរពាក្យសម្ងាត់រួចរាល់!")
        self.accept()
    
    def show_error(self, message):
        self.status_label.setText(message)
        self.status_label.show()
        QTimer.singleShot(3000, lambda: self.status_label.hide())

# ==================== ថ្នាក់ UserManagementDialog ====================
class UserManagementDialog(QDialog):
    def __init__(self, parent, db):
        super().__init__(parent)
        self.parent_window = parent
        self.db = db
        self.setWindowTitle('គ្រប់គ្រងអ្នកប្រើប្រាស់')
        self.setModal(True)
        self.setMinimumWidth(850)
        self.setMinimumHeight(550)
        
        font = QFont('Khmer OS', 11)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QTableWidget {
                font-size: 12px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
            }
            QPushButton {
                padding: 8px 15px;
                font-size: 13px;
                min-height: 35px;
            }
            QPushButton[text="បន្ថែមអ្នកប្រើប្រាស់ថ្មី"] {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton[text="កែប្រែ"] {
                background-color: #FF9800;
                color: white;
            }
            QPushButton[text="លុប"] {
                background-color: #f44336;
                color: white;
            }
            QPushButton[text="ធ្វើឱ្យថ្មី"] {
                background-color: #2196F3;
                color: white;
            }
            QPushButton:hover {
                opacity: 0.9;
            }
        """)
        
        self.initUI()
        self.load_users()
    
    def initUI(self):
        layout = QVBoxLayout()
        
        title = QLabel("គ្រប់គ្រងអ្នកប្រើប្រាស់")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #4CAF50; padding: 10px;")
        layout.addWidget(title)
        
        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)
        
        add_btn = QPushButton("➕ បន្ថែមអ្នកប្រើប្រាស់ថ្មី")
        add_btn.clicked.connect(self.add_user)
        toolbar.addWidget(add_btn)
        
        edit_btn = QPushButton("✏️ កែប្រែ")
        edit_btn.clicked.connect(self.edit_user)
        toolbar.addWidget(edit_btn)
        
        delete_btn = QPushButton("🗑️ លុប")
        delete_btn.clicked.connect(self.delete_user)
        toolbar.addWidget(delete_btn)
        
        refresh_btn = QPushButton("🔄 ធ្វើឱ្យថ្មី")
        refresh_btn.clicked.connect(self.load_users)
        toolbar.addWidget(refresh_btn)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(6)
        self.users_table.setHorizontalHeaderLabels(['ID', 'ឈ្មោះអ្នកប្រើ', 'ឈ្មោះពេញ', 'តួនាទី', 'ស្ថានភាព', 'បង្កើតនៅ'])
        self.users_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.users_table.setAlternatingRowColors(True)
        
        self.users_table.setColumnWidth(0, 50)
        self.users_table.setColumnWidth(1, 150)
        self.users_table.setColumnWidth(2, 200)
        self.users_table.setColumnWidth(3, 120)
        self.users_table.setColumnWidth(4, 100)
        self.users_table.setColumnWidth(5, 150)
        
        self.users_table.verticalHeader().setDefaultSectionSize(35)
        
        layout.addWidget(self.users_table)
        
        button_layout = QHBoxLayout()
        close_btn = QPushButton("បិទ")
        close_btn.setStyleSheet("background-color: #f44336; color: white; padding: 10px 20px; font-weight: bold;")
        close_btn.clicked.connect(self.accept)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def load_users(self):
        try:
            self.db.cursor.execute("""
                SELECT id, username, full_name, role, 
                       CASE WHEN active = 1 THEN 'សកម្ម' ELSE 'មិនសកម្ម' END,
                       created_at
                FROM users
                ORDER BY id
            """)
            users = self.db.cursor.fetchall()
            
            self.users_table.setRowCount(len(users))
            for row, user in enumerate(users):
                for col, value in enumerate(user):
                    item = QTableWidgetItem(str(value))
                    if col == 3:
                        if value == 'admin':
                            item.setForeground(QBrush(QColor(76, 175, 80)))
                            item.setText("អ្នកគ្រប់គ្រង")
                        else:
                            item.setForeground(QBrush(QColor(33, 150, 243)))
                            item.setText("អ្នកលក់")
                    elif col == 4:
                        if value == 'សកម្ម':
                            item.setForeground(QBrush(QColor(76, 175, 80)))
                        else:
                            item.setForeground(QBrush(QColor(244, 67, 54)))
                    self.users_table.setItem(row, col, item)
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចផ្ទុកទិន្នន័យ: {str(e)}')
    
    def add_user(self):
        dialog = AddUserDialog(self, self.db, self.parent_window.current_user)
        if dialog.exec_():
            self.load_users()
    
    def edit_user(self):
        current_row = self.users_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមជ្រើសរើសអ្នកប្រើប្រាស់ដែលចង់កែប្រែ!')
            return
        
        user_id = int(self.users_table.item(current_row, 0).text())
        dialog = EditUserDialog(self, self.db, user_id)
        if dialog.exec_():
            self.load_users()
    
    def delete_user(self):
        current_row = self.users_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'សូមជ្រើសរើសអ្នកប្រើប្រាស់ដែលចង់លុប!')
            return
        
        user_id = int(self.users_table.item(current_row, 0).text())
        username = self.users_table.item(current_row, 1).text()
        
        if user_id == self.parent_window.current_user.id:
            QMessageBox.warning(self, 'ព្រមាន', 'អ្នកមិនអាចលុបគណនីខ្លួនឯងបានទេ!')
            return
        
        reply = QMessageBox.question(self, 'បញ្ជាក់ការលុប',
                                    f'តើអ្នកពិតជាចង់លុបអ្នកប្រើប្រាស់ "{username}" មែនទេ?',
                                    QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                self.db.cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
                self.db.conn.commit()
                self.load_users()
                QMessageBox.information(self, 'ជោគជ័យ', 'លុបអ្នកប្រើប្រាស់រួចរាល់!')
            except Exception as e:
                QMessageBox.critical(self, 'កំហុស', f'មិនអាចលុបអ្នកប្រើប្រាស់:\n{str(e)}')

# ==================== ថ្នាក់ AddUserDialog ====================
class AddUserDialog(QDialog):
    def __init__(self, parent, db, current_user):
        super().__init__(parent)
        self.db = db
        self.current_user = current_user
        self.setWindowTitle('បន្ថែមអ្នកប្រើប្រាស់ថ្មី')
        self.setModal(True)
        self.setFixedSize(500, 500)
        
        font = QFont('Khmer OS', 12)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
                padding: 2px 0;
            }
            QLineEdit {
                padding: 12px 15px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
                min-height: 20px;
            }
            QLineEdit:focus {
                border-color: #4CAF50;
            }
            QLineEdit::placeholder {
                color: #999;
                font-style: italic;
            }
            QComboBox {
                padding: 12px 15px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
            }
            QCheckBox {
                font-size: 14px;
                padding: 5px;
            }
            QPushButton {
                padding: 12px;
                border: none;
                border-radius: 5px;
                font-size: 15px;
                font-weight: bold;
                min-height: 40px;
            }
            QPushButton[text="រក្សាទុក"] {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton[text="បោះបង់"] {
                background-color: #f44336;
                color: white;
            }
            QPushButton:hover {
                opacity: 0.9;
            }
        """)
        
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        
        title = QLabel("បន្ថែមអ្នកប្រើប្រាស់ថ្មី")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4CAF50; qproperty-alignment: AlignCenter; padding: 10px;")
        layout.addWidget(title)
        
        layout.addWidget(QLabel("ឈ្មោះអ្នកប្រើប្រាស់:"))
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("បញ្ចូលឈ្មោះអ្នកប្រើប្រាស់...")
        layout.addWidget(self.username_input)
        
        layout.addWidget(QLabel("ឈ្មោះពេញ:"))
        self.fullname_input = QLineEdit()
        self.fullname_input.setPlaceholderText("បញ្ចូលឈ្មោះពេញ...")
        layout.addWidget(self.fullname_input)
        
        layout.addWidget(QLabel("ពាក្យសម្ងាត់:"))
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText("បញ្ចូលពាក្យសម្ងាត់...")
        layout.addWidget(self.password_input)
        
        layout.addWidget(QLabel("បញ្ជាក់ពាក្យសម្ងាត់:"))
        self.confirm_input = QLineEdit()
        self.confirm_input.setEchoMode(QLineEdit.Password)
        self.confirm_input.setPlaceholderText("បញ្ចូលពាក្យសម្ងាត់ម្តងទៀត...")
        layout.addWidget(self.confirm_input)
        
        layout.addWidget(QLabel("តួនាទី:"))
        self.role_combo = QComboBox()
        self.role_combo.addItem("អ្នកលក់", "cashier")
        self.role_combo.addItem("អ្នកគ្រប់គ្រង", "admin")
        layout.addWidget(self.role_combo)
        
        self.active_check = QCheckBox("សកម្ម")
        self.active_check.setChecked(True)
        layout.addWidget(self.active_check)
        
        layout.addStretch()
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        save_btn = QPushButton("រក្សាទុក")
        save_btn.clicked.connect(self.save_user)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton("បោះបង់")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #f44336; font-size: 13px; padding: 5px; background-color: #ffebee; border-radius: 4px;")
        self.status_label.hide()
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
    
    def save_user(self):
        username = self.username_input.text().strip()
        fullname = self.fullname_input.text().strip()
        password = self.password_input.text()
        confirm = self.confirm_input.text()
        role = self.role_combo.currentData()
        active = 1 if self.active_check.isChecked() else 0
        
        if not username or not fullname or not password:
            self.show_error("សូមបញ្ចូលព័ត៌មានឱ្យបានពេញលេញ!")
            return
        
        if password != confirm:
            self.show_error("ពាក្យសម្ងាត់មិនត្រូវគ្នា!")
            return
        
        if len(password) < 6:
            self.show_error("ពាក្យសម្ងាត់ត្រូវមានយ៉ាងហោចណាស់ 6 តួអក្សរ!")
            return
        
        try:
            self.db.cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
            if self.db.cursor.fetchone():
                self.show_error("ឈ្មោះអ្នកប្រើប្រាស់នេះមានរួចហើយ!")
                return
            
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.db.cursor.execute("""
                INSERT INTO users (username, password, full_name, role, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (username, hashed_password, fullname, role, active, current_time, current_time))
            
            self.db.conn.commit()
            self.accept()
            
        except Exception as e:
            self.show_error(f"មានបញ្ហា: {str(e)}")
    
    def show_error(self, message):
        self.status_label.setText(message)
        self.status_label.show()
        QTimer.singleShot(3000, lambda: self.status_label.hide())

# ==================== ថ្នាក់ EditUserDialog ====================
class EditUserDialog(QDialog):
    def __init__(self, parent, db, user_id):
        super().__init__(parent)
        self.db = db
        self.user_id = user_id
        self.setWindowTitle('កែប្រែព័ត៌មានអ្នកប្រើប្រាស់')
        self.setModal(True)
        self.setFixedSize(500, 450)
        
        font = QFont('Khmer OS', 12)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
                padding: 2px 0;
            }
            QLineEdit, QLabel[readonly] {
                padding: 12px 15px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
                min-height: 20px;
            }
            QLabel[readonly] {
                background-color: #f5f5f5;
                border: 1px solid #ddd;
            }
            QComboBox {
                padding: 12px 15px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                background-color: white;
            }
            QCheckBox {
                font-size: 14px;
                padding: 5px;
            }
            QPushButton {
                padding: 12px;
                border: none;
                border-radius: 5px;
                font-size: 15px;
                font-weight: bold;
                min-height: 40px;
            }
            QPushButton[text="រក្សាទុក"] {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton[text="បោះបង់"] {
                background-color: #f44336;
                color: white;
            }
            QPushButton[text="ប្តូរពាក្យសម្ងាត់"] {
                background-color: #2196F3;
                color: white;
            }
            QPushButton:hover {
                opacity: 0.9;
            }
        """)
        
        self.load_user_data()
        self.initUI()
    
    def load_user_data(self):
        self.db.cursor.execute("SELECT username, full_name, role, active FROM users WHERE id = ?", (self.user_id,))
        self.user_data = self.db.cursor.fetchone()
    
    def initUI(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        
        title = QLabel("កែប្រែព័ត៌មានអ្នកប្រើប្រាស់")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4CAF50; qproperty-alignment: AlignCenter; padding: 10px;")
        layout.addWidget(title)
        
        layout.addWidget(QLabel("ឈ្មោះអ្នកប្រើប្រាស់:"))
        self.username_label = QLabel(self.user_data[0])
        self.username_label.setProperty("readonly", True)
        layout.addWidget(self.username_label)
        
        layout.addWidget(QLabel("ឈ្មោះពេញ:"))
        self.fullname_input = QLineEdit()
        self.fullname_input.setText(self.user_data[1])
        layout.addWidget(self.fullname_input)
        
        layout.addWidget(QLabel("តួនាទី:"))
        self.role_combo = QComboBox()
        self.role_combo.addItem("អ្នកលក់", "cashier")
        self.role_combo.addItem("អ្នកគ្រប់គ្រង", "admin")
        
        index = 0 if self.user_data[2] == 'cashier' else 1
        self.role_combo.setCurrentIndex(index)
        layout.addWidget(self.role_combo)
        
        self.active_check = QCheckBox("សកម្ម")
        self.active_check.setChecked(bool(self.user_data[3]))
        layout.addWidget(self.active_check)
        
        change_pass_btn = QPushButton("ប្តូរពាក្យសម្ងាត់")
        change_pass_btn.clicked.connect(self.change_password)
        layout.addWidget(change_pass_btn)
        
        layout.addStretch()
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        save_btn = QPushButton("រក្សាទុក")
        save_btn.clicked.connect(self.save_user)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton("បោះបង់")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #f44336; font-size: 13px; padding: 5px; background-color: #ffebee; border-radius: 4px;")
        self.status_label.hide()
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
    
    def save_user(self):
        fullname = self.fullname_input.text().strip()
        role = self.role_combo.currentData()
        active = 1 if self.active_check.isChecked() else 0
        
        if not fullname:
            self.show_error("សូមបញ្ចូលឈ្មោះពេញ!")
            return
        
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.db.cursor.execute("""
                UPDATE users 
                SET full_name = ?, role = ?, active = ?, updated_at = ?
                WHERE id = ?
            """, (fullname, role, active, current_time, self.user_id))
            
            self.db.conn.commit()
            self.accept()
            
        except Exception as e:
            self.show_error(f"មានបញ្ហា: {str(e)}")
    
    def change_password(self):
        dialog = ChangePasswordDialog(self.db, self.user_id)
        dialog.exec_()
    
    def show_error(self, message):
        self.status_label.setText(message)
        self.status_label.show()
        QTimer.singleShot(3000, lambda: self.status_label.hide())

# ==================== ថ្នាក់ ImageLabel និង ProductImageWidget ====================
class ImageLabel(QLabel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self.setAlignment(Qt.AlignCenter)
        self.setText("ចុចទ្វេដងដើម្បីជ្រើសរើសរូបភាព")
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #cccccc;
                border-radius: 5px;
                background-color: #f9f9f9;
                min-height: 180px;
                max-height: 180px;
                min-width: 180px;
                max-width: 180px;
                padding: 5px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                font-size: 12px;
            }
            QLabel:hover {
                border-color: #4CAF50;
                background-color: #e8f5e8;
            }
        """)
        
    def mouseDoubleClickEvent(self, event):
        if self.parent_dialog:
            self.parent_dialog.select_image()

class ProductImageWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self.image_path = None
        self.initUI()
        
    def initUI(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)
        
        title = QLabel("រូបភាពទំនិញ")
        title.setStyleSheet("font-weight: bold; font-size: 14px; color: #333;")
        layout.addWidget(title)
        
        self.image_label = ImageLabel(self)
        layout.addWidget(self.image_label, 0, Qt.AlignCenter)
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(5)
        
        self.select_btn = QPushButton("ជ្រើសរើសរូបភាព")
        self.select_btn.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        self.select_btn.setStyleSheet("background-color: #2196F3; color: white; border: none; padding: 6px 10px; border-radius: 3px; font-size: 12px; min-height: 30px;")
        self.select_btn.clicked.connect(self.select_image)
        
        self.clear_btn = QPushButton("សម្អាត")
        self.clear_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogResetButton))
        self.clear_btn.setStyleSheet("background-color: #FF9800; color: white; border: none; padding: 6px 10px; border-radius: 3px; font-size: 12px; min-height: 30px;")
        self.clear_btn.clicked.connect(self.clear_image)
        self.clear_btn.setEnabled(False)
        
        button_layout.addWidget(self.select_btn)
        button_layout.addWidget(self.clear_btn)
        
        layout.addLayout(button_layout)
        
        self.info_label = QLabel("មិនទាន់មានរូបភាព")
        self.info_label.setStyleSheet("color: #666666; font-size: 11px; font-style: italic;")
        self.info_label.setWordWrap(True)
        layout.addWidget(self.info_label)
        
        layout.addStretch()
        self.setLayout(layout)
    
    def select_image(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self, 
            "ជ្រើសរើសរូបភាពទំនិញ", 
            "", 
            "Images (*.png *.jpg *.jpeg *.bmp *.gif *.webp)"
        )
        
        if file_path:
            self.load_image(file_path)
    
    def load_image(self, file_path):
        try:
            pixmap = QPixmap(file_path)
            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(
                    170, 170, 
                    Qt.KeepAspectRatio, 
                    Qt.SmoothTransformation
                )
                self.image_label.setPixmap(scaled_pixmap)
                self.image_label.setText("")
                self.image_path = file_path
                self.info_label.setText(f"រូបភាព: {os.path.basename(file_path)}")
                self.clear_btn.setEnabled(True)
                
                if self.parent_dialog and hasattr(self.parent_dialog, 'show_status'):
                    self.parent_dialog.show_status("បានផ្ទុករូបភាពរួចរាល់", "success")
            else:
                QMessageBox.warning(self, "កំហុស", "មិនអាចផ្ទុករូបភាពបានទេ")
        except Exception as e:
            QMessageBox.critical(self, "កំហុស", f"មានបញ្ហាក្នុងការផ្ទុករូបភាព:\n{str(e)}")
    
    def clear_image(self):
        self.image_label.clear()
        self.image_label.setText("ចុចទ្វេដងដើម្បីជ្រើសរើសរូបភាព")
        self.image_path = None
        self.info_label.setText("មិនទាន់មានរូបភាព")
        self.clear_btn.setEnabled(False)
        
        if self.parent_dialog and hasattr(self.parent_dialog, 'show_status'):
            self.parent_dialog.show_status("បានសម្អាតរូបភាព", "warning")
    
    def get_image_path(self):
        return self.image_path
    
    def set_image_from_path(self, image_path):
        if image_path and os.path.exists(image_path):
            self.load_image(image_path)

# ==================== ថ្នាក់ AddProductDialog ====================
class AddProductDialog(QDialog):
    def __init__(self, parent=None, product_data=None):
        super().__init__(parent)
        self.parent_window = parent
        self.currency_manager = CurrencyManager()
        self.product_data = product_data
        self.setWindowTitle('បន្ថែមទំនិញថ្មី' if not product_data else 'កែប្រែទំនិញ')
        self.setModal(True)
        self.setMinimumWidth(1100)
        self.setMinimumHeight(800)
        
        font = QFont('Khmer OS', 11)
        self.setFont(font)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox {
                padding: 10px;
                border: 2px solid #e0e0e0;
                border-radius: 5px;
                font-size: 13px;
                font-family: 'Khmer OS', 'Arial', sans-serif;
                min-height: 20px;
                background-color: white;
            }
            QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                border-color: #4CAF50;
            }
            QLineEdit::placeholder, QTextEdit::placeholder {
                color: #999;
                font-style: italic;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px;
            }
            QLabel {
                font-size: 13px;
                padding: 2px 0;
            }
            QPushButton {
                padding: 12px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
                min-height: 40px;
            }
            QPushButton:hover {
                opacity: 0.9;
            }
            QCheckBox {
                font-size: 13px;
                padding: 5px;
            }
        """)
        
        self.initUI()
        
        if product_data:
            self.load_product_data()
    
    def initUI(self):
        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        left_panel = QWidget()
        left_panel.setMaximumWidth(250)
        left_panel.setStyleSheet("background-color: #f8f9fa; border-radius: 5px; padding: 10px;")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(10)
        
        self.image_widget = ProductImageWidget(self)
        left_layout.addWidget(self.image_widget)
        
        right_panel = QWidget()
        right_panel.setStyleSheet("background-color: white; border-radius: 5px; padding: 10px;")
        right_layout = QVBoxLayout(right_panel)
        
        title = QLabel("ព័ត៌មានទំនិញ")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4CAF50; padding: 5px;")
        right_layout.addWidget(title)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        scroll.setStyleSheet("QScrollArea { background-color: transparent; }")
        
        scroll_widget = QWidget()
        form_layout = QFormLayout(scroll_widget)
        form_layout.setSpacing(15)
        form_layout.setLabelAlignment(Qt.AlignRight)
        form_layout.setFormAlignment(Qt.AlignLeft)
        
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("បញ្ចូលឈ្មោះទំនិញ...")
        form_layout.addRow('ឈ្មោះទំនិញ:', self.name_input)
        
        self.category_input = QComboBox()
        self.category_input.setEditable(True)
        self.category_input.addItems(['សៀវភៅ', 'សៀវភៅកត់ត្រា', 'ប៊ិក', 'ខ្មៅដៃ', 'បន្ទាត់', 'កាបូប', 'ម៉ាស៊ីនគិតលេខ', 'ក្រដាស', 'សម្ភារៈផ្សេងៗ'])
        self.category_input.setPlaceholderText("ជ្រើសរើសឬបញ្ចូលប្រភេទថ្មី...")
        form_layout.addRow('ប្រភេទ:', self.category_input)
        
        code_layout = QHBoxLayout()
        code_layout.setSpacing(10)
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("បញ្ចូលកូដទំនិញ ៦ខ្ទង់ ឬស្កេន Barcode...")
        self.code_input.setMaxLength(6)
        self.code_input.setValidator(QRegExpValidator(QRegExp("[0-9]{0,6}")))
        
        self.generate_code_btn = QPushButton("បង្កើតកូដ")
        self.generate_code_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogContentsView))
        self.generate_code_btn.setStyleSheet("background-color: #4CAF50; color: white; border: none; padding: 8px 12px; border-radius: 4px; font-weight: bold; font-size: 12px; min-width: 80px;")
        self.generate_code_btn.clicked.connect(self.generate_barcode)
        
        code_layout.addWidget(self.code_input)
        code_layout.addWidget(self.generate_code_btn)
        form_layout.addRow('កូដទំនិញ (៦ខ្ទង់):', code_layout)
        
        price_frame = QGroupBox("តម្លៃទំនិញ (គិតជារៀល)")
        price_frame.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #4CAF50;
            }
        """)
        price_layout = QGridLayout()
        price_layout.setVerticalSpacing(10)
        price_layout.setHorizontalSpacing(15)
        
        price_layout.addWidget(QLabel("តម្លៃដើម (៛):"), 0, 0)
        self.cost_price_input = QSpinBox()
        self.cost_price_input.setRange(0, 100000000)
        self.cost_price_input.setSingleStep(100)
        self.cost_price_input.setValue(0)
        self.cost_price_input.setSuffix(" ៛")
        price_layout.addWidget(self.cost_price_input, 0, 1)
        
        price_layout.addWidget(QLabel("តម្លៃលក់រាយ (៛):"), 1, 0)
        self.retail_price_input = QSpinBox()
        self.retail_price_input.setRange(0, 100000000)
        self.retail_price_input.setSingleStep(100)
        self.retail_price_input.setValue(0)
        self.retail_price_input.setSuffix(" ៛")
        price_layout.addWidget(self.retail_price_input, 1, 1)
        
        price_layout.addWidget(QLabel("តម្លៃលក់ដុំ (៛):"), 2, 0)
        self.wholesale_price_input = QSpinBox()
        self.wholesale_price_input.setRange(0, 100000000)
        self.wholesale_price_input.setSingleStep(100)
        self.wholesale_price_input.setValue(0)
        self.wholesale_price_input.setSuffix(" ៛")
        price_layout.addWidget(self.wholesale_price_input, 2, 1)
        
        price_layout.addWidget(QLabel("បរិមាណតិចបំផុតសម្រាប់លក់ដុំ:"), 3, 0)
        self.wholesale_min_qty = QSpinBox()
        self.wholesale_min_qty.setRange(1, 1000)
        self.wholesale_min_qty.setValue(10)
        self.wholesale_min_qty.setSuffix(" ដុំ")
        price_layout.addWidget(self.wholesale_min_qty, 3, 1)
        
        price_frame.setLayout(price_layout)
        form_layout.addRow(price_frame)
        
        discount_frame = QGroupBox("បញ្ចុះតម្លៃ")
        discount_frame.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #FF9800;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #FF9800;
            }
        """)
        discount_layout = QGridLayout()
        discount_layout.setVerticalSpacing(10)
        discount_layout.setHorizontalSpacing(15)
        
        self.has_discount_check = QCheckBox("មានការបញ្ចុះតម្លៃ")
        self.has_discount_check.toggled.connect(self.toggle_discount_fields)
        discount_layout.addWidget(self.has_discount_check, 0, 0, 1, 2)
        
        discount_layout.addWidget(QLabel("ប្រភេទបញ្ចុះតម្លៃ:"), 1, 0)
        self.discount_type = QComboBox()
        self.discount_type.addItems(["ភាគរយ (%)", "ចំនួនទឹកប្រាក់ (៛)"])
        self.discount_type.currentTextChanged.connect(self.calculate_discount)
        discount_layout.addWidget(self.discount_type, 1, 1)
        
        discount_layout.addWidget(QLabel("តម្លៃបញ្ចុះ:"), 2, 0)
        self.discount_value = QDoubleSpinBox()
        self.discount_value.setRange(0, 100000000)
        self.discount_value.setSingleStep(100)
        self.discount_value.setValue(0)
        self.discount_value.valueChanged.connect(self.calculate_discount)
        discount_layout.addWidget(self.discount_value, 2, 1)
        
        discount_layout.addWidget(QLabel("តម្លៃបន្ទាប់ពីបញ្ចុះ:"), 3, 0)
        self.final_price_display = QLabel("0៛")
        self.final_price_display.setStyleSheet("font-weight: bold; color: #4CAF50;")
        discount_layout.addWidget(self.final_price_display, 3, 1)
        
        discount_frame.setLayout(discount_layout)
        form_layout.addRow(discount_frame)
        
        stock_frame = QGroupBox("ស្តុកទំនិញ")
        stock_frame.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #9C27B0;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #9C27B0;
            }
        """)
        stock_layout = QGridLayout()
        stock_layout.setVerticalSpacing(10)
        stock_layout.setHorizontalSpacing(15)
        
        stock_layout.addWidget(QLabel("ស្តុកលក់រាយ:"), 0, 0)
        self.stock_retail_input = QSpinBox()
        self.stock_retail_input.setRange(0, 999999)
        self.stock_retail_input.setValue(0)
        self.stock_retail_input.setSuffix(" ដុំ")
        stock_layout.addWidget(self.stock_retail_input, 0, 1)
        
        stock_layout.addWidget(QLabel("ស្តុកលក់ដុំ:"), 1, 0)
        self.stock_wholesale_input = QSpinBox()
        self.stock_wholesale_input.setRange(0, 999999)
        self.stock_wholesale_input.setValue(0)
        self.stock_wholesale_input.setSuffix(" ដុំ")
        stock_layout.addWidget(self.stock_wholesale_input, 1, 1)
        
        stock_layout.addWidget(QLabel("ស្តុកអប្បបរមា:"), 2, 0)
        self.min_stock = QSpinBox()
        self.min_stock.setRange(0, 999999)
        self.min_stock.setValue(5)
        self.min_stock.setSuffix(" ដុំ")
        stock_layout.addWidget(self.min_stock, 2, 1)
        
        stock_frame.setLayout(stock_layout)
        form_layout.addRow(stock_frame)
        
        self.description_input = QTextEdit()
        self.description_input.setMaximumHeight(80)
        self.description_input.setPlaceholderText("បញ្ចូលការពិពណ៌នាបន្ថែម...")
        form_layout.addRow('បរិយាយ:', self.description_input)
        
        options_group = QGroupBox("ជម្រើសបន្ថែម")
        options_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #2196F3;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #2196F3;
            }
        """)
        options_layout = QHBoxLayout()
        options_layout.setSpacing(20)
        
        self.tax_check = QCheckBox("គិតពន្ធ")
        self.tax_check.setChecked(True)
        options_layout.addWidget(self.tax_check)
        
        self.discountable_check = QCheckBox("អាចបញ្ចុះតម្លៃបាន")
        self.discountable_check.setChecked(True)
        options_layout.addWidget(self.discountable_check)
        
        self.active_check = QCheckBox("សកម្ម")
        self.active_check.setChecked(True)
        options_layout.addWidget(self.active_check)
        
        options_group.setLayout(options_layout)
        form_layout.addRow(options_group)
        
        scroll.setWidget(scroll_widget)
        right_layout.addWidget(scroll)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("padding: 8px; border-radius: 4px; margin-top: 5px; font-size: 12px;")
        self.status_label.hide()
        right_layout.addWidget(self.status_label)
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.save_btn = QPushButton("💾 រក្សាទុក")
        self.save_btn.setStyleSheet("background-color: #4CAF50; color: white; border: none; padding: 12px 25px; border-radius: 5px; font-weight: bold; font-size: 14px;")
        self.save_btn.clicked.connect(self.save_product)
        
        self.print_barcode_btn = QPushButton("🏷️ បោះពុម្ព Barcode")
        self.print_barcode_btn.setStyleSheet("background-color: #FF9800; color: white; border: none; padding: 12px 25px; border-radius: 5px; font-weight: bold; font-size: 14px;")
        self.print_barcode_btn.clicked.connect(self.print_barcode)
        self.print_barcode_btn.setEnabled(False)
        
        self.clear_btn = QPushButton("🔄 សម្អាត")
        self.clear_btn.setStyleSheet("background-color: #9C27B0; color: white; border: none; padding: 12px 25px; border-radius: 5px; font-weight: bold; font-size: 14px;")
        self.clear_btn.clicked.connect(self.clear_form)
        
        cancel_btn = QPushButton("❌ បោះបង់")
        cancel_btn.setStyleSheet("background-color: #f44336; color: white; border: none; padding: 12px 25px; border-radius: 5px; font-weight: bold; font-size: 14px;")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.print_barcode_btn)
        button_layout.addWidget(self.clear_btn)
        button_layout.addWidget(cancel_btn)
        button_layout.addStretch()
        
        right_layout.addLayout(button_layout)
        
        main_layout.addWidget(left_panel)
        main_layout.addWidget(right_panel, 2)
        
        self.toggle_discount_fields(False)
    
    def toggle_discount_fields(self, enabled):
        self.discount_type.setEnabled(enabled)
        self.discount_value.setEnabled(enabled)
        if not enabled:
            self.discount_value.setValue(0)
            self.update_final_price()
    
    def calculate_discount(self):
        if self.has_discount_check.isChecked() and self.discount_value.value() > 0:
            retail_price = self.retail_price_input.value()
            discount_val = self.discount_value.value()
            
            if self.discount_type.currentText() == "ភាគរយ (%)":
                discount_amount = retail_price * (discount_val / 100)
                final_price = retail_price - discount_amount
            else:
                discount_amount = discount_val
                final_price = retail_price - discount_val
            
            if final_price < 0:
                final_price = 0
            
            self.final_price_display.setText(f"{final_price:,.0f}៛")
        else:
            self.final_price_display.setText(f"{self.retail_price_input.value():,.0f}៛")
    
    def update_final_price(self):
        self.calculate_discount()
    
    def show_status(self, message, status_type="info"):
        colors = {
            "success": "#4CAF50",
            "warning": "#FF9800",
            "error": "#f44336",
            "info": "#2196F3"
        }
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"padding: 8px; border-radius: 4px; margin-top: 5px; background-color: {colors.get(status_type, '#2196F3')}; color: white; font-weight: bold; font-size: 12px;")
        self.status_label.show()
        
        QTimer.singleShot(3000, lambda: self.status_label.hide())
    
    def generate_barcode(self):
        barcode = BarcodeManager.generate_barcode(6)
        self.code_input.setText(barcode)
        self.show_status(f"បានបង្កើតកូដ: {barcode}", "success")
    
    def save_product_image(self):
        image_path = self.image_widget.get_image_path()
        if image_path and os.path.exists(image_path):
            file_extension = os.path.splitext(image_path)[1].lower()
            if not file_extension:
                file_extension = '.jpg'
            new_filename = f"{uuid.uuid4().hex}{file_extension}"
            new_path = os.path.join('product_images', new_filename)
            
            try:
                shutil.copy2(image_path, new_path)
                return new_path
            except Exception as e:
                QMessageBox.warning(self, "ព្រមាន", f"មិនអាចរក្សាទុករូបភាព:\n{str(e)}")
                return None
        return None
    
    def validate_inputs(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, 'ព្រមាន', 'សូមបញ្ចូលឈ្មោះទំនិញ!')
            self.name_input.setFocus()
            return False
        
        barcode = self.code_input.text().strip()
        if barcode and len(barcode) != 6:
            QMessageBox.warning(self, 'ព្រមាន', 'កូដទំនិញត្រូវមាន ៦ខ្ទង់!')
            self.code_input.setFocus()
            return False
        
        if self.cost_price_input.value() < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'តម្លៃដើមមិនអាចតិចជាង 0!')
            self.cost_price_input.setFocus()
            return False
        
        if self.retail_price_input.value() < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'តម្លៃលក់រាយមិនអាចតិចជាង 0!')
            self.retail_price_input.setFocus()
            return False
        
        if self.wholesale_price_input.value() < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'តម្លៃលក់ដុំមិនអាចតិចជាង 0!')
            self.wholesale_price_input.setFocus()
            return False
        
        if self.stock_retail_input.value() < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'ចំនួនស្តុកលក់រាយមិនអាចតិចជាង 0!')
            self.stock_retail_input.setFocus()
            return False
        
        if self.stock_wholesale_input.value() < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'ចំនួនស្តុកលក់ដុំមិនអាចតិចជាង 0!')
            self.stock_wholesale_input.setFocus()
            return False
        
        if self.min_stock.value() < 0:
            QMessageBox.warning(self, 'ព្រមាន', 'ស្តុកអប្បបរមាមិនអាចតិចជាង 0!')
            self.min_stock.setFocus()
            return False
        
        return True
    
    def save_product(self):
        if not self.validate_inputs():
            return
        
        try:
            barcode = self.code_input.text().strip() if self.code_input.text().strip() else None
            
            if not self.product_data and barcode:
                self.parent_window.db.cursor.execute(
                    "SELECT id FROM products WHERE barcode = ?", 
                    (barcode,)
                )
                existing = self.parent_window.db.cursor.fetchone()
                if existing:
                    reply = QMessageBox.question(
                        self, 'ព្រមាន',
                        f'មានទំនិញដែលមានកូដ {barcode} រួចហើយ។\nតើចង់បន្តទេ?',
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply == QMessageBox.No:
                        return
            
            image_path = self.save_product_image()
            
            has_discount = 1 if self.has_discount_check.isChecked() else 0
            discount_percent = 0
            discount_amount = 0
            
            if has_discount and self.discount_value.value() > 0:
                if self.discount_type.currentText() == "ភាគរយ (%)":
                    discount_percent = self.discount_value.value()
                    discount_amount = self.retail_price_input.value() * (discount_percent / 100)
                else:
                    discount_amount = self.discount_value.value()
                    discount_percent = (discount_amount / self.retail_price_input.value() * 100) if self.retail_price_input.value() > 0 else 0
            
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            if self.product_data:
                old_image_path = None
                if len(self.product_data) > 15:
                    old_image_path = self.product_data[15]
                
                if image_path and old_image_path and os.path.exists(old_image_path):
                    try:
                        os.remove(old_image_path)
                    except:
                        pass
                
                self.parent_window.db.cursor.execute("""
                    UPDATE products 
                    SET name=?, category=?, barcode=?, cost_price=?, 
                        selling_price_retail=?, selling_price_wholesale=?,
                        wholesale_min_qty=?, discount_percent=?, discount_amount=?,
                        has_discount=?, stock_retail=?, stock_wholesale=?,
                        min_stock=?, description=?, image_path=?, 
                        taxable=?, discountable=?, active=?, updated_at=?
                    WHERE id=?
                """, (
                    self.name_input.text().strip(),
                    self.category_input.currentText(),
                    barcode,
                    self.cost_price_input.value(),
                    self.retail_price_input.value(),
                    self.wholesale_price_input.value(),
                    self.wholesale_min_qty.value(),
                    discount_percent,
                    discount_amount,
                    has_discount,
                    self.stock_retail_input.value(),
                    self.stock_wholesale_input.value(),
                    self.min_stock.value(),
                    self.description_input.toPlainText().strip(),
                    image_path if image_path else old_image_path,
                    1 if self.tax_check.isChecked() else 0,
                    1 if self.discountable_check.isChecked() else 0,
                    1 if self.active_check.isChecked() else 0,
                    current_time,
                    self.product_data[0]
                ))
                message = 'កែប្រែទំនិញរួចរាល់!'
                self.saved_product_id = self.product_data[0]
                
            else:
                self.parent_window.db.cursor.execute("""
                    INSERT INTO products (
                        name, category, barcode, cost_price, 
                        selling_price_retail, selling_price_wholesale,
                        wholesale_min_qty, discount_percent, discount_amount, has_discount,
                        stock_retail, stock_wholesale, min_stock,
                        description, image_path, 
                        taxable, discountable, active, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    self.name_input.text().strip(),
                    self.category_input.currentText(),
                    barcode,
                    self.cost_price_input.value(),
                    self.retail_price_input.value(),
                    self.wholesale_price_input.value(),
                    self.wholesale_min_qty.value(),
                    discount_percent,
                    discount_amount,
                    has_discount,
                    self.stock_retail_input.value(),
                    self.stock_wholesale_input.value(),
                    self.min_stock.value(),
                    self.description_input.toPlainText().strip(),
                    image_path,
                    1 if self.tax_check.isChecked() else 0,
                    1 if self.discountable_check.isChecked() else 0,
                    1 if self.active_check.isChecked() else 0,
                    current_time,
                    current_time
                ))
                self.saved_product_id = self.parent_window.db.cursor.lastrowid
                message = 'បន្ថែមទំនិញថ្មីរួចរាល់!'
            
            self.parent_window.db.conn.commit()
            
            QMessageBox.information(self, 'ជោគជ័យ', message)
            
            self.parent_window.load_products()
            if self.parent_window.current_user.is_admin():
                self.parent_window.load_inventory()
            
            self.print_barcode_btn.setEnabled(True)
            
            reply = QMessageBox.question(self, 'បោះពុម្ព Barcode', 
                                        'តើចង់បោះពុម្ព Barcode ឥឡូវនេះទេ?',
                                        QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.print_barcode()
            else:
                self.accept()
            
        except sqlite3.IntegrityError as e:
            if "UNIQUE constraint failed" in str(e):
                QMessageBox.critical(self, 'កំហុស', 'កូដទំនិញនេះមានរួចហើយ!\nសូមបញ្ចូលកូដផ្សេងទៀត។')
            else:
                QMessageBox.critical(self, 'កំហុស', f'មានបញ្ហាក្នុងការរក្សាទុកទិន្នន័យ:\n{str(e)}')
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មានបញ្ហាក្នុងការរក្សាទុកទិន្នន័យ:\n{str(e)}')
            import traceback
            traceback.print_exc()
    
    def print_barcode(self):
        if hasattr(self, 'saved_product_id'):
            product_data = (
                self.saved_product_id,
                self.name_input.text().strip(),
                self.code_input.text().strip(),
                self.retail_price_input.value(),
                self.image_widget.get_image_path()
            )
            
            dialog = BarcodePrintDialog(self.parent_window, product_data, 1)
            dialog.exec_()
    
    def load_product_data(self):
        if not self.product_data:
            return
        
        try:
            self.saved_product_id = self.product_data[0]
            
            if len(self.product_data) > 1 and self.product_data[1]:
                self.name_input.setText(str(self.product_data[1]))
            
            if len(self.product_data) > 2 and self.product_data[2]:
                self.category_input.setCurrentText(str(self.product_data[2]))
            
            if len(self.product_data) > 3 and self.product_data[3]:
                barcode = str(self.product_data[3])
                if len(barcode) > 6:
                    barcode = barcode[-6:]
                self.code_input.setText(barcode)
            
            if len(self.product_data) > 4 and self.product_data[4] is not None:
                try:
                    self.cost_price_input.setValue(int(float(self.product_data[4])))
                except (ValueError, TypeError):
                    self.cost_price_input.setValue(0)
            
            if len(self.product_data) > 5 and self.product_data[5] is not None:
                try:
                    self.retail_price_input.setValue(int(float(self.product_data[5])))
                except (ValueError, TypeError):
                    self.retail_price_input.setValue(0)
            
            if len(self.product_data) > 6 and self.product_data[6] is not None:
                try:
                    self.wholesale_price_input.setValue(int(float(self.product_data[6])))
                except (ValueError, TypeError):
                    self.wholesale_price_input.setValue(0)
            
            if len(self.product_data) > 7 and self.product_data[7] is not None:
                try:
                    self.wholesale_min_qty.setValue(int(self.product_data[7]))
                except (ValueError, TypeError):
                    self.wholesale_min_qty.setValue(10)
            
            if len(self.product_data) > 8 and self.product_data[8] is not None:
                discount_percent = float(self.product_data[8])
            
            if len(self.product_data) > 9 and self.product_data[9] is not None:
                discount_amount = float(self.product_data[9])
            
            has_discount = False
            if len(self.product_data) > 10 and self.product_data[10] is not None:
                has_discount = bool(self.product_data[10])
            
            self.has_discount_check.setChecked(has_discount)
            if has_discount:
                if discount_percent > 0:
                    self.discount_type.setCurrentText("ភាគរយ (%)")
                    self.discount_value.setValue(discount_percent)
                elif discount_amount > 0:
                    self.discount_type.setCurrentText("ចំនួនទឹកប្រាក់ (៛)")
                    self.discount_value.setValue(discount_amount)
            
            if len(self.product_data) > 11 and self.product_data[11] is not None:
                try:
                    self.stock_retail_input.setValue(int(self.product_data[11]))
                except (ValueError, TypeError):
                    self.stock_retail_input.setValue(0)
            
            if len(self.product_data) > 12 and self.product_data[12] is not None:
                try:
                    self.stock_wholesale_input.setValue(int(self.product_data[12]))
                except (ValueError, TypeError):
                    self.stock_wholesale_input.setValue(0)
            
            if len(self.product_data) > 13 and self.product_data[13] is not None:
                try:
                    self.min_stock.setValue(int(self.product_data[13]))
                except (ValueError, TypeError):
                    self.min_stock.setValue(5)
            else:
                self.min_stock.setValue(5)
            
            if len(self.product_data) > 14 and self.product_data[14]:
                self.description_input.setText(str(self.product_data[14]))
            
            if len(self.product_data) > 15 and self.product_data[15]:
                image_path = str(self.product_data[15])
                if os.path.exists(image_path):
                    self.image_widget.set_image_from_path(image_path)
            
            if len(self.product_data) > 16 and self.product_data[16] is not None:
                self.tax_check.setChecked(bool(self.product_data[16]))
            else:
                self.tax_check.setChecked(True)
            
            if len(self.product_data) > 17 and self.product_data[17] is not None:
                self.discountable_check.setChecked(bool(self.product_data[17]))
            else:
                self.discountable_check.setChecked(True)
            
            if len(self.product_data) > 18 and self.product_data[18] is not None:
                self.active_check.setChecked(bool(self.product_data[18]))
            else:
                self.active_check.setChecked(True)
            
            self.print_barcode_btn.setEnabled(True)
            self.update_final_price()
            self.show_status("បានផ្ទុកទិន្នន័យទំនិញសម្រាប់កែប្រែ", "success")
            
        except Exception as e:
            QMessageBox.critical(self, 'កំហុស', f'មិនអាចផ្ទុកទិន្នន័យទំនិញ:\n{str(e)}')
            import traceback
            traceback.print_exc()
    
    def clear_form(self):
        self.name_input.clear()
        self.category_input.setCurrentIndex(0)
        self.code_input.clear()
        self.cost_price_input.setValue(0)
        self.retail_price_input.setValue(0)
        self.wholesale_price_input.setValue(0)
        self.wholesale_min_qty.setValue(10)
        self.has_discount_check.setChecked(False)
        self.discount_type.setCurrentIndex(0)
        self.discount_value.setValue(0)
        self.stock_retail_input.setValue(0)
        self.stock_wholesale_input.setValue(0)
        self.min_stock.setValue(5)
        self.description_input.clear()
        self.image_widget.clear_image()
        self.tax_check.setChecked(True)
        self.discountable_check.setChecked(True)
        self.active_check.setChecked(True)
        self.print_barcode_btn.setEnabled(False)
        self.update_final_price()
        
        if hasattr(self, 'saved_product_id'):
            delattr(self, 'saved_product_id')
        
        self.show_status("បានសម្អាតទម្រង់បែបបទរួចរាល់", "success")

def start_auth_flow():
    db = StudyMaterialsDB()
    
    while True:
        welcome = WelcomeDialog()
        choice = welcome.exec_()
        
        if choice == 1:
            login_dialog = LoginDialog(db)
            result = login_dialog.exec_()
            
            if result == QDialog.Accepted:
                return login_dialog.current_user
            elif result == 2:
                continue
            else:
                return None
                
        elif choice == 2:
            register_dialog = RegisterDialog(db, None)
            result = register_dialog.exec_()
            
            if result == QDialog.Accepted:
                continue
            elif result == 2:
                continue
            else:
                return None
        else:
            return None

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    font = QFont('Khmer OS', 10)
    app.setFont(font)
    
    current_user = start_auth_flow()
    
    if current_user:
        window = MainWindow(current_user)
        window.show()
        sys.exit(app.exec_())
    else:
        sys.exit()